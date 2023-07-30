# import settings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
from selenium.webdriver.support.select import Select
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
# from bs4 import BeautifulSoup
import warnings
import argparse
import os
from random import randint
import sys
import unicodedata as ud
import uuid
import string
from os import walk
import fitz
from datetime import date
from sys import platform
import json

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config


def explicit_wait():
    time.sleep(randint(1, 2))

def clearlist(*args):
    for varlist in args:
        varlist.clear()

def data_generator(ws):
    allData = {}
    wcode = []
    wshipper = []
    wdesc = []
    wsize = []
    wtotal = []
    wmanufact = []
    wmanufact_addr = []
    wmanufact_city = []
    wconsignee = []
    wconsignee_addr = []
    wconsignee_city = []
    wconsignee_postal = []
    wconsignee_state = []
    wconsignee_stact = []
    wsubmitter = []
    wsubmitter_add = []
    wsubmitter_cityetc = []
    wsubmitter_country = []
    wentryid = ws['B{}'.format(2)].value
    for i in range(2, ws.max_row + 1):
        if wentryid != ws['B{}'.format(i)].value:# and ws['B{}'.format(i)].value != None:
            rid = uuid.uuid4().hex
            allData[rid] = {'data':list(zip(wshipper, wcode, wdesc, wsize, wtotal, wmanufact, wmanufact_addr, wmanufact_city, wconsignee, wconsignee_addr, wconsignee_city, wconsignee_postal, wconsignee_stact, wconsignee_state, wsubmitter, wsubmitter_add, wsubmitter_cityetc, wsubmitter_country)),
            'count' : len(wcode)} 
            wentryid = ws['B{}'.format(i)].value
            # print(wshipper, len(wshipper))
            # break
            clearlist(wshipper, wcode, wdesc, wsize, wtotal, wmanufact, wmanufact_addr, wmanufact_city, wconsignee, wconsignee_addr, wconsignee_city, wconsignee_postal, wconsignee_stact, wconsignee_state, wsubmitter, wsubmitter_add, wsubmitter_cityetc, wsubmitter_country)
        if ws['B{}'.format(i)].value == None:
            break
        wshipper.append(str(ws['B{}'.format(i)].value).strip())
        wcode.append(str(ws['F{}'.format(i)].value).strip())
        strdesc= ud.normalize('NFKD', str(ws['G{}'.format(i)].value).strip()).encode('ascii', 'ignore').decode('ascii')
        wdesc.append(strdesc)
        wsize.append(str(ws['H{}'.format(i)].value).strip())
        wtotal.append(str(ws['I{}'.format(i)].value).strip())
        strmanufact = ud.normalize('NFKD', str(ws['K{}'.format(i)].value).strip()).encode('ascii', 'ignore').decode('ascii')
        wmanufact.append(strmanufact)
        strmanufact_addr = ud.normalize('NFKD', str(ws['L{}'.format(i)].value).strip()).encode('ascii', 'ignore').decode('ascii')
        wmanufact_addr.append(strmanufact_addr)
        wmanufact_city.append(str(ws['M{}'.format(i)].value).strip())
        wconsignee.append(str(ws['N{}'.format(i)].value).strip())
        wconsignee_addr.append(str(ws['O{}'.format(i)].value).strip())
        wconsignee_city.append(str(ws['P{}'.format(i)].value).strip())
        wconsignee_postal.append(str(ws['Q{}'.format(i)].value).strip())
        wconsignee_state.append(str(ws['R{}'.format(i)].value).strip())
        wconsignee_stact.append(str(ws['S{}'.format(i)].value).strip())
        wsubmitter.append(str(ws['T{}'.format(i)].value).strip())
        wsubmitter_add.append(str(ws['U{}'.format(i)].value).strip())
        wsubmitter_cityetc.append(str(ws['V{}'.format(i)].value).strip())
        wsubmitter_country.append(str(ws['W{}'.format(i)].value).strip())

    # rid = uuid.uuid4().hex
    # allData[rid] = {'data':list(zip(wshipper, wcode, wdesc, wsize, wtotal, wmanufact, wmanufact_addr, wmanufact_city, wconsignee, wconsignee_addr, wconsignee_city, wconsignee_postal, wconsignee_stact, wconsignee_state)),
    # 'count' : len(wcode)} 
    return allData

def parse(profilename, datatable, datearrival, pdffolder):
    warnings.filterwarnings("ignore", category=UserWarning)
    config = getConfig()
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir={}".format(getProfiles()[profilename]['chrome_user_data']))
    options.add_argument("profile-directory={}".format(getProfiles()[profilename]['chrome_profile']))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    options.add_argument("--disable-notifications")
    # options.add_argument("user-agent=" + ua.random )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    download_dir = pdffolder
    
    profile = {"plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
                "download.default_directory": download_dir, "download.extensions_to_open": "applications/pdf", 
                'profile.default_content_setting_values.automatic_downloads': 1}
    options.add_experimental_option("prefs", profile)

    driver = webdriver.Chrome(service=Service(CM(version="114.0.5735.90").install()), options=options)
    os.system("cls")
    first = True
    for item in datatable:
        if first:
            # FIRST LOGIN
            first = False
            loginurl = "https://www.access.fda.gov/oaa/logonFlow.htm?execution=e1s1"
            # https://www.access.fda.gov/oaa/logonFlow.htm?execution=e1s3
            driver.get(loginurl)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input[id='understand']")))
            driver.find_element(By.CSS_SELECTOR, "input[id='understand']").click()
            explicit_wait()
            driver.find_element(By.CSS_SELECTOR, "a[id='login']").click()
            explicit_wait()
            driver.find_element(By.CSS_SELECTOR, "a[title='Prior Notice System Interface']").click()
            explicit_wait()
            driver.find_element(By.CSS_SELECTOR, "img[alt='Create New Web Entry']").click()
            explicit_wait()
        else:
            driver.find_element(By.CSS_SELECTOR, "img[alt='Create WebEntry Button']").click()
        
        # SET WEB ENTRY
        pncount = str(datatable[item]['count'])
        
        print("PN Web Entry", datatable[item]['data'][0][14], "Started.. ", "(" + pncount + " Products)")
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='webEntry.entryType.code']")).select_by_visible_text('Consumption')
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "img[alt='Next Button']").click()
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='generateIdFlag']").click()
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='webEntry.intendedPNCount']").send_keys(pncount)
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='webEntry.portOfArrival.portCode']").send_keys("3310")
        explicit_wait()
        ldate = datearrival.split("-")
        strdate = "{}/{}/{}".format(ldate[1], ldate[2], ldate[0])
        driver.find_element(By.CSS_SELECTOR, "input[name='anticipatedArrivalDate']").send_keys(strdate)
        explicit_wait()
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='hourValue']")).select_by_visible_text('09')
        time.sleep(1)
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='minValue']")).select_by_visible_text('00')
        explicit_wait()
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='submitterSameAsRoleId']")).select_by_visible_text('Yes')
        explicit_wait()
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='importerSameAsRoleId']")).select_by_visible_text('Yes')
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "img[alt='Enter Submitter Button']").click()
        explicit_wait()

        # added
        # input('')
        driver.find_element(By.CSS_SELECTOR, "input[name='submitterSameAsRoleId']").click()
        explicit_wait()
        wsubmitter = datatable[item]['data'][0][14]
        wsubmitter_add = datatable[item]['data'][0][15]
        wsubmitter_cityetc = datatable[item]['data'][0][16]
        wsubmitter_clist = wsubmitter_cityetc.split("/")
        # input('pause')
        
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.name']").clear() 
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.name']").send_keys(wsubmitter)
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.address.address1']").clear()
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.address.address1']").send_keys(wsubmitter_add)
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.address.city']").clear()
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.address.city']").send_keys(wsubmitter_clist[0])
        explicit_wait()
        Select(driver.find_element(By.CSS_SELECTOR, "select[id='requiring work']")).select_by_value("{}-{}".format("CA", wsubmitter_clist[1]))
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.address.zipMailCode']").clear()
        driver.find_element(By.CSS_SELECTOR, "input[name='submitter.address.zipMailCode']").send_keys(wsubmitter_clist[2])
        explicit_wait()
        #---


        driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='useNewAddr'][value='1']").click()
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "img[alt='OK Button']").click()
        explicit_wait()
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='motCode']")).select_by_visible_text('Land, Truck')
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "img[alt='Enter Carrier Button']").click()
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "input[name='carrier.name']").send_keys("DYKP")
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
        # input("pause")
        explicit_wait()
        driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
        explicit_wait()

        for data in datatable[item]['data']:
            try:
                driver.find_element(By.CSS_SELECTOR, "img[alt='Create PN Button']").click()
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='shippingCountryCode']")).select_by_visible_text('Canada  (CA)')
                explicit_wait()
                wcode = data[1]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.product.fdaProductCode']").send_keys(wcode)
                explicit_wait()
                wdesc = data[2]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.product.productCommonName']").send_keys(wdesc)
                explicit_wait()
                wsize = data[3]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.baseUnitNumber']").send_keys(wsize)
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='uomCode']")).select_by_visible_text('Grams')
                explicit_wait()
                wtot = data[4]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.packageItem0.containerQty']").send_keys(wtot)
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='containerTypeCode0']")).select_by_visible_text('Box')
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='pnArticle.pnFacilities.producer.address.country.countryCode']")).select_by_visible_text('Canada  (CA)')
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Enter Manufacturer Button']").click()
                explicit_wait()
                wmanuc = data[5]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.producer.name']").send_keys(wmanuc)
                explicit_wait()
                wmanuc_addr = data[6]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.producer.address.address1']").send_keys(wmanuc_addr)
                explicit_wait()
                wmanuc_city = data[7]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.producer.address.city']").send_keys(wmanuc_city)
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.producer.regExemptFlag']").click()
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[id='requiring work']")).select_by_value("11")
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='pnArticle.pnFacilities.shipper.address.country.countryCode']")).select_by_visible_text('Canada  (CA)')

                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Enter Shipper Button']").click()
                explicit_wait()

                Select(driver.find_element(By.CSS_SELECTOR, "select[id='State']")).select_by_value("8")
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()

                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='pnArticle.pnFacilities.owner.address.country.countryCode']")).select_by_visible_text('Canada  (CA)')
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Enter Owner Button']").click()
                explicit_wait()
                Select(driver.find_element(By.CSS_SELECTOR, "select[id='State']")).select_by_value("8")
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Enter Consignee Button']").click()
                explicit_wait()
                wcon = data[8]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.consignee.name']").send_keys(wcon)
                explicit_wait()
                wcon_addr = data[9]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.consignee.address.address1']").send_keys(wcon_addr)
                explicit_wait()
                wcon_city = data[10]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.consignee.address.city']").send_keys(wcon_city)
                explicit_wait()
                wcon_postal = data[11]
                driver.find_element(By.CSS_SELECTOR, "input[name='pnArticle.pnFacilities.consignee.address.zipMailCode']").send_keys(wcon_postal)
                explicit_wait()
                wcon_state = data[12]
                Select(driver.find_element(By.CSS_SELECTOR, "select[name='pnArticle.pnFacilities.consignee.address.subdivision.code']")).select_by_visible_text(wcon_state)
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Save Button']").click()
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='PN Save Button']").click()
                explicit_wait()
                print(wcode, "Saved")
                driver.find_element(By.CSS_SELECTOR, "img[alt='Next Button']").click()
                explicit_wait()
                driver.find_element(By.CSS_SELECTOR, "img[alt='Cancel Button']").click()
                explicit_wait()
            except:
                input("Error Found..")
        # break
        # input("pause")
        try:
            driver.find_element(By.CSS_SELECTOR, "img[alt='Complete Web Entry Button']").click()
            explicit_wait()
            driver.find_element(By.CSS_SELECTOR, "img[alt='Next Button']").click()
            explicit_wait()
            driver.find_element(By.CSS_SELECTOR, "img[alt='Print Summary Button']").click()
            time.sleep(2)
            # input("pause")
            print("PN Web Entry", datatable[item]['data'][0][14], "End.\n")
        except:
            input("Error found")
        # input("pause")
    time.sleep(5)
    driver.close()

def format_filename(s):
    valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
    filename = ''.join(c for c in s if c in valid_chars)
    filename = filename.replace(' ','_') # I don't like spaces in filenames.
    return filename

def get_delimeter():
    delimeter = "/"    
    if platform == "win32":
        delimeter = "\\"
    return delimeter


def pdf_rename(pdffolder):
    filelist = []
    print("Renaming Files started")
    filenames = next(walk(pdffolder), (None, None, []))[2]  # [] if no file
    delimeter = get_delimeter()
    for filename in filenames:
        if filename.find("filename") != -1:
            doc = fitz.open(pdffolder + delimeter + filename)
            page = doc[0]
            search = page.get_text("blocks", clip=[100.6500015258789, 271.04034423828125, 185.60845947265625, 283.09893798828125])
            tmpname = search[0][4].replace(".", "")
            strdate = str(date.today())
            pdfsubmitter = format_filename("{}_{}.{}".format(tmpname, strdate, "pdf"))
            doc.close()
            isExist = os.path.exists(pdffolder + delimeter + pdfsubmitter)
            if isExist == False:
                print("rename", pdffolder + delimeter + filename)
                os.rename(pdffolder + delimeter + filename, pdffolder + delimeter + pdfsubmitter)
                filelist.append(pdffolder + delimeter + pdfsubmitter)
            else:
                print("File destination exist, rename failed")
    return filelist

def webentry_update(workbook, worksheet, filelist, xlsfilename):
    print("Update Web Entry Identification Started..")
    time.sleep(1)
    for pdffile in filelist:
        doc = fitz.open(pdffile)
        page = doc[0]
        submitter = page.get_text("block", clip=[100.6500015258789, 271.04034423828125, 185.60845947265625, 283.09893798828125]).strip()
        entry_id = page.get_text("block", clip=(152.7100067138672, 202.04034423828125, 230.7493438720703, 214.09893798828125)).strip()

        # print(submitter, entry_id)
        for i in range(2, worksheet.max_row + 1):
            if worksheet['B{}'.format(i)].value == None:
                break
            if worksheet['T{}'.format(i)].value.strip() == submitter:
                worksheet['A{}'.format(i)].value = entry_id
        workbook.save(xlsfilename)
        print(submitter, "Updated")
        time.sleep(1)
    print("Update Web Entry Identification Finished..")


def main():
    parser = argparse.ArgumentParser(description="FDA Entry")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-s', '--sheet', type=str,help="Sheet Name")
    parser.add_argument('-dt', '--date', type=str,help="Arrival Date")
    parser.add_argument('-d', '--data', type=str,help="Chrome User Data Directory")
    parser.add_argument('-o', '--output', type=str,help="PDF output folder")
    args = parser.parse_args()
    if args.input[-5:] != '.xlsx':
        input('File input have to XLSX file')
        sys.exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        input('Please check XLSX file')
        sys.exit()
    if len(args.date) != 10:
        input('Date Arrival is wrong')
        sys.exit()

    isExist = os.path.exists(args.output)
    if isExist == False :
        input('Please make sure PDF folder is exist')
        sys.exit()

    wb = load_workbook(filename=args.input , read_only=False)#, keep_vba=True, data_only=True)
    # ws = wb.worksheets[0]
    try:
        ws = wb[args.sheet]
    except:
        input('Sheet Name is not found!')
        sys.exit()
    # input(args)
    os.system('cls')
    print('File Selected:', args.input)
    datagen = data_generator(ws)

    parse(args.data, datagen, args.date, args.output)
    time.sleep(2)
    filelist = pdf_rename(args.output)
    
    webentry_update(wb, ws, filelist, args.input)
    input('All Data Processed done.')

if __name__ == '__main__':
    main()
