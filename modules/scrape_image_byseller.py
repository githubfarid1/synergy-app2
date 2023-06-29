import json
from urllib.parse import urlencode, urlparse
import time
import os
# import xlwings as xw
import argparse
from sys import platform
from playwright.sync_api import sync_playwright
import random
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup 
import re
from urllib import request


userAgentStrings = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13.4; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (X11; Linux i686; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (X11; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (X11; Fedora; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/113.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13.4; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (X11; Linux i686; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (Linux x86_64; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (X11; Fedora; Linux x86_64; rv:102.0) Gecko/20100101 Firefox/102.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.57",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.57"
]
resource_types1 = ["stylesheet", "script", "image", "font"] 
resource_types2 = ["stylesheet",  "image", "font"] 
resource_types3 = ["stylesheet", "script", "image", "font"] 

def file_delimeter():
    delimeter = "/"    
    if platform == "win32":
        delimeter = "\\"
    return delimeter

def block_aggressively1(route): 
	if (route.request.resource_type in resource_types1): 
		route.abort() 
	else: 
		route.continue_() 

def block_aggressively2(route): 
	if (route.request.resource_type in resource_types2): 
		route.abort() 
	else: 
		route.continue_() 

def block_aggressively3(route): 
	if (route.request.resource_type in resource_types3): 
		route.abort() 
	else: 
		route.continue_() 

def parse(fileinput, imagedir, postal):
    # os.system('cls')
    notfound = ['Sorry! Something went wrong!', 'Amazon.com']
    print('File Selected:', fileinput)
    scrapebyseller_source = fileinput
    wb = load_workbook(filename=scrapebyseller_source , read_only=False)
    ws = wb['Sheet1']
    for irow in range(1, ws.max_row + 1):
        if ws['A{}'.format(irow)].value == None:
            break
        sheet_name = ws['A{}'.format(irow)].value
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            wb.remove(sheet)
        wb.create_sheet(sheet_name)
        print(sheet_name, 'Created..')
        ws2 = wb[sheet_name]
        asins = []
        baseurl = ws['B{}'.format(irow)].value
        pno = 0
        index = 0
        print("Page Found..", end=" ", flush=True)

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False, timeout=10000)
            context = browser.new_context(user_agent=random.choice(userAgentStrings))
            page = context.new_page()
            # page2 = context.new_page()
            # page.route("**/*", block_aggressively2) 
            page.goto(baseurl)
            page.query_selector("a#nav-global-location-popover-link").click()
            page.wait_for_selector("input.GLUX_Full_Width").fill(postal)
            page.query_selector("span[data-action='GLUXPostalUpdateAction']").click()
            page.wait_for_selector("div.a-popover-footer span[data-action='GLUXConfirmAction']").click()
            time.sleep(2)
            try:
                last = page.wait_for_selector("span[class='s-pagination-item s-pagination-disabled']")
            except:
                print("0")
                continue

            # maxpage = 0
            # if last == None:
            #     pass
            # else:
            maxpage = int(last.text_content())
            print(str(maxpage))
            # input("")

            for ipage in range(1, maxpage+1):
                print(sheet_name, "Page {} scraping...".format(ipage))
                try:
                     browser.close()
                except:
                     pass
                browser = p.chromium.launch(headless=False, timeout=10000)
                context = browser.new_context(user_agent=random.choice(userAgentStrings))
                page = context.new_page()
                page2 = context.new_page()
                # page.route("**/*", block_aggressively2)
                page.goto(baseurl) 
                page.query_selector("a#nav-global-location-popover-link").click()
                page.wait_for_selector("input.GLUX_Full_Width").fill(postal)
                page.query_selector("span[data-action='GLUXPostalUpdateAction']").click()
                page.wait_for_selector("div.a-popover-footer span[data-action='GLUXConfirmAction']").click()
                time.sleep(2)
                url = baseurl + '&page={}'.format(ipage)
                page.goto(url) 
                html = page.content()
                soup = BeautifulSoup(html,"html.parser")
                if soup.find('div', class_='s-main-slot') != None:
                    searchs = soup.find('div', class_='s-main-slot').find_all('div')
                    for search in searchs:
                        if search.has_attr('data-asin') and search['data-asin'] != '':
                            namestr = ''
                            asin = search['data-asin']
                            name = search.find('span',class_='a-size-medium a-color-base a-text-normal')
                            if name:
                                try:
                                    namestr = name.text
                                except:
                                    namestr = ""
                            alink = search.find("a", class_='a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal')
                            pricestr = ""
                            if alink.has_attr('href'):
                                itemurl = "https://www.amazon.com{}".format(alink['href'])
                                # page2.route("**/*", block_aggressively3) 
                                page2.goto(itemurl)
                                html = page2.content()
                                soup = BeautifulSoup(html,"html.parser")
                                try:
                                    pricestr = soup.find("span", class_="priceToPay").find("span", class_="a-offscreen").text.replace("$","")
                                except Exception as er:
                                    pricestr = ""
                            try:
                                m = re.search(r"'colorImages': ({.*})", html)
                                datastr = m.group(0).replace("'colorImages': { 'initial': ", "")[:-1]
                                datadict = json.loads(datastr)
                                imagelist = []
                                for d in datadict:
                                    lastkey = ""
                                    for key, value in d["main"].items():
                                        lastkey = key
                                    imagelist.append(lastkey)
                                icount = 1
                                for img in imagelist:
                                    urlp = urlparse(img)
                                    filename, file_extension = os.path.splitext(urlp.path)
                                    request.urlretrieve(img,  f'{imagedir}{asin}_{str(icount)}{file_extension}')
                                    
                                    icount += 1
                            except:
                                 pass
                                                        
                            print(asin, namestr, pricestr)
                            index += 1
                            ws2['A{}'.format(index)].value = asin
                            ws2['B{}'.format(index)].value = namestr
                            ws2['C{}'.format(index)].value = pricestr
                            # input("pa")
                else:
                     input("pause")
                print(sheet_name, 'page {}'.format(ipage), 'End..')


    wb.save(fileinput)
    input('Finished...')

def main():
    parser = argparse.ArgumentParser(description="Scrape with Images By Amazon Seller")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-d', '--dir', type=str,help="Image directory")
    parser.add_argument('-w', '--website', type=str,help="Website")
    parser.add_argument('-p', '--postal', type=str,help="Deliver to postal")

    args = parser.parse_args()
    if args.input[-5:] != '.xlsx':
        print('File input have to XLSX file')
        exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        print('Please check XLSX file')
        exit()
    
    isExist = os.path.exists(args.dir)
    if isExist == False :
        print('Please check Images Folder')
        exit()
    parse(args.input, args.dir + file_delimeter(), args.postal)
    
if __name__ == '__main__':
    main()

