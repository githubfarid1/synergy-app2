import fitz
import os
from openpyxl import Workbook, load_workbook
import sys
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from sys import platform
from datetime import date, datetime, timedelta
import time
import glob
import shutil
import easyocr
from pathlib import Path
from pdf2image import convert_from_path
import numpy
import pandas as pd
from random import randint
import json
import warnings

warnings.filterwarnings("ignore", category=Warning)
def clearlist(*args):
    for varlist in args:
        varlist.clear()

def explicit_wait():
    time.sleep(randint(1, 3))

def clear_screan():
    # return
    try:
        if platform == "win32":
            os.system("cls")
        else:    
            os.system("clear")
    except Exception as er:
        print(er, "Command is not supported")

def pause(mess=""):
    input(mess)

def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def getDownloadFolder():
    download_folder = os.path.expanduser('~/Downloads')    
    if platform == "win32":
        download_folder = os.getenv('USERPROFILE') + r'\Downloads'
    return download_folder

def killAllChrome():
    if platform == "win32":
        os.system("taskkill /f /im chrome.exe")

def file_delimeter():
    dm = "/" 
    if platform == "win32":
        dm = "\\"
    return dm

def checkdimension(value=""):
    valuelist = value.upper().split('X')
    if len(valuelist) != 3:
        return False
    
    for v in valuelist:
        if v.isnumeric() == False:
            return False
    
    return True


def join_pdfs(source_folder, output_folder, tag='Labels'):
    merger = PdfMerger()
    print("Merging PDF files in one PDF File..", end=" ", flush=True)
    time.sleep(1)
    file_delimeter = "/" 
    if platform == "win32":
        file_delimeter = "\\"
    time.sleep(1)
    now = datetime.now()
    fname = "{} {}.pdf".format(now.strftime("%B %d"), tag)
    # pdfoutput_folder_combined = pdfoutput_folder # + file_delimeter + "combined"
    tmpname = "{}{}{}.pdf".format(source_folder, file_delimeter, "tmp")
    isExist = os.path.exists(tmpname)
    if isExist:
        os.remove(tmpname)
    resultfile = output_folder + file_delimeter + fname
    pdffiles = glob.glob(source_folder + file_delimeter + "*.pdf")
    if len(pdffiles) != 0:
        dictfiles = {}
        for pdffile in pdffiles:
            try:
                basefilename = os.path.basename(pdffile)
                dictfiles[int(basefilename.replace(".pdf",""))] = pdffile
            except:
                continue
        sortedfiles = dict(sorted(dictfiles.items()))
        for file in sortedfiles:
            merger.append(sortedfiles[file])
        merger.write(resultfile)
        print("Finished")
        return resultfile
    else:
        print("No pdf files was found:", source_folder + file_delimeter)
        return ""

def add_page_numbers(pdffile):
    print("Add page numbering...", end=" ", flush=True)
    time.sleep(1)
    tmpfile = "__tmp.pdf" 
    shutil.copy(pdffile, tmpfile)
    doc = fitz.open(tmpfile)
    for i in range(0, doc.page_count):
        page = doc[i]
        page.insert_text((590.2469787597656, 400.38037109375), "{}".format(str(i+1)), rotate=90, fontsize=9)
    doc.save(pdffile)
    doc.close()
    os.remove(tmpfile)
    print("Finished")

def generate_xls_from_pdf(fileinput, addressfile):
    print("Generate new XLS file from PDF File...", end=" ", flush=True)
    # breakpoint()
    addressdict = pd.read_csv(addressfile, usecols=['Consignee', 'Address']).to_dict('index')
    pdfFileObject = open(fileinput, 'rb')
    pdfReader = PdfReader(pdfFileObject)
    ocrreader = easyocr.Reader(['en'], gpu=False, verbose=False)
    # print(" No. Of Pages :", pdfReader.numPages)
    filepath = fileinput[:-4] + ".xlsx"
    wb = Workbook()
    ws = wb.active

    ws['A1'].value = 'Box'
    ws['B1'].value = 'Tracking ID'
    ws['C1'].value = 'Weight'
    ws['D1'].value = 'Consignee'
    ws['E1'].value = 'Address'
    ws['F1'].value = 'Distributor'
    ws['G1'].value = 'Shipment ID'
    ws['H1'].value = 'Consignee Check'

    for i in range(0, len(pdfReader.pages)):
        if os.path.exists(Path("pdftmp.pdf")):
            os.remove(Path("pdftmp.pdf"))
        pdfWriter = PdfWriter()
        pdf = pdfReader.pages[i]
        pdfWriter.add_page(pdf)
        with Path("pdftmp.pdf").open(mode="wb") as output_file:
            pdfWriter.write(output_file)
        images = convert_from_path(Path('pdftmp.pdf'))
        imgcrop = images[0].crop(box = (180,750,750,900))
        # imgcrop.save(Path("pdftmp.png"))
        # breakpoint()
        res = ocrreader.readtext(numpy.array(imgcrop)  , detail = 0)
        # res = ocrreader.readtext(Path("pdftmp.png")  , detail = 0)

        tracking_id = res[1].strip().replace('TRACKING #','').replace(' ','').replace(":","")
        # print(tracking_id, fileinput, i)
        lines = pdf.extract_text(space_width=200).split("\n")
        # print(lines)
        if platform == "win32":
            weight = lines[1].split('-')[1].strip().split("lb")[0]
        else:
            weight = lines[0].split('-')[1].strip().replace("lb","")    
            
        to = lines[7]
        address = lines[8]
        shipper = lines[2]
        code = lines[12]
        box = lines[-2].replace(":", " ")
        consignee = ""
        for idx, addict in addressdict.items():
            if str(addict['Address'])[0:21].lower() == address[0:21].lower():
                consignee = addict['Consignee']
                break

        ws['A{}'.format(i + 2)].value = box
        ws['B{}'.format(i + 2)].value = tracking_id
        ws['C{}'.format(i + 2)].value = weight
        ws['D{}'.format(i + 2)].value = to
        ws['E{}'.format(i + 2)].value = address
        ws['F{}'.format(i + 2)].value = shipper
        ws['G{}'.format(i + 2)].value = code
        ws['H{}'.format(i + 2)].value = consignee
    wb.save(filepath)
    print("Finished")

def copysheet(source, destination, cols, sheetsource, sheetdestination, tracksheet, xlbook):
    print("Insert {} and {} to {}...".format(sheetdestination, tracksheet, destination), end=" ", flush=True)
    # wb1 = load_workbook(filename=destination, read_only=False, keep_vba=True, data_only=True)
    wb2 = load_workbook(filename=source, read_only=False, keep_vba=True, data_only=True)
    try:
        del xlbook.sheets[sheetdestination]
    except:
        pass
    xlbook.sheets.add(sheetdestination)
    ws1 = xlbook.sheets[sheetdestination]
    # wb1.create_sheet(sheetdestination)
    # ws1 = wb1[sheetdestination]
    wstrack = xlbook.sheets[tracksheet]
    ws2 = wb2[sheetsource]
    for i in range(1, ws2.max_row + 1):
        for col in cols:
            ws1['{}{}'.format(col, i)].value = ws2['{}{}'.format(col, i)].value

    # for i in range(2, wstrack.max_row + 1):
    #     for j in range(2, ws2.max_row + 1):
    #         if wstrack['O{}'.format(i)].value == ws2['A{}'.format(j)].value:
    #             wstrack['M{}'.format(i)].value = ws2['B{}'.format(j)].value
    #             wstrack['A{}'.format(i)].value = ws2['H{}'.format(j)].value
    #             break
    
    for j in range(2, ws2.max_row + 1):
        wstrack['A{}'.format(j)].value = ws2['H{}'.format(j)].value
        wstrack['M{}'.format(j)].value = ws2['B{}'.format(j)].value

    # wb1.save(destination)
    # wb1.close()    
    print("Finished")

if __name__ == "__main__":
    pdfoutput = "/home/farid/dev/python/synergy-github/data/sample/Feb 21/combined"
    addressfile = "/home/farid/dev/python/synergy-github/synergy-gui/address.csv"
    resultfile = join_pdfs(pdfoutput_folder=pdfoutput)
    if resultfile != "":
        add_page_numbers(resultfile)
        generate_xls_from_pdf(resultfile, addressfile)
    input("End Process..")    

