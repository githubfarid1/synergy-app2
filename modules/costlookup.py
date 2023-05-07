from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM

# import cred
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup 
from random import randint
import os
import psutil
import warnings
import argparse
import json

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def parse(fileinput, profile):
    config = getConfig()
    warnings.filterwarnings("ignore", category=UserWarning) 
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")
    # options.add_experimental_option('debuggerAddress', 'localhost:9251')
    options.add_argument("user-data-dir={}".format(getProfiles()[profile]['chrome_user_data']))
    options.add_argument("profile-directory={}".format(getProfiles()[profile]['chrome_profile']))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    # options.add_argument("user-agent=" + ua.random )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(service=Service(CM().install()), options=options)
    os.system('cls')
    print('File Selected:', fileinput)
    cl_source = fileinput

    wb = load_workbook(filename=cl_source , read_only=False)
    ws = wb['Sheet1']
    first = True
    for i in range(2, ws.max_row + 1):
        if ws['A{}'.format(i)].value == None:
            break
        if first:
            driver.get('https://www.amazon.com/')
            first = False
            time.sleep(2)
        asin = ws['A{}'.format(i)].value
        
        url = 'https://www.amazon.com/dp/{}'.format(asin)
        driver.get(url)
        # input('p')
        html = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
        soup = BeautifulSoup(html,"html.parser")
        found = True
        try:
            priceall = soup.find('div', {'id':'corePrice_feature_div'}).find('span', class_='a-offscreen').text
            price = priceall.replace('$','')
            title = soup.find('span', {'id':'productTitle'}).text.strip()
            # print(asin, price)
            found = True
        except:
            found = False
            # print(asin, 'Not Found')
            pass
        if found:
            try:
                shipping = soup.find('div', {'id':'deliveryBlockMessage'}).find('span')['data-csa-c-delivery-price'].replace('$','')
            except:
                shipping = '0.00'
                pass
            try:
                shipfrom = soup.find('div', {'class':'tabular-buybox-text', 'tabular-attribute-name':'Ships from'}).text.strip()
            except:
                shipfrom = ''
                pass
            try:
                soldby = soup.find('div', {'class':'tabular-buybox-text', 'tabular-attribute-name':'Sold by'}).text.strip()
            except:
                soldby = ''
                pass
            print(asin, price, shipping, shipfrom, soldby)
            if shipping == 'FREE':
                shipping = '0.00'
            ws['B{}'.format(i)].value = float(price)
            ws['C{}'.format(i)].value = float(shipping)
            ws['E{}'.format(i)].value = shipfrom
            ws['F{}'.format(i)].value = soldby


        else:
            print(asin, 'not found')
            ws['G{}'.format(i)].value = 'Not Found'

        time.sleep(randint(3, 6))

    wb.save(cl_source)
    driver.close()
    input('Finished...')

def main():
    parser = argparse.ArgumentParser(description="Cost Lookup")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-d', '--data', type=str,help="Chrome User Data Directory")
    args = parser.parse_args()
    if args.input[-5:] != '.xlsx':
        print('File input have to XLSX file')
        exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        print('Please check XLSX file')
        exit()
    
    parse(args.input, args.data)
    
if __name__ == '__main__':
    main()

