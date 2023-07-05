from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup 
import pandas as pd
import os
import psutil
import warnings
import argparse
import json
from urllib.parse import urlparse

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def parse(fileinput, profile):
    warnings.filterwarnings("ignore", category=UserWarning) 
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")
    # options.add_experimental_option('debuggerAddress', 'localhost:9251')
    # options.add_argument("user-data-dir={}".format(chrome_data)) #Path to your chrome profile
    options.add_argument("user-data-dir={}".format(getProfiles()[profile]['chrome_user_data']))
    options.add_argument("profile-directory={}".format(getProfiles()[profile]['chrome_profile']))

    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    # options.add_argument("user-agent=" + ua.random )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(service=Service(CM().install()), options=options)
    os.system('cls')
    print('File Selected:', fileinput)
    scrapebyseller_source = fileinput


    wb = load_workbook(filename=scrapebyseller_source , read_only=False)
    ws = wb['Sheet1']
    # Use the active cell when the file was loaded
    # ws = wb.active
    for i in range(1, ws.max_row + 1):
        if ws['A{}'.format(i)].value == None:
            break
        # print(ws['A{}'.format(i)].value)
        sheet_name = ws['A{}'.format(i)].value
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            wb.remove(sheet)
        wb.create_sheet(sheet_name)
        print(sheet_name, 'Created..')
        
        ws2 = wb[sheet_name]

        asins = []
        url = ws['B{}'.format(i)].value
        time.sleep(2)
        driver.get(url)
        page = 0
        index = 0
        first = True
        while True:
            # print(url)
            html = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
            soup = BeautifulSoup(html,"html.parser")
            # input('pause')
            if soup.find('div', class_='s-main-slot') == None:
                try:
                    linkdom = soup.find("div", {"id":"seller-info-storefront-link"}).find("a", class_="a-link-normal")
                    domain = urlparse(url).netloc
                    # input("{}{}".format(domain, linkdom['href']))
                    driver.get("https://{}{}".format(domain, linkdom['href']) )
                    html = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
                    soup = BeautifulSoup(html,"html.parser")
                    input("pause")
                except:
                    break


            if soup.find('div', class_='s-main-slot') != None:
                
                # div#seller-info-storefront-link  a.a-link-normal
                searchs = soup.find('div', class_='s-main-slot').find_all('div')
                for search in searchs:
                    if search.has_attr('data-asin') and search['data-asin'] != '':
                        # asins.append(search['data-asin'])
                        name = ''
                        asin = search['data-asin']
                        search.find('span',class_='a-size-medium a-color-base a-text-normal')
                        if search:
                            name = search.find('span',class_='a-size-medium a-color-base a-text-normal').text
                        index += 1
                        ws2['A{}'.format(index)].value = asin
                        ws2['B{}'.format(index)].value = name
                print(sheet_name, 'page {}'.format(page+1), 'Scrapped..')
                # time.sleep(2)
                # break
                if soup.select('.s-pagination-next.s-pagination-disabled') == []:
                    page += 1
                    url = url + '&page={}'.format(page)
                    time.sleep(2)
                    driver.get(url)
                    
                    # break
                    # urls.append(url)
                else:
                    break
            else:
                break
    wb.save(scrapebyseller_source)
    driver.close()
    input('Finished...')


def main():
    parser = argparse.ArgumentParser(description="Scrape By Amazon Seller")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-d', '--data', type=str,help="Chrome User Data Directory")
    parser.add_argument('-w', '--website', type=str,help="Website")

    args = parser.parse_args()
    if args.input[-5:] != '.xlsx':
        print('File input have to XLSX file')
        exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        print('Please check XLSX file')
        exit()
    
    parse(args.input, args.data)
    
if __name__ == '__main__':
    main()

