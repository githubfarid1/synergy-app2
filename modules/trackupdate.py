# import settings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
# from bs4 import BeautifulSoup
import warnings
import argparse
import os
import json
import requests
from datetime import datetime
import calendar


def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def getcanapost(trackid):
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'en-US,en;q=0.9,ja-JP;q=0.8,ja;q=0.7,id;q=0.6',
        'Authorization': 'Basic Og==',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
        'Referer': 'https://www.canadapost-postescanada.ca/track-reperage/en',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Linux"',
    }

    response = requests.get(
        'https://www.canadapost-postescanada.ca/track-reperage/rs/track/json/package/{}/detail'.format(trackid),
        headers=headers,
    )
    
    data = json.loads(response.text)
    newest = data['events'][0]
    regcd = newest['locationAddr']['regionCd']
    if regcd == "":
        try:
            regcd = newest['locationAddr']['countryNmEn']
        except:
            regcd = ""

    datetime_str = newest['datetime']['date'] + " " + newest['datetime']['time']
    dt = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    if regcd != "":
        # text = f"{calendar.month_abbr[dt.month]} {dt.day} {dt.strftime('%I:%M %p')} {newest['descEn']} {newest['locationAddr']['city'].capitalize() }, {regcd}"
        dayname = dt.strftime("%a")
        text = {
            "time":f"{dayname}, {calendar.month_abbr[dt.month]} {dt.day}, {dt.year} {dt.strftime('%I:%M %p')}",
            "location":f"{newest['descEn']} {newest['locationAddr']['city'].capitalize() }, {regcd}",
            "evdetail": f"{newest['descEn']}"
        }
    else:
        text = {
            "time":f"{dayname}, {calendar.month_abbr[dt.month]} {dt.day}, {dt.year} {dt.strftime('%I:%M %p')}",
            "location":"",
            "evdetail": f"{newest['descEn']}"
        }

    return text

def parse(fileinput, profile):
    warnings.filterwarnings("ignore", category=UserWarning) 
    trackupdate_source = fileinput
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")
    # options.add_experimental_option('debuggerAddress', 'localhost:9251')
    options.add_argument("user-data-dir={}".format(getProfiles()[profile]['chrome_user_data']))
    options.add_argument("profile-directory={}".format(getProfiles()[profile]['chrome_profile']))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    # options.add_argument("user-agent=" + ua.random )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(service=Service(CM().install()), options=options)
    os.system('cls')
    print('File Selected:', fileinput)

    wb = load_workbook(filename=trackupdate_source, read_only=False, keep_vba=True, data_only=True)
    ws = wb['dyk_manifest_template']
    # Use the active cell when the file was loaded
    ws = wb.active
    first = True
    for i in range(2, ws.max_row + 1):
        if ws['R{}'.format(i)].value == None:
            break
        order_id = ws['R{}'.format(i)].value
        url = 'https://sellercentral.amazon.com/orders-v3/order/{}'.format(order_id) # 111-9589748-6199459
        driver.get(url)
        time.sleep(2)
        print('order ID: ', order_id, end=".. ", flush=True)
        try:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[data-test-id='tracking-id-value'], span[data-test-id='tracking-id-value']")))
            print("OK")
        except:
            print("Not Found", end="\n\n")
            continue
        try:
            tracking_id = driver.find_element(By.CSS_SELECTOR, "a[data-test-id='tracking-id-value']").text
        except:
            tracking_id = driver.find_element(By.CSS_SELECTOR, "span[data-test-id='tracking-id-value']").text

        # try:
        try:
            weight = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[8]/div/div/div[1]/div[3]/div/div/div[2]/div[2]/div[3]/div/div[2]').text
        except:
            try:
                weight = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[7]/div/div[1]/div[1]/div[3]/div/div/div[2]/div[2]/div[3]/div/div[2]').text
            except:
                weight = ""

        # SHIPPING COST
        try:
            cost = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[8]/div/div/div[1]/div[3]/div/div/div[2]/div[3]/div/div[2]/div[2]/span').text.replace('$','')
            
        except:
            # 114-5921481-0720211
            try:
                cost = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[7]/div/table/tbody/tr/td[7]/div/table[1]/tbody/div[3]/div[2]/span').text.replace('$','')
            except:
                try:
                    cost = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[7]/div/div/div[1]/div[3]/div/div/div[2]/div[3]/div/div[2]/div[2]/span').text.replace('$','')
                except:
                    cost = ''
        
        # SHIPPING SERVICE
        try:
            service = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[8]/div/div/div[1]/div[2]/div/div[3]/div/div[2]').text
        except:
            try:
                service = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[7]/div/div/div[1]/div[2]/div/div[3]/div/div[2]').text
            except:
                service = ''
        
        try:
            carrier = driver.find_element(By.XPATH,'//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[8]/div/div/div[1]/div[2]/div/div[2]/div[1]/div[2]').text
        except:
            pass

        # print(carrier)
        print(tracking_id,weight, cost, service)
        ws['M{}'.format(i)].value = tracking_id
        ws['N{}'.format(i)].value = weight
        ws['O{}'.format(i)].value = cost
        ws['P{}'.format(i)].value = service

        if carrier == 'Canada Post':
            carrierinfo = getcanapost(tracking_id)
            ws['Z{}'.format(i)].value = carrierinfo['time']
            ws['AA{}'.format(i)].value = carrierinfo['location']
            ws['AB{}'.format(i)].value = carrierinfo['evdetail']
            print(carrierinfo['time'])
        else:
            try:
                try:
                    trackbutton = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[7]/div/div/div[1]/div[2]/div/div[2]/div[2]/div[2]/div/span/a/i'))).click()
                except:
                    trackbutton = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="MYO-app"]/div/div[1]/div[1]/div/div[8]/div/div/div[1]/div[2]/div/div[2]/div[2]/div[2]/div/span/a/i'))).click()
                # input('wait')
                
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="a-popover-content-1"]/div/table/tr[1]')))
                time.sleep(1.5)
                tracktable = driver.find_element(By.XPATH, '//*[@id="a-popover-content-1"]/div/table/tr[2]').find_elements(By.CSS_SELECTOR, 'td')
                timetr = tracktable[0].text
                loctr = tracktable[1].text
                eventtr = tracktable[2].text
                print(timetr)
                ws['Z{}'.format(i)].value = timetr
                ws['AA{}'.format(i)].value = loctr
                ws['AB{}'.format(i)].value = eventtr
            except:
                ws['Z{}'.format(i)].value = ''
                ws['AA{}'.format(i)].value = ''
                ws['AB{}'.format(i)].value = ''

        print("")
    wb.save(trackupdate_source)

    input('Process Finished...')

def main():
    parser = argparse.ArgumentParser(description="Tracking Update")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-d', '--data', type=str,help="Chrome User Data Directory")
    args = parser.parse_args()
    if not (args.input[-5:] == '.xlsx' or args.input[-5:] == '.xlsm'):
        print('File input have to XLSM or XLSX file')
        exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        print('Please check XLSX file')
        exit()
    
    parse(args.input, args.data)
    
if __name__ == '__main__':
    main()

