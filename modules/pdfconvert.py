# from cv2 import add
from pdf2image import convert_from_path
import easyocr
import numpy
from openpyxl import Workbook, load_workbook
import PyPDF2
from pathlib import Path
import os
import io
import warnings
import argparse


def parse(fileinput, fileoutput):
    warnings.filterwarnings("ignore", category=UserWarning) 
    pdfFileObject = open(fileinput, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObject)
    reader = easyocr.Reader(['en'], gpu=False, verbose=False)

    print(" No. Of Pages :", pdfReader.numPages)
    print('')
    filepath = fileoutput
    wb = Workbook()
    ws = wb.active

    for i in range(pdfReader.numPages):
        if os.path.exists(Path("pdftmp.pdf")):
            os.remove(Path("pdftmp.pdf"))
        pdfWriter = PyPDF2.PdfFileWriter()
        pdf = pdfReader.getPage(i)
        pdfWriter.addPage(pdf)
        with Path("pdftmp.pdf").open(mode="wb") as output_file:
            pdfWriter.write(output_file)
        images = convert_from_path(Path('pdftmp.pdf'))
        imgcrop = images[0].crop(box = (180,750,750,900))
        res = reader.readtext(numpy.array(imgcrop)  , detail = 0)
        tracking_id = res[1].strip().replace('TRACKING #','').replace(' ','')
        lines = pdf.extract_text().partition('\n')
        # print(pdf.extract_text())
        weight = lines[0].partition('-')[2].strip()
        to = lines[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[0].strip()
        address = lines[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[0].strip()
        shipper = lines[2].partition('\n')[2].partition('\n')[0].strip()
        code = lines[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[0].strip()
        boxint = lines[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[2].partition('\n')[-1].partition('\n')
        if boxint[0] == '':
            box = 'Box {}'.format(boxint[-1].partition('\n')[-1].strip())
        else:
            box = 'Box {}'.format(boxint[0].strip())


        print(box, tracking_id, weight, to, address, shipper, code)
        ws['A{}'.format(i + 1)].value = box
        ws['B{}'.format(i + 1)].value = tracking_id
        ws['C{}'.format(i + 1)].value = weight
        ws['D{}'.format(i + 1)].value = to
        ws['E{}'.format(i + 1)].value = address
        ws['F{}'.format(i + 1)].value = shipper
        ws['G{}'.format(i + 1)].value = code

    wb.save(filepath)
    input('Finished...')

def main():
    parser = argparse.ArgumentParser(description="Pdf Convert")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-o', '--output', type=str,help="File Output")
    args = parser.parse_args()
    if args.input[-4:] != '.pdf':
        print('File input have to PDF file')
        exit()
    
    if args.output[-5:] != '.xlsx' :
        print('File input have to XLSX file')
        exit()
    parse(args.input, args.output)
    
if __name__ == '__main__':
    main()