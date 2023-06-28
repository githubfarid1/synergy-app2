# https://sellercentral.amazon.com/revcal?ref=RC1&
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM

# import cred
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import requests
import os
import psutil
import warnings
import argparse
import json

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def parse(fileinput, profile, country):
    warnings.filterwarnings("ignore", category=UserWarning) 
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir={}".format(getProfiles()[profile]['chrome_user_data']))
    options.add_argument("profile-directory={}".format(getProfiles()[profile]['chrome_profile']))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    # options.add_argument("user-agent=" + ua.random )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(service=Service(CM().install()), options=options)

    driver.get('https://sellercentral.amazon.com/revcal?ref=RC1&')
    
    cookies = {}
    for cookie in driver.get_cookies():
        cookies[cookie['name']] = cookie['value']
    # print(cookies)
    # exit()
    os.system('cls')
    print('File Selected:', fileinput)
    statistic_source = fileinput
    wb = load_workbook(filename=statistic_source)
    ws = wb['Sheet1']
    # Use the active cell when the file was loaded
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        if ws['A{}'.format(i)].value == None:
            break
        print(ws['A{}'.format(i)].value)

        headers = {
            'authority': 'sellercentral.amazon.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            # Requests sorts cookies= alphabetically
            # 'cookie': 'ubid-main=130-3908922-3270858; sid="XetUu4ai7bifPzLfLHP8Vw==|5fnXbjUwIS6qb21zO6OtGu9Mq2TvJMVAoJzoErtnhE8="; __Host-mselc=H4sIAAAAAAAA/6tWSs5MUbJSSsytyjPUS0xOzi/NK9HLT85M0XM0DA0KDncOcgly8bWMUNJRykVSmZtalJyRCFKq52jkZ+pp6GTi6Gvq5OkGUpeNrLAApCQkLMDF29M7wsDFNUipFgAKAR4EdQAAAA==; session-id=146-7378674-3450147; csm-hit=tb:EETFMRQ23JVJNSS7S2P7+s-N98CQT4MKY68RGJB1NT0|1658705269743&t:1658705269743&adb:adblk_no; session-id-time=2289425277l; session-token=bhWDdtJdp8kJ3acOrlXTTmx0x6BJi4nF4gIhHbFPrSdGzyhIDUzVEB7+Z8370q6+KiLzlyf/3HuXp0HfQeo7u2EXCYBWTWNWdqninHDcxmtHSt8m6PvxzCHONCYZ9OnDWlxkSX0GUFM/o18g6ffzb+wviDax+otltFjbhc3pCvrQJhBoow3hbEAtM/eYGycyYKocacymYTG8oKUgjAMipx4KsM4fmIsro0aIbRZtlouidjEw4IBlntFPRYGTId0V; x-main="tGJbBvopA2ojz59XsYYhNNpwvx8Nzgw26NeX?bU?ELJVaamnWgnqyL9PSxOkkzYi"; at-main=Atza|IwEBIM-peKpCY-cfyymPPp2_tya9lTQumvakYjS1ePQC9kQhYYcXkMUyINVQl0-ATUvYjyTnhJh2xApoxiqYg0nQ_EsZqPupogVCQdd8QQM7SB0YNykyuQiTLn1iVNXbPH3DQXtcoMvTF_bKRSXyjpx6_oLbElGmxt7rGzXlctmS_HbiuuLhN9QMJcZbhhHimfTSBp6MxbU5xHg0APVXqwdR1jCC9pFnPoDa7-Lmp2IOalDKFw; sess-at-main="QCWuILPzA3OInwTQ+G1is0Fos2W3CuuxwXOtRVJ6wvc="; sst-main=Sst1|PQGJqoRjyjqPMDWrAM8edRGnCesTf1Zx5A7orf94ozeCdfZJ2gvTP6mNIDX5a5aplT-T6PstKgHHjUejfrJp3mIgHmLvE0mIIaGQQZsM0qKM6OpmmBylbMdYkAITBN9BWObidGxtGu_Aa2kyQ1uA9p02Fm8SMPlKc1us8OnHlnq6Mc-LbXtA7ydXNVs9hHvI4sVjw1ST8xoks4kyxuoq84StFlvhZAuSiCLFCua9lkkRw_2YTgkDHVgY_c6rVeumklTt2f25S0qcPv6uqT6_RKR6-qcbmEhzqHSIR14c5SwVGC0',
            'referer': 'https://sellercentral.amazon.com/revcal?ref=RC1&',
            'sec-ch-ua': '".Not/A)Brand";v="99", "Google Chrome";v="103", "Chromium";v="103"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
        }

        params = {
            'searchKey': '{}'.format(ws['A{}'.format(i)].value),
            'countryCode': '{}'.format(country),
            'locale': 'en-US',
        }
        session = requests.Session()
        response = session.get('https://sellercentral.amazon.com/revenuecalculator/productmatch', params=params, cookies=cookies, headers=headers)
        data = response.json()
        csrftoken = response.headers['anti-csrftoken-a2z']
        try:
            name = data['data']['otherProducts']['products'][0]['title']
            weight = data['data']['otherProducts']['products'][0]['weight']
            dim1 = data['data']['otherProducts']['products'][0]['length']	
            dim2 = data['data']['otherProducts']['products'][0]['width']	
            dim3 = data['data']['otherProducts']['products'][0]['height']	
            asin = data['data']['otherProducts']['products'][0]['asin']	
            try:
                merchantsku = data['data']['otherProducts']['products'][0]['merchantSku']	
            except:
                merchantsku = ''
            try:
                fnsku = data['data']['otherProducts']['products'][0]['fnsku']	
            except:
                fnsku = ''
            gl = data['data']['otherProducts']['products'][0]['gl']
            try:	
                specialdel = data['data']['otherProducts']['products'][0]['specialDeliveryRequirement']	
            except:
                specialdel = ''
            dimensionunit = data['data']['otherProducts']['products'][0]['dimensionUnit']	
            weightunit = data['data']['otherProducts']['products'][0]['weightUnit']
        except:
                try:
                    name = data['data']['myProducts']['products'][0]['title']
                    weight = data['data']['myProducts']['products'][0]['weight']
                    dim1 = data['data']['myProducts']['products'][0]['length']	
                    dim2 = data['data']['myProducts']['products'][0]['width']	
                    dim3 = data['data']['myProducts']['products'][0]['height']	
                    asin = data['data']['myProducts']['products'][0]['asin']
                    try:
                        merchantsku = data['data']['myProducts']['products'][0]['merchantSku']	
                    except:
                        merchantsku = ''
                    try:
                        fnsku = data['data']['myProducts']['products'][0]['fnsku']	
                    except:
                        fnsku = ''
                    gl = data['data']['myProducts']['products'][0]['gl']
                    try:	
                        specialdel = data['data']['myProducts']['products'][0]['specialDeliveryRequirement']
                    except:
                        specialdel = ''
                    dimensionunit = data['data']['myProducts']['products'][0]['dimensionUnit']	
                    weightunit = data['data']['myProducts']['products'][0]['weightUnit']
                except:
                    print('Product is Not Found')
                    # exit()
                    continue

        headers = {
            'authority': 'sellercentral.amazon.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            # Requests sorts cookies= alphabetically
            # 'cookie': 'ubid-main=130-3908922-3270858; sid="XetUu4ai7bifPzLfLHP8Vw==|5fnXbjUwIS6qb21zO6OtGu9Mq2TvJMVAoJzoErtnhE8="; __Host-mselc=H4sIAAAAAAAA/6tWSs5MUbJSSsytyjPUS0xOzi/NK9HLT85M0XM0DA0KDncOcgly8bWMUNJRykVSmZtalJyRCFKq52jkZ+pp6GTi6Gvq5OkGUpeNrLAApCQkLMDF29M7wsDFNUipFgAKAR4EdQAAAA==; ubid-acbus=131-6094560-3308422; s_pers=%20s_fid%3D257D3B89A54B7044-0C5F4BF3315CC695%7C1816646551778%3B%20s_dl%3D1%7C1658881951780%3B%20gpv_page%3DUS%253ASD%253ASOA-home%7C1658881951786%3B%20s_ev15%3D%255B%255B%2527SCSOAlogin%2527%252C%25271658880151791%2527%255D%255D%7C1816646551791%3B; s_sess=%20s_cc%3Dtrue%3B%20s_ppvl%3DUS%25253ASD%25253ASOA-home%252C11%252C11%252C969%252C1920%252C969%252C1920%252C1080%252C1%252CL%3B%20c_m%3DSCSOAloginundefinedSeller%2520Centralundefined%3B%20s_sq%3D%3B%20s_ppv%3DUS%25253ASD%25253ASOA-home%252C11%252C11%252C969%252C1920%252C969%252C1920%252C1080%252C1%252CL%3B; session-id=144-2412903-6436916; csm-hit=tb:31VKHZ4N39MX1385RQZA+s-B5X64YFXQFMBXD0P7WCJ|1658880155418&t:1658880155418&adb:adblk_no; session-id-time=2289600159l; x-main="rizVGQ5wiTeO?9UXtuENR6ai8ZxZ2Y7Le2VnAeL3gJhHbvwwRDhGpyL1cw07F1BQ"; at-main=Atza|IwEBIB0QaXQj-ea8tdAcFr5NONqPRtymU8m3EVY-eWF6QTtgnJkOJdwWlWHZC9ahknE4g5XT8YZo9pyUCn0u32zc7xFPHPi2VqfWj0PnYPm79fYjeeRGfE0KyEaHtmdLiXm8xZrc-aQcICoIWJFirk0CGOAvOrxvrz3rEMDe6rIk7cZ7lSpjSqZUeoF-fUtoIzFw1KGbHg8_JAY6yxCn3qgW00DBWb9P2SpraizINkk8qCihsQ; sess-at-main="dizRf3EoKB3wvltqrkCTyxAeHxYE9DB+1SiBS0jOU94="; sst-main=Sst1|PQHQMJUWs1WkgR1tHr3Hkl3LCbUE7ejF65PYrh0RTV_zPgc-IpYKVbEafLeXZS1aVjaU3cPm8GYEwfTtkDSo1Q3sxEtxSB8iiccE2_ogv_l6v9f2OZ9J-W5xaVCErOdgS4HGzLjDLxm3VsXxtG1CNOJsQyvIgF8gcMPlXsziXs3MkC3qcXJCISD58FPtRAOv8S8V9XQuQ3qrRyhZ3dKKG-R8TrnBxXAWyyrzyrSlDW8Hhlo19OqyRANRFJQ647QTMsnE82qr6QXYIIt72BVhJtgD5lT5f2G_hnJLwOv3gKjbARU; session-token=6bNEGg4e0SWhWqgEkWf71qxbuaQi9YNvZ1GjaypQ6yf1iF8bAen2yOA3EM08DAjiSWAIxTcHtecnooo78PtZM3AeTnywv/caWFYI7K1xOA6AqEr628PBCUS3HD1pcDoQSC6BbAmC+Gx4P0n0ATjDWwJMTB2q0abVoYiOaIck+Fqe1wzL16iadf1n0s9rvBhGU5y6u8u0xjCTSmLPcRWEnqhZ1JjPWAzi',
            'referer': 'https://sellercentral.amazon.com/revcal?ref=RC1&',
            'sec-ch-ua': '".Not/A)Brand";v="99", "Google Chrome";v="103", "Chromium";v="103"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
        }
        params = {
            'asin': '{}'.format(ws['A{}'.format(i)].value),
            'countryCode': 'US',
            'locale': 'en-US',
        }
        # print(params)
        # time.sleep(5)
        session = requests.Session()
        response = session.get('https://sellercentral.amazon.com/revenuecalculator/getadditionalpronductinfo', params=params, cookies=cookies, headers=headers)
        data = response.json()
        # print(data)
        try:
            afnprice = data['data']['price']['amount']
            mfnprice = data['data']['price']['amount']
        except:
            afnprice = 0
            mfnprice = 0
        try:    
            currency = data['data']['price']['currency']
        except:
            currency= 'USD'
        try:
            mfnshipping = data['data']['shipping']['amount']
        except:
            mfnshipping = 0
        params = {
            'locale': 'en-US',
        }

        json_data = {
            'countryCode': 'US',
            'itemInfo': {
                'asin': asin,
                # 'merchantSku':  merchantsku,
                # 'fnsku': fnsku,
                'glProductGroupName': gl,
                # 'specialDeliveryRequirement': specialdel,
                'packageLength': '{}'.format(dim1),
                'packageWidth': '{}'.format(dim2),
                'packageHeight': '{}'.format(dim3),
                'dimensionUnit': dimensionunit,
                'packageWeight': '{}'.format(weight),
                'weightUnit': weightunit,
                'afnPriceStr': '{}'.format(afnprice),
                'mfnPriceStr': '{}'.format(mfnprice),
                'mfnShippingPriceStr': '{}'.format(mfnshipping),
                'currency': currency,
                'isNewDefined': False,
            },
            'programIdList': [
                'Core',
                'MFN',
            ],
        }
        # print(json_data)

        if merchantsku != '':
            json_data['itemInfo']['merchantSku'] = merchantsku
        if fnsku != '':
            json_data['itemInfo']['fnsku'] = fnsku
        if specialdel != '':
            json_data['itemInfo']['specialDeliveryRequirement'] = specialdel

        headers = {
            'authority': 'sellercentral.amazon.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            'anti-csrftoken-a2z': csrftoken, #'hFJ+oU4193GHU6Aj4DfnNkZn4hQNHCIMUz6w5qJ3AH+KAAAAAGLgiYcwZmEyOWE0Ni1kYzFiLTQ5YjUtOTlhZC1mNDBhOTcxMjI2MDg=',
            'content-type': 'application/json; charset=UTF-8',
            # Requests sorts cookies= alphabetically
            # 'cookie': 'ubid-main=130-3908922-3270858; sid="XetUu4ai7bifPzLfLHP8Vw==|5fnXbjUwIS6qb21zO6OtGu9Mq2TvJMVAoJzoErtnhE8="; __Host-mselc=H4sIAAAAAAAA/6tWSs5MUbJSSsytyjPUS0xOzi/NK9HLT85M0XM0DA0KDncOcgly8bWMUNJRykVSmZtalJyRCFKq52jkZ+pp6GTi6Gvq5OkGUpeNrLAApCQkLMDF29M7wsDFNUipFgAKAR4EdQAAAA==; ubid-acbus=131-6094560-3308422; s_pers=%20s_fid%3D257D3B89A54B7044-0C5F4BF3315CC695%7C1816646551778%3B%20s_dl%3D1%7C1658881951780%3B%20gpv_page%3DUS%253ASD%253ASOA-home%7C1658881951786%3B%20s_ev15%3D%255B%255B%2527SCSOAlogin%2527%252C%25271658880151791%2527%255D%255D%7C1816646551791%3B; s_sess=%20s_cc%3Dtrue%3B%20s_ppvl%3DUS%25253ASD%25253ASOA-home%252C11%252C11%252C969%252C1920%252C969%252C1920%252C1080%252C1%252CL%3B%20c_m%3DSCSOAloginundefinedSeller%2520Centralundefined%3B%20s_sq%3D%3B%20s_ppv%3DUS%25253ASD%25253ASOA-home%252C11%252C11%252C969%252C1920%252C969%252C1920%252C1080%252C1%252CL%3B; session-id=144-2412903-6436916; csm-hit=tb:31VKHZ4N39MX1385RQZA+s-B5X64YFXQFMBXD0P7WCJ|1658880155418&t:1658880155418&adb:adblk_no; session-id-time=2289600159l; x-main="rizVGQ5wiTeO?9UXtuENR6ai8ZxZ2Y7Le2VnAeL3gJhHbvwwRDhGpyL1cw07F1BQ"; at-main=Atza|IwEBIB0QaXQj-ea8tdAcFr5NONqPRtymU8m3EVY-eWF6QTtgnJkOJdwWlWHZC9ahknE4g5XT8YZo9pyUCn0u32zc7xFPHPi2VqfWj0PnYPm79fYjeeRGfE0KyEaHtmdLiXm8xZrc-aQcICoIWJFirk0CGOAvOrxvrz3rEMDe6rIk7cZ7lSpjSqZUeoF-fUtoIzFw1KGbHg8_JAY6yxCn3qgW00DBWb9P2SpraizINkk8qCihsQ; sess-at-main="dizRf3EoKB3wvltqrkCTyxAeHxYE9DB+1SiBS0jOU94="; sst-main=Sst1|PQHQMJUWs1WkgR1tHr3Hkl3LCbUE7ejF65PYrh0RTV_zPgc-IpYKVbEafLeXZS1aVjaU3cPm8GYEwfTtkDSo1Q3sxEtxSB8iiccE2_ogv_l6v9f2OZ9J-W5xaVCErOdgS4HGzLjDLxm3VsXxtG1CNOJsQyvIgF8gcMPlXsziXs3MkC3qcXJCISD58FPtRAOv8S8V9XQuQ3qrRyhZ3dKKG-R8TrnBxXAWyyrzyrSlDW8Hhlo19OqyRANRFJQ647QTMsnE82qr6QXYIIt72BVhJtgD5lT5f2G_hnJLwOv3gKjbARU; session-token=lLknNnsXnyvGhJKBUzEtyp3Gjx8F0+NWo5zLRgroj8DFu1UIuAClu/9GESM2Y6XV9GbiwUrO3tkyXjkBkJ13XgkI1UoXC1BXB7Bj1UEPB/JptXXhfaylhS4rVqeq0bl+caeOMUFSJmZ5oAT1HCN7cO1NnTtLdEEzgBva5+iylTKIAOtWbLIvs9kCFWdWAC9x0gvc9Q5rhTZArUAX75p1aVJaGiw8Ch6B',
            'origin': 'https://sellercentral.amazon.com',
            'referer': 'https://sellercentral.amazon.com/revcal?ref=RC1&',
            'sec-ch-ua': '".Not/A)Brand";v="99", "Google Chrome";v="103", "Chromium";v="103"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
        }

        
        if dim1 == 0 or dim2 == 0 or dim3 == 0:
            feeamount = 0    
        else:
            session = requests.Session()
            response = session.post('https://sellercentral.amazon.com/revenuecalculator/getfees', params=params, cookies=cookies, headers=headers, json=json_data)
            # print(response)
            data = response.json()
            feeamount = data['data']['programFeeResultMap']['Core']['otherFeeInfoMap']['FulfillmentFee']['total']['amount']
        
            
        print(name, weight, dim1, dim2, dim3, feeamount)
        ws['B{}'.format(i)].value = name
        ws['C{}'.format(i)].value = round(weight,3)
        ws['D{}'.format(i)].value = round(dim1,3)
        ws['E{}'.format(i)].value = round(dim2,3)
        ws['F{}'.format(i)].value = round(dim3,3)
        ws['G{}'.format(i)].value = feeamount
        # if i == 10:
        #     break

        time.sleep(2)

        # exit()
    wb.save(statistic_source)
    input('Finished...')


def main():
    parser = argparse.ArgumentParser(description="Statistics")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-d', '--profile', type=str,help="Chrome Profile Selected")
    parser.add_argument('-c', '--country', type=str,help="Country")

    args = parser.parse_args()

    if args.input[-5:] != '.xlsx':
        print('File input have to XLSX file')
        exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        print('Please check XLSX file')
        exit()
    
    parse(args.input, args.profile, args.country)
    
if __name__ == '__main__':
    main()

