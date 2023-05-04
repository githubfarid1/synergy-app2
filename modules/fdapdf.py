import sys
import fitz
import shutil
import os
import argparse
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import unicodedata as ud
from sys import platform
from os import walk
import json

POSX1CODE1 = 12.0
POSX2CODE1 = 32.01737594604492
POSX1DESC = 71.0999984741211
POSX2DESC = 271.55694580078125
POSX1LOC = 277.95001220703125
POSX2LOC = 290.450927734375
POSX1DATE = 396.1499938964844
POSX2DATE = 478.6723937988281
POSX1CODE2 = 514.3499755859375
POSX2CODE2 = 594.415771484375

def clearlist(*args):
    for varlist in args:
        varlist.clear()


def combine_allpdf(pdffiles, pdfoutput):
    print("Merging PDF files started..")
    file_delimeter = "/"    
    if platform == "win32":
        file_delimeter = "\\"

    time.sleep(1)
    dictfiles = {}
    result = fitz.open()
    for pdffile in pdffiles:
        basefilename = os.path.basename(pdffile)
        dictfiles[int(basefilename.replace(".pdf",""))] = pdffile
    sortedfiles = dict(sorted(dictfiles.items()))
    # print(sortedfiles)

    for file in sortedfiles:
        print(os.path.basename(sortedfiles[file]), "merged")
        mfile = fitz.open(sortedfiles[file])
        result.insert_pdf(mfile)
        time.sleep(1)
    result.save(pdfoutput + file_delimeter + "combine.pdf")
    print(pdfoutput + file_delimeter + "combine.pdf", "Created!")
    print("Merging PDF files finished..")

class PriorPdf:
    def __init__(self, filename, page, xlsname, sname, pdfoutput) -> None:
        print(os.path.basename(filename), "Initialated..")
        time.sleep(1)
        self.__filename = filename
        self.__page = page
        self.pnlist = []
        self.__xlsname = xlsname
        self.__sname = sname
        self.__pdfoutput = pdfoutput
        self.__savedfiles = []
        self.__delimeter = "/"    
        if platform == "win32":
            self.__delimeter = "\\"

        try:
            self.__workbook = load_workbook(filename=xlsname, read_only=False)#, keep_vba=True, data_only=True)
            self.__worksheet = self.__workbook[sname]
        except:
            input("XLSX file or Sheet name not found")
            sys.exit()
        self.__data_generator()
        self.__newpdf_generator()
    
    def __newpdf_generator(self):
        foldername = self.filename.replace(".pdf",'')
        isExist = os.path.exists(foldername)
        if not isExist:
            os.makedirs(foldername)
        boxlist = []
        for ds in self.pnlist:
            tmpboxes = ds['boxes'].replace('Box','').split(',')
            for tmpbox in tmpboxes:
                fname = "{}.pdf".format(tmpbox.strip())
                boxlist.append(fname)

        for box in set(boxlist):
            shutil.copy(self.filename, "{}{}{}".format(foldername, self.file_delimeter, box) )
        
    def tester(self):
        
        foldername = os.path.dirname(self.filename)
        fname = "16.pdf"
        doc = fitz.open(foldername + self.file_delimeter + fname)
        pdfpage = doc[0]
        rects = pdfpage.search_for('Mars Fun Size Bars, (12.9g/0.5 oz.) x 50 pack, Peanut Free {Imported from Canada}')
        
        splitter = rects[0][2]
        rdata = []
        tmpdata = []
        # print(splitter)
        first = True
        for rect in rects:
            # print(rect)
            # print(rect[2])
            if first:
                tmpdata.append(rect)
                first = False
            else:    
                if rect[2] == splitter:
                    rdata.append(tmpdata)
                    tmpdata = []
                    tmpdata.append(rect)
                    # pass
                else:
                    tmpdata.append(rect)
        
        rdata.append(tmpdata)
        print(rdata)

    def insert_text(self):
        print("")
        print("Inserting filenames text into PDF started..")
        time.sleep(1)
        pdffolder = self.filename.replace(".pdf",'')
        filenames = next(walk(pdffolder), (None, None, []))[2]  # [] if no file
        if platform == "linux" or platform == "linux2":
            pdffolder = pdffolder + "/"
        elif platform == "win32":
            pdffolder = pdffolder + "\\"
        red = fitz.utils.getColor("red")
        for filename in filenames:
            shutil.copy(pdffolder + self.file_delimeter + filename, pdffolder + self.file_delimeter + "tmp.pdf")
            doc = fitz.open(pdffolder + self.file_delimeter + "tmp.pdf")
            for i in range(0, doc.page_count):
                page = doc[i]
                page.insert_text((520.2469787597656, 803.38037109375), filename, color=red)
            doc.save(pdffolder + self.file_delimeter + filename)
            self.savedfiles.append(pdffolder + self.file_delimeter + filename)
            print(filename, "inserted.")
            time.sleep(1)
            doc.close()    
        os.remove(pdffolder + self.file_delimeter + "tmp.pdf")
        print("Inserting the filenames finished..", end="\n----------------------------------------\n\n")
    
    def __research_text(self, pdfpage, text):
        for i in range(0, len(text)+1):
            rect = pdfpage.search_for(text[0:i],flags=(fitz.TEXT_PRESERVE_WHITESPACE))
            if rect == []:
                break
        lastfound = text[0:i-1]
        tail = text.replace(lastfound, "")
        textsearch = lastfound + " " + tail
        return pdfpage.search_for(textsearch,flags=(fitz.TEXT_PRESERVE_WHITESPACE))
        # pass

    def highlightpdf_generator(self):
        foldername = self.filename.replace(".pdf",'')
        print("")
        print("Highlight text Process Starting...")
        time.sleep(1)
        code2set = set()
        for idx, ds in enumerate(self.pnlist):
            time.sleep(1)
            print(ds['webentry_id'], ds['boxes'])
            tmpboxes = ds['boxes'].replace('Box','').split(',')
            for tmpbox in tmpboxes:
                fname = "{}.pdf".format(tmpbox.strip())
                shutil.copy(foldername + self.file_delimeter + fname, foldername + self.file_delimeter + "tmp.pdf")
                doc = fitz.open(foldername + self.file_delimeter + "tmp.pdf")
                for i in range(0, doc.page_count):
                    pdfpage = doc[i]
                    rects = pdfpage.search_for(ds['description'], flags=(fitz.TEXT_PRESERVE_WHITESPACE))
                    
                    if rects == []:
                        rects = self.__research_text(pdfpage, ds['description'])
                        if rects != []:
                            break
                    else:
                        break
                if rects == []:
                    input("Item not found, Report to administrator")
                    sys.exit()
                
                splitter = rects[0][2]
                #DON'T REMOVE THIS
                # try:
                #     splitter = rects[0][2]
                # except:
                #     words = ds['description'].split()
                #     res = ''
                #     for i in range(1, len(words)):
                #         rects = pdfpage.search_for(' '.join(words[:i]))
                #         if rects == []:
                            
                #             res = ' '.join(words[:i-1])
                #             break
                #     rects = pdfpage.search_for(res)
                #     rects[0][3] = rects[0][3] + 10
                #     splitter = rects[0][2]


                # try:
                #     splitter = rects[0][2]
                #     # print(ds['description'])
                # except:
                #     words = ds['description'].split()
                #     pos = -1
                #     while True:
                #         sent = ' '.join(words[:pos])
                #         try:
                #             # print(sent)
                #             rects = pdfpage.search_for(sent)
                #             print(rects)
                #             splitter = rects[0][2]
                #             break
                #         except:
                #             pos = pos-1
                #             continue
                # print(rects)            
                
                rdata = []
                tmpdata = []
            
                first = True
                for rect in rects:
                    # print(rect[2])
                    if first:
                        tmpdata.append(rect)
                        first = False
                    else:    
                        if rect[2] == splitter:
                            rdata.append(tmpdata)
                            tmpdata = []
                            tmpdata.append(rect)
                            # pass
                        else:
                            tmpdata.append(rect)
                
                rdata.append(tmpdata)

                if len(rdata) > 1:
                    for rd in rdata:
                        pncode1s = pdfpage.get_text("blocks", clip=(POSX1CODE1, rd[0][1]-10, POSX2CODE1, rd[0][3]+10))
                        pncode2s = pdfpage.get_text("blocks", clip=(POSX1CODE2, rd[0][1]-10, POSX2CODE2, rd[0][3]+10))
                        locs = pdfpage.get_text("blocks", clip=(POSX1LOC, rd[0][1]-10, POSX2LOC, rd[0][3]+10))
                        dates = pdfpage.get_text("blocks", clip=(POSX1DATE, rd[0][1]-10, POSX2DATE, rd[0][3]+10))
                        # print(pncode2s[0][4])
                        if pncode2s[0][4].strip() in code2set:
                            continue
                        else:
                            code2set.add(pncode2s[0][4].strip())
                            pnnumber = pncode2s[0][4].strip()
                            r = fitz.Rect(pncode2s[0][0], pncode2s[0][1], pncode2s[0][2], pncode2s[0][3])
                            pdfpage.add_highlight_annot(r)
                            r = fitz.Rect(pncode1s[0][0], pncode1s[0][1], pncode1s[0][2], pncode1s[0][3])
                            pdfpage.add_highlight_annot(r)
                            r = fitz.Rect(locs[0][0], locs[0][1], locs[0][2], locs[0][3])
                            pdfpage.add_highlight_annot(r)
                            r = fitz.Rect(dates[0][0], dates[0][1], dates[0][2], dates[0][3])
                            pdfpage.add_highlight_annot(r)
                            for rx in rd:
                                r = fitz.Rect(rx)
                                pdfpage.add_highlight_annot(r)

                            doc.save(foldername + self.file_delimeter + fname)

                            break
                else:
                    pncode1s = pdfpage.get_text("blocks", clip=(POSX1CODE1, rdata[0][0][1]-10, POSX2CODE1, rdata[0][0][3]+10))
                    pncode2s = pdfpage.get_text("blocks", clip=(POSX1CODE2, rdata[0][0][1]-10, POSX2CODE2, rdata[0][0][3]+10))
                    locs = pdfpage.get_text("blocks", clip=(POSX1LOC, rdata[0][0][1]-10, POSX2LOC, rdata[0][0][3]+10))
                    dates = pdfpage.get_text("blocks", clip=(POSX1DATE, rdata[0][0][1]-10, POSX2DATE, rdata[0][0][3]+10))
                    pnnumber = pncode2s[0][4].strip()
                    r = fitz.Rect(pncode2s[0][0], pncode2s[0][1], pncode2s[0][2], pncode2s[0][3])
                    pdfpage.add_highlight_annot(r)
                    r = fitz.Rect(pncode1s[0][0], pncode1s[0][1], pncode1s[0][2], pncode1s[0][3])
                    pdfpage.add_highlight_annot(r)
                    r = fitz.Rect(locs[0][0], locs[0][1], locs[0][2], locs[0][3])
                    pdfpage.add_highlight_annot(r)
                    r = fitz.Rect(dates[0][0], dates[0][1], dates[0][2], dates[0][3])
                    pdfpage.add_highlight_annot(r)

                    for rx in rdata[0]:
                        r = fitz.Rect(rx)
                        pdfpage.add_highlight_annot(r)

                    doc.save(foldername + self.file_delimeter + fname)
                    #pnlist
                self.pnlist[idx]['pnnumber'] = pnnumber
        try:
            doc.close()
            os.remove(foldername + self.file_delimeter + "tmp.pdf")
        except:
            input("No Data found, make sure you have input `Web Entry Identifier` in the Excel file..")
            sys.exit()
    
    def save_to_xls(self):
        for i in range(2, self.worksheet.max_row + 1):
            strdesc = ud.normalize('NFKD', str(self.worksheet['G{}'.format(i)].value).strip()).encode('ascii', 'ignore').decode('ascii')
            if strdesc == None:
                break
            for pn in self.pnlist:
                if self.worksheet['A{}'.format(i)].value == pn['webentry_id'] and strdesc == pn['description'] and self.worksheet['N{}'.format(i)].value == pn['consignee']:
                     self.worksheet['X{}'.format(i)].value = pn['pnnumber']
                     break
        try:        
            self.workbook.save(self.xlsname)
        except:
            input("Save to excel Failed!!. Make sure you have closed it. Run the script again.")
            sys.exit()
             

    def __data_generator(self):
        allData = {}
        wdesc = []
        wbox = []
        wconsignee = []
        wentryid = self.worksheet['A{}'.format(2)].value
        # print(wentryid)
        for i in range(2, self.worksheet.max_row + 1):
            if wentryid != self.worksheet['A{}'.format(i)].value:# and ws['B{}'.format(i)].value != None:
                allData[wentryid] = {'data':list(zip(wbox, wdesc, wconsignee))} 
                wentryid = self.worksheet['A{}'.format(i)].value
                clearlist(wbox, wdesc, wconsignee)
            if self.worksheet['G{}'.format(i)].value == None:
                break
            strdesc= ud.normalize('NFKD', str(self.worksheet['G{}'.format(i)].value).strip()).encode('ascii', 'ignore').decode('ascii')
            wdesc.append(strdesc)
            wbox.append(str(self.worksheet['D{}'.format(i)].value).strip())
            wconsignee.append(str(self.worksheet['N{}'.format(i)].value).strip())

        doc = fitz.open(self.filename)
        pdfpage = doc[0]
        self.pnlist.clear()
        entry_id = pdfpage.get_text("block", clip=(152.7100067138672, 202.04034423828125, 230.7493438720703, 214.09893798828125)).strip()
        for ds in allData:
            if ds == entry_id:
                for det in allData[ds]['data']:
                    dict = {
                        'webentry_id': entry_id,
                        'boxes': det[0],
                        'description': det[1],
                        'consignee': det[2],
                        'pnnumber': '' 
                    }
                    self.pnlist.append(dict)
        # jsn = json.dumps(self.pnlist)
        # print(jsn)
        # sys.exit()
    def combine_allpdf(self, pdffiles):
        print("Merging PDF files started..")
        time.sleep(1)
        dictfiles = {}
        result = fitz.open()
        for pdffile in pdffiles:
            basefilename = os.path.basename(pdffile)
            dictfiles[int(basefilename.replace(".pdf",""))] = pdffile
        sortedfiles = dict(sorted(dictfiles.items()))
        # print(sortedfiles)

        for file in sortedfiles:
            print(os.path.basename(sortedfiles[file]), "merged")
            mfile = fitz.open(sortedfiles[file])
            result.insert_pdf(mfile)
            time.sleep(1)
        result.save(self.pdfoutput_folder + self.file_delimeter + "combine.pdf")
        print(self.pdfoutput_folder + self.file_delimeter + "combine.pdf", "Created!")
        print("Merging PDF files finished..")

    @property
    def filename(self):
        return self.__filename


    @filename.setter
    def filename(self, value):
        self.__filename.set(value)

    @property
    def xlsname(self):
        return self.__xlsname


    @xlsname.setter
    def xlsname(self, value):
        self.__xlsname.set(value)

    @property
    def page(self):
        return self.__page

    @page.setter
    def page(self, value):
        self.__page.set(value)

    @property
    def sname(self):
        return self.__sname

    @sname.setter
    def sname(self, value):
        self.__sname.set(value)

    @property
    def word_list(self):
        return self.__word_list

    @property
    def workbook(self):
        return self.__workbook
   
    @property
    def worksheet(self):
        return self.__worksheet

    @property
    def pdfpage(self):
        return self.__pdfpage

    @property
    def savedfiles(self):
        return self.__savedfiles

    @property
    def pdfoutput_folder(self):
        return self.__pdfoutput

    @property
    def file_delimeter(self):
        return self.__delimeter
def file_delimeter():
    delimeter = "/"    
    if platform == "win32":
        delimeter = "\\"
    return delimeter
    
def del_non_annot_page(pdffiles, pdffolder):
    print("Removing Non Highlight Pages..")
    tmpfile = pdffolder + file_delimeter() + "tmp.pdf"
    for pdffile in pdffiles:
        shutil.copy(pdffile, tmpfile)
        doc = fitz.open(pdffolder + file_delimeter() + "tmp.pdf")
        selected = []
        for idx, page in enumerate(doc):
            for annot in page.annots():
                selected.append(idx)
                break
        selected.append(0)
        selected = list(dict.fromkeys(selected))
        selected.sort()
        doc.select(selected)
        doc.save(pdffile)
        print(os.path.basename(pdffile), "passed.")
        time.sleep(1)
    isExist = os.path.exists(tmpfile)
    doc.close()
    if isExist:    
        os.remove(tmpfile)    
    print("")

def main():
    parser = argparse.ArgumentParser(description="FDA PDF Extractor")
    parser.add_argument('-pdf', '--pdfinput', type=str,help="PDF File Input")
    parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
    parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    parser.add_argument('-output', '--pdfoutput', type=str,help="PDF output folder")
    args = parser.parse_args()
    # if args.pdfinput[-4:] != '.pdf':
    #     input('1st file input have to PDF file')
    #     sys.exit()
    # if args.xlsinput[-5:] != '.xlsx':
    #     input('2nd File input have to XLSX file')
    #     sys.exit()
        
    if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
        input('input the right XLSX or XLSM file')
        sys.exit()

    filelist = args.pdfinput.replace("('", '').replace("')","").replace("',)","").split("', '")
    for idx, filename in enumerate(filelist):
        isExist = os.path.exists(filename.strip())
        if not isExist:
            input(filename.strip() + " does not exist")
            sys.exit()
        else:
            filelist[idx] = filename.strip()

    isExist = os.path.exists(args.xlsinput)
    if not isExist:
        input(args.xlsinput + " does not exist")
        sys.exit()
    isExist = os.path.exists(args.pdfoutput)
    if not isExist:
        input(args.pdfoutput + " folder does not exist")
        sys.exit()

    allsavedfiles = []
    for filename in filelist:
        prior = PriorPdf(filename, 0, args.xlsinput, args.sheetname, args.pdfoutput)
        prior.highlightpdf_generator()
        prior.save_to_xls()
        prior.insert_text()
        allsavedfiles.extend(prior.savedfiles)
    setall = set(allsavedfiles)
    if len(setall) != len(allsavedfiles):
        input("Combining all pdf files Failed because there are one or more files is has the same name.")
    else:
        del_non_annot_page(allsavedfiles, args.pdfoutput)
        combine_allpdf(allsavedfiles, args.pdfoutput)
    input("data generating completed...")

if __name__ == '__main__':
    main()


  