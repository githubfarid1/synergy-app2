import requests
# import settings
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import time
import warnings
import argparse
import os

def parse(fileinput, chrome_data):
    warnings.filterwarnings("ignore", category=UserWarning)
    # cookies = settings.walmart_cookies
    headers = {
        'authority': 'www.walmart.ca',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-language': 'en-US,en;q=0.9,ja-JP;q=0.8,ja;q=0.7,id;q=0.6',
        'cache-control': 'max-age=0',
        # Requests sorts cookies= alphabetically
        # 'cookie': 'localStoreInfo=eyJwb3N0YWxDb2RlIjoiTDVWMk42IiwibG9jYWxTdG9yZUlkIjoiMTA2MSIsInNlbGVjdGVkU3RvcmVJZCI6IjEwNjEiLCJzZWxlY3RlZFN0b3JlTmFtZSI6IkhlYXJ0bGFuZCBTdXBlcmNlbnRyZSIsImZ1bGZpbGxtZW50U3RvcmVJZCI6IjEwNjEiLCJmdWxmaWxsbWVudFR5cGUiOiJJTlNUT1JFX1BJQ0tVUCIsImFzc2lnbmVkRmFsbGJhY2siOnRydWV9; deliveryCatchment=1061; walmart.nearestPostalCode=L5V2N6; walmart.shippingPostalCode=L5V2N6; defaultNearestStoreId=1061; headerType=whiteGM; vtc=cB3MtzAKzN9lrCMa0WP6gI; walmart.nearestLatLng="43.60822,-79.69387"; userSegment=40-percent; walmart.id=1a214437-89eb-4acb-a74a-7286853f4ddd; _gcl_au=1.1.981439298.1658271372; DYN_USER_ID=8fbd9d3c-f225-496f-a510-3e0d91a955b0; WM_SEC.AUTH_TOKEN=MTAyOTYyMDE4onPtxd1R9jkfaQfplXNM7asZvMxJ8q1HHNacT/CEqN9j0oE7Sdrm2VHl/Uh+VxL/+QvmSnvZ6Prwv20xKOaiK9G1ZW5JCjpQSRfdpHFGjOIXx+8ymisUxQx/eD5mHfhqj8OFN4dileb20bpDLeCIlSFd/Hsc7bnSe4+TLU2zbj29Yvukdblo2XIkqlJxsUc0hB/KkBC7B+0rrw4+JOTVETAqiLywgBYxbVZuaUPuL/bb/SoGFgAYL9DGZ8K45WCXM/FHGZ2dCNmxWrdkwqEKroZMTa42zTQOmBGJAW0EVdapHCxT5ZIYiB8yJC06hDWzamUxtd0UbTEz8Bw7QZFLxXsCoP033oSzI+1E3Bxyv/HMrPMEbgHCPLKaipmbw1E0sCvIZ2JKjrvPmvmQrQsvTEr1eX9YGQ0laieVMoEr348=; LT=1658271372514; DYN_USER_ID.ro=8fbd9d3c-f225-496f-a510-3e0d91a955b0; pxcts=fcddfb96-07b5-11ed-b698-45466856786d; _pxvid=fcddee61-07b5-11ed-b698-45466856786d; AMCVS_C4C6370453309C960A490D44%40AdobeOrg=1; _ga=GA1.2.1162591189.1658271374; _cs_c=1; _cs_cvars=%7B%221%22%3A%5B%22appName%22%2C%22product-page%22%5D%7D; s_ecid=MCMID%7C29511177273493667792583508179618052430; s_cc=true; wmt.c=0; NoCookie=true; BVImplmain_site=2036; cartId=fce2356c-5fe6-48d1-b5fe-9959793bd75f; cookiePolicy=true; s_sq=%5B%5BB%5D%5D; BVBRANDID=39bab97f-9c68-4e36-9f23-faabaaa6749e; TS017d5bf6=01538efd7c5c6997b7e0377cc25d4a5e960e1d15462d474c6f5aaf77abdd83e498fe5550bf44417874c25fa2970a42d2a0e3e815ea; TS01170c9f=01538efd7c5c6997b7e0377cc25d4a5e960e1d15462d474c6f5aaf77abdd83e498fe5550bf44417874c25fa2970a42d2a0e3e815ea; ENV=ak-scus-t1-prod; bstc=VxRErSdYb7bjHBfMaJeyQ4; xpa=--jYu|-99OH|2P1Of|2lwWQ|59Tck|5NWkV|7Xi3l|8C3ux|9LHzK|9jPmV|ATSnS|D9LOQ|EF9Bs|Ewx8F|F6obB|FtRSv|Gi9em|H33MR|H67lQ|HRTLp|I9Jw3|IHCVo|IVqRT|J6lpV|JXvsb|JoXRc|JvV8u|KeAz7|LVSOt|MbHLZ|NOaJP|NPyyt|NbKZN|Nmfwy|OVStb|PHaOk|Qoi7g|T4DD-|TTwy3|V3_qS|W-D6M|YmdKw|_vHmT|_vY-K|a4fYU|aVJBH|cFFZi|ja1-z|jeBOs|ldwJI|m0qtG|mOlOu|mfKI8|pCsPF|pQ73W|rWfNO|rkVFM|sZixZ|t18ca|tnV6l|uBlRm|u_M4R|v-FRz|vm9yl|wKkZZ|xafoR|yAFBx|zhR2K; exp-ck=--jYu12P1Of159Tck15NWkV17Xi3l18C3ux39jPmV1ATSnS3D9LOQ2EF9Bs1Ewx8F1F6obB1FtRSv1H33MR1H67lQ6I9Jw31IHCVo1IVqRT1JXvsb1JoXRc1JvV8u1KeAz71MbHLZ1NPyyt4NbKZN5Nmfwy1OVStb1Qoi7g1T4DD-5TTwy31V3_qS1YmdKw1_vHmT1ja1-z1mfKI81pCsPF1pQ73W1rWfNO4rkVFM1sZixZ1t18ca1tnV6l3u_M4R2v-FRz6vm9yl1wKkZZ1yAFBx1zhR2K6; TS0196c61b=01538efd7c140251298434802589714ec9e15c3cf5141715bfc11fc3566927ff5e830fe01b080f021d277ac02560b1898083525327; ak_bmsc=F2E7D14E9B57B7B5BB535EE585552DFA~000000000000000000000000000000~YAAQL6IAF/cIQSOCAQAAhmd/KBDYVg/39n77aCNHsXj6gsZsFOyYEZgb42ywSpXszmCRZ6Te3X5w+ahliGc2UqN3uNf+l+r4W4rD3ocwAA4fP+I1Dsms+ddOZRNPSuL2DYph35NZkvynpt4wOR/DQNg5KxJ5gYU/MQCqqnXcCljMBbUcff7bWtWKgf2s8YuLnVvwZxhVSPglxglfF/Umesh4KXyphHwkmBLI/uSSFs00WOzqrVAlS6+rCMY5jjFGl4hqE8uyjGHYAYjK1Vy4nwBXLUbxNV1u/9xFaX5QWMhUw5lvYyQjIvxrWSUMIr/5paTqgTZvtU8djLbVJDbYv+fFp+wfCVWb6/3EGqg1T1hWryuD1dioGxEf9arfUJ51fKoqOp+7TBoj7nk=; xpm=1%2B1658536813%2BcB3MtzAKzN9lrCMa0WP6gI~%2B0; s_gnr=1658536814836-Repeat; _cs_mk_aa=0.7028220796911733_1658536814839; WM.USER_STATE=GUEST|Guest; authDuration={"lat":"1658536816660000","lt":"1658536816660000"}; NEXT_GEN.ENABLED=1; AMCV_C4C6370453309C960A490D44%40AdobeOrg=-1124106680%7CMCIDTS%7C19197%7CMCMID%7C29511177273493667792583508179618052430%7CMCAAMLH-1659141615%7C3%7CMCAAMB-1659141615%7CRKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y%7CMCOPTOUT-1658544015s%7CNONE%7CMCAID%7CNONE%7CvVersion%7C5.2.0; cto_bundle=pSJseF80QWRiNm1Ic1RxUkZwS1pnY1dYYkIxeEk5OWxYWXFUeDZ1JTJCWlphQWl3TVdqTzBYMUZnS3NRMSUyRm9SRWtKR3g4aWlzS3c5bVFWcTdJRk9EWiUyQktzd2RKOFBkdyUyRnNBeGpLMXVLMUlPbGdlZVF0NUo3emptZHZJclVUMSUyRnlHcDg5cDRSb3AyYWpIJTJGNTU2T1V0JTJGaVVjbUdwUSUzRCUzRA; _cs_id=99dce351-6402-ae0e-9c22-e5e4e742c495.1658271374.7.1658536815.1658536825.1.1692435374226; _cs_s=2.0.0.1658538615989; _gid=GA1.2.1346150038.1658536816; s_visit=1; gpv_Page=Product%3ADoritos%20Nacho%20Cheese%20Tortilla%20Chips; _px3=a6eb2425434d633a4a28379b15aa29504a7a03776b9f625cdd2b593a6b6838fd:O8tqudaiF/DcEVi1K31vjtMP3yzmRpmvaCxXy+eGnyrVvn2wX2p/Un6ZD1rn2VkFmjB++BTtli1MxhjjO8FyUw==:1000:6I8Es4ImfdSFN6vMycTaw/gVDKBgeQaOiUh0uAzL6bIRLnfhsm5gAgTFB8VbuclwPDghDZBgUyA6AL3+p9MAueASRghcdHahKqumU++ruCXT3SKwjihdYFCMziXpnIrVD8OWOjG7v2xcMgwaEx0Vu8BO6KiY23FXy6LFup1I3R0SYCvBTD0bi+2Jik3b+Tb7T1Sen/MSGyYsqwjIHLCAjg==; _gat=1; kndctr_C4C6370453309C960A490D44_AdobeOrg_identity=CiYyOTUxMTE3NzI3MzQ5MzY2Nzc5MjU4MzUwODE3OTYxODA1MjQzMFIPCJjGtMWhMBgBKgRTR1Az8AGB7v3DojA=; kndctr_C4C6370453309C960A490D44_AdobeOrg_cluster=sgp3; bm_sv=CA563A9D1523E172DCCFA466C51EFA33~YAAQL6IAF08LQSOCAQAA2px/KBCVwaGx6qk0C8+9IEXm6Q6Nz/eBmQFByz162567W2AXNPvC2OXTO9UhuyxH0CLumLAkwcJld90AORSqQwHcaA+IyRdCjTVE1bxlNqgP8N5oiGuHxjq60ZSTPNMubhlwJKGHmN3XbypqqDlbTW+VjOBOljvToTDnJsIkiqTkM3qH3UEtiGfx5fjLfdrocEjbTkyDl3W3m0pPS36RORDyeBVsWQZZQ08zuuZJqCjw~1; seqnum=7',
        'sec-ch-ua': '".Not/A)Brand";v="99", "Google Chrome";v="103", "Chromium";v="103"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    }

    # print(cookies)
    # UPDATE COOKIES
    os.system('cls')
    print('File Selected:', fileinput)

    sess = requests.Session()
    sess.get('https://www.walmart.ca/en/ip/6000189529187', headers=headers)
    cookies = sess.cookies.get_dict() 


    wb = load_workbook(filename=fileinput, read_only=False)
    ws = wb['Sheet1']
    for i in range(3, ws.max_row + 1):
        if ws['A{}'.format(i)].value == None:
            break

        url = ws['A{}'.format(i)].value
        s1 = requests.Session()
        response = s1.get(url, cookies=cookies, headers=headers)
        found = response.text.find('"upc":["')
        # print(found)
        upc = ''
        if found != -1:
            start = 8
            while True:
                if response.text[found+start] != '"':
                    upc += response.text[found+start]
                else:
                    break
                start += 1
            # upc = response.text[found+8:found+8+10]
        print(url)
        try:
            postal1 = ws['B1'].value.strip()
            s2 = requests.Session()
            response = s2.get("https://www.walmart.ca/api/product-page/geo-location?postalCode={}".format(postal1), cookies=cookies, headers=headers)
            # print(response)
            lat = response.json()
            response = s2.get("https://www.walmart.ca/api/product-page/find-in-store?latitude={}&longitude={}&lang=en&upc={}".format(lat['lat'], lat['lng'], upc), cookies=cookies, headers=headers)    
            data1 = response.json()
            cell = ['B','C','D']
            for idx, d in enumerate(data1['info']):
                if d['availabilityStatus'] == 'AVAILABLE':
                    ws['{}{}'.format(cell[idx], i)].value = 'a'   
                else:
                    ws['{}{}'.format(cell[idx], i)].value = 'x'   



            postal2 = ws['E1'].value.strip()
            response = s2.get("https://www.walmart.ca/api/product-page/geo-location?postalCode={}".format(postal2), cookies=cookies, headers=headers)
            lat = response.json()
            response = s2.get("https://www.walmart.ca/api/product-page/find-in-store?latitude={}&longitude={}&lang=en&upc={}".format(lat['lat'], lat['lng'], upc), cookies=cookies, headers=headers)    
            data2 = response.json()
            cell = ['E','F','G']
            for idx, d in enumerate(data2['info']):
                if d['availabilityStatus'] == 'AVAILABLE':
                    ws['{}{}'.format(cell[idx], i)].value = 'a'   
                else:
                    ws['{}{}'.format(cell[idx], i)].value = 'x'   

            postal3 = ws['H1'].value.strip()
            response = s2.get("https://www.walmart.ca/api/product-page/geo-location?postalCode={}".format(postal3), cookies=cookies, headers=headers)
            lat = response.json()
            response = s2.get("https://www.walmart.ca/api/product-page/find-in-store?latitude={}&longitude={}&lang=en&upc={}".format(lat['lat'], lat['lng'], upc), cookies=cookies, headers=headers)    
            data3 = response.json()
            cell = ['H','I','J']
            for idx, d in enumerate(data3['info']):
                if d['availabilityStatus'] == 'AVAILABLE':
                    ws['{}{}'.format(cell[idx], i)].value = 'a'   
                else:
                    ws['{}{}'.format(cell[idx], i)].value = 'x'   

        except:
            print('Please open chrome browser', url, 'and pass captcha manually, then try run the script again!')
            input('')
            exit()
        time.sleep(2)
        # break
    # exit()
    wb.save(fileinput)
    input('Finished...')


def main():
    parser = argparse.ArgumentParser(description="Scrape Walmart")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    parser.add_argument('-d', '--data', type=str,help="Chrome User Data Directory")
    args = parser.parse_args()
    if args.input[-5:] != '.xlsx':
        print('File input have to XLSX file')
        exit()
    isExist = os.path.exists(args.data)
    if isExist == False :
        print('Please check Chrome User Data Directory')
        exit()
    isExist = os.path.exists(args.input)
    if isExist == False :
        print('Please check XLSX file')
        exit()
    
    parse(args.input, args.data)
    
if __name__ == '__main__':
    main()

