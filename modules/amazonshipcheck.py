from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
from selenium.webdriver.common.action_chains import ActionChains
import sys
import os
import argparse
import time
from openpyxl import Workbook, load_workbook
from sys import platform
import json
from random import randint
from datetime import date, datetime, timedelta
import warnings
import logging
from pathlib import Path

logger = logging.getLogger()
logger.setLevel(logging.NOTSET)

logger2 = logging.getLogger()
logger2.setLevel(logging.NOTSET)


def clearlist(*args):
    for varlist in args:
        varlist.clear()

def explicit_wait():
    time.sleep(randint(1, 3))

def clear_screan():
    try:
        if platform == "win32":
            os.system("cls")
        else:    
            os.system("clear")
    except Exception as er:
        print(er, "Command is not supported")

def pause(mess=""):
    input(mess)

def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def killAllChrome():
    if platform == "win32":
        os.system("taskkill /f /im chrome.exe")

def checkdimension(value=""):
    valuelist = value.upper().split('X')
    if len(valuelist) != 3:
        return False
    
    for v in valuelist:
        if v.isnumeric() == False:
            return False
    
    return True
    
def file_delimeter():
    dm = "/" 
    if platform == "win32":
        dm = "\\"
    return dm

class AmazonShipmentCheck:
    def __init__(self, xlsfile, sname, profile) -> None:
        try:
            self.__workbook = load_workbook(filename=xlsfile, read_only=False, keep_vba=True, data_only=True)
            self.__worksheet = self.__workbook[sname]
        except Exception as e:
            logger.error(e)
            input("XLSX file or Sheet name not found")
            sys.exit()
        self.__profile = profile

        self.__datalist = []
        self.__xlsfile = xlsfile
        self.__delimeter = "/" 
        if platform == "win32":
            self.__delimeter = "\\"
        clear_screan()

        self.__driver = self.__browser_init()
        self.__data_generator()

    def __browser_init(self):
        config = getConfig()
        warnings.filterwarnings("ignore", category=UserWarning)
        options = webdriver.ChromeOptions()
        # options = Options()
        # options.add_argument("--headless")
        options.add_argument("user-data-dir={}".format(getProfiles()[self.profile]['chrome_user_data'])) 
        options.add_argument("profile-directory={}".format(getProfiles()[self.profile]['chrome_profile']))
        options.add_argument('--no-sandbox')
        options.add_argument("--log-level=3")
        # options.add_argument("--window-size=1200, 900")
        options.add_argument('--start-maximized')
        options.add_argument("--disable-notifications")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        # return webdriver.Chrome(service=Service(CM(version="114.0.5735.90").install()), options=options)
        return webdriver.Chrome(service=Service(executable_path=os.path.join(os.getcwd(), "chromedriver", "chromedriver.exe")), options=options)

    def __data_generator(self):
        print("Data Mounting... ", end="")
        shipmentlist = []
        for i in range(2, self.worksheet.max_row + 1):
            shipment_row = str(self.worksheet['A{}'.format(i)].value)
            if shipment_row.find('Shipment') != -1:
                # print(shipment_row, i)
                startrow = i
                y = i
                shipment_empty = True
                while True:
                    y += 1
                    # skip if shipment_id was filled
                    if ''.join(str(self.worksheet['B{}'.format(y)].value)).strip() == 'Shipment ID':
                        if ''.join(str(self.worksheet['E{}'.format(y)].value)).strip() != 'None':
                            shipment_empty = False

                    if str(self.worksheet['B{}'.format(y)].value) == 'Tracking Number':
                        endrow = y + 1
                        i = y + 1
                        break
                
                if shipment_empty == True:
                    shipmentlist.append({'begin':startrow, 'end':endrow})
                else:
                    logger2.critical(shipment_row + " Shipment ID Found, Skipped")

        # print(json.dumps(shipmentlist))
        for index, shipmentdata in enumerate(shipmentlist):
            shipmentlist[index]['submitter'] = self.worksheet['B{}'.format(shipmentdata['begin'])].value
            shipmentlist[index]['address'] = self.worksheet['B{}'.format(shipmentdata['begin']+1)].value
            shipmentlist[index]['name'] = self.worksheet['A{}'.format(shipmentdata['begin'])].value
            boxes = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
            boxcount = 0
            for box in boxes:
                if self.worksheet['{}{}'.format(box, shipmentdata['begin'])].value != None:
                    boxcount += 1
                else:
                    break
            if boxcount == 0:
                del shipmentlist[index]
                continue
            shipmentlist[index]['boxcount'] = boxcount
            start = shipmentdata['begin'] + 2
            shipmentlist[index]['weightboxes'] = []
            shipmentlist[index]['dimensionboxes'] = []
            shipmentlist[index]['nameboxes'] = []
            shipmentlist[index]['items'] = []

            # get weightboxes
            rowsearch = 0
            for i in range(start, shipmentdata['end']):
                if self.worksheet['B{}'.format(i)].value == 'Weight':
                    rowsearch = i
                    break
            
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['weightboxes'].append(self.worksheet['{}{}'.format(box, rowsearch)].value)

            # get dimensionboxes
            rowsearch = 0
            for i in range(start, shipmentdata['end']):
                if self.worksheet['B{}'.format(i)].value == 'Dimensions':
                    rowsearch = i
                    break
            
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['dimensionboxes'].append(self.worksheet['{}{}'.format(box, rowsearch)].value)

            #get nameboxes
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['nameboxes'].append(str(self.worksheet['{}{}'.format(box, shipmentdata['begin'])].value))

            ti = -1
            for i in range(start, shipmentdata['end']):
                ti += 1
                if self.worksheet['A{}'.format(i)].value == None or str(self.worksheet['A{}'.format(i)].value).strip() == '':
                    break
                # shipmentlist[index]['items'].append()
                
                dict = {
                    'id': self.worksheet['A{}'.format(i)].value,
                    'name': self.worksheet['B{}'.format(i)].value,
                    'total': self.worksheet['C{}'.format(i)].value,
                    'expiry': str(self.worksheet['D{}'.format(i)].value),
                    'boxes':[],

                }

                shipmentlist[index]['items'].append(dict)
                for ke, box in enumerate(boxes):
                    if ke == boxcount:
                        break
                    if self.worksheet['{}{}'.format(box, i)].value == None or str(self.worksheet['{}{}'.format(box, i)].value).strip() == '':
                        shipmentlist[index]['items'][ti]['boxes'].append(0)
                    else:                           
                        shipmentlist[index]['items'][ti]['boxes'].append(self.worksheet['{}{}'.format(box, i)].value)

        
        #cleansing
        idxdel = []
        for index, shipmentdata in enumerate(shipmentlist):
            try:
                cheat = shipmentdata['name']
            except:
                idxdel.append(index)
        
        for idx in idxdel:
            for index, shipmentdata in enumerate(shipmentlist):
                try:
                    cheat = shipmentdata['name']
                except:
                    del shipmentlist[index]
                
            # pass
        
        self.datalist = shipmentlist
        explicit_wait()
        print("Passed")

    def data_sanitizer(self):
        print("Try to login... ", end="")
        url = "https://sellercentral.amazon.ca/fba/sendtoamazon?ref=fbacentral_nav_fba"
        self.driver.get(url)
        print("Check SKU page ready... ", end="")
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
            checksku = self.driver.find_element(By.CSS_SELECTOR,"kat-tabs[id='skuTabs']").find_element(By.CSS_SELECTOR, "kat-tab-header[tab-id='3']").find_element(By.CSS_SELECTOR, "span[slot='label']").text
            if checksku != 'SKUs ready to send (0)':
                raise Exception('SKUs ready to send is not 0')
        except Exception as er:
            logger.error(er)
            logger.info("Trying to click start new link..")
            print("Trying to click start new link..", end="")
            try:
                self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
                print("Passed")
            except Exception as e:
                logger.error(e)
                print("Failed")
                sys.exit()

        explicit_wait()
        print('Checking Excel Data..')
        #before
        # self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").click()
        # after
        button = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']")))
        button.click()

        explicit_wait()
        # input("pause2")
        shadow_host = self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']")
        shadow_root = shadow_host.shadow_root
        shadow_root.find_element(By.CSS_SELECTOR, "kat-option[tabindex='-1'").click()

        # self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='MSKU']").click()
        explicit_wait()
        idxdel = []
        logger2.critical("\n")
        logger2.critical("Trying to input all shipment to Amazon")
        logger2.critical("-------------------------")

        for idx, dlist in enumerate(self.datalist):
            print(dlist['name'], "... ", end="")
            # logger2.critical("{} Inspect...".format(dlist['name']))

            shipping_name = dlist['name']
            error = False
            errorlist = []
            notelist = []
            defsubmitter = self.driver.find_element(By.CSS_SELECTOR, "div[class='textBlock-60ch break-words']").text

            submitter = dlist['submitter'].split("(")[0].strip()
            addresstmp = dlist['address']
            addresslist = addresstmp[addresstmp.find("(")+1:addresstmp.find(")")].strip().split(" ")
            address = addresslist[0] + " " + addresslist[1]# + " " + addresslist[2]
            address_found = False
            if defsubmitter.find(submitter) != -1 and defsubmitter.find(address) != -1:
                address_found = True
            else:
                addresslink = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[data-testid='ship-from-another-address-link']")))
                addresslink.click()
                try:
                    WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='selected-address-tile']")))
                except:
                    input("Address list failed to open")
                    sys.exit()

                selects = self.driver.find_elements(By.CSS_SELECTOR, "div[class='address-tile']")
                explicit_wait()
                address_found = False
                # breakpoint()
                for idx, sel  in enumerate(selects):
                    txt = sel.find_element(By.CSS_SELECTOR, "div[class='tile-address']").text
                    if txt.find(submitter) != -1 and txt.find(address) != -1:
                        # shadow_host = self.driver.find_element(By.
                        # CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']")
                        # shadow_root = shadow_host.shadow_root
                        # shadow_root.find_element(By.CSS_SELECTOR, "kat-option[tabindex='-1'").click()
                        # input("p1")
                        # breakpoint()
                        shadow_host = sel.find_element(By.CSS_SELECTOR, "kat-button.tile-selection-button")
                        try:
                            shadow_host.click()
                        except:
                            breakpoint()
                            shadow_root = shadow_host.shadow_root
                            shadow_root.find_element(By.CSS_SELECTOR, "button.button").click()
                        address_found = True
                        break

            if not address_found:
                errorlist.append("Address or Submitter not Found")
                # logger2.critical("Address or Submitter not Found")
                error = True
                self.driver.find_element(By.CSS_SELECTOR, "div[class='selected-address-tile']").find_element(By.CSS_SELECTOR, "button[class='secondary']").send_keys(Keys.ESCAPE)
            
            if len(dlist['dimensionboxes']) == 0:
                error = True
                errmsg = "dimension value is Empty"
                # print(errmsg)
                errorlist.append(errmsg)

            for dim in dlist['dimensionboxes']:
                if checkdimension(dim) == False:
                    error = True
                    errmsg = "{} dimension box value is wrong".format(dim)
                    # print(errmsg)
                    errorlist.append(errmsg)
 
            wbox = "".join(str(x) for x in dlist['weightboxes'])
            if wbox.isnumeric() == False:
                error = True
                errmsg = "Weight box value is wrong"
                # print(errmsg)
                errorlist.append(errmsg)

            for idx2, item in enumerate(dlist['items']):
                shadow_host = self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']")
                shadow_root = shadow_host.shadow_root
                xlssku = item['id'].upper()
                shadow_root.find_element(By.CSS_SELECTOR, "input").clear()
                shadow_root.find_element(By.CSS_SELECTOR, "input").send_keys(xlssku)
                searchinput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[data-testid='search-input-link']")))

                # self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").clear()
                # xlssku = item['id'].upper()
                # self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").send_keys(xlssku)
                # explicit_wait()
                # searchinput = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[data-testid='search-input-link']")))

                # breakpoint()
                searchinput.click()
                explicit_wait()
                cols = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='sku-row-information-details']")
                sku = ''
                try:
                    sku = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='msku']").find_element(By.CSS_SELECTOR, "span").text
                except:
                    error = True
                    errorlist.append(xlssku + ' Not Found')
                    # logger2.critical('{} Not Found'.format(xlssku))

                if xlssku != sku:
                    errorlist.append(sku + ' Not Match')
                    # logger2.critical('{} Not Match'.format(sku))
                    error = True
                errormsg = ''
                try:
                    errormsg = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='sku-action-info']").find_element(By.CSS_SELECTOR, "span[data-testid='sku-action-error-text']").text
                except:
                    pass
                if errormsg != '':
                    errorlist.append(errormsg)
                    # logger2.critical('Error: {}'.format(errormsg))
                    error = True

                try:
                    individual = WebDriverWait(cols[0], 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='packing-template-dropdown']")))
                except Exception as e:
                    logger.error(e)
                    input(xlssku + " Not Found or internet error. Please check the SKU and run the script again!")
                    sys.exit()

                # breakpoint()
                # if individual.text.find('Individual units') == -1:
                # input("pause")
                # if individual.find_element(By.CSS_SELECTOR, "kat-option[value='1']").get_attribute("aria-selected") == "false":
                shadow_host = self.driver.find_element(By.CSS_SELECTOR,"kat-dropdown[data-testid='packing-template-dropdown']")
                shadow_root = shadow_host.shadow_root
                if shadow_root.find_element(By.CSS_SELECTOR, "div.kat-select-container").get_attribute("title") != "Individual units":
                
                    individual.click()
                    explicit_wait()

                    # shadow_root = individual.shadow_root
                    
                    individual.find_element(By.CSS_SELECTOR, "kat-option[data-testid='packing-template-Individual-units']").click()
                    
                    # individual.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-name='Individual units']").click()
                # breakpoint()
                # input("pause")
                if cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='sku-action-info'").text.find('Prep not required') == -1:
                    # breakpoint()
                    try:
                        try:
                            infoprep = cols[0].find_element(By.CSS_SELECTOR, "kat-link[data-testid='sku-action-info-prep-missing-link']")
                        except:
                            infoprep = cols[0].find_element(By.CSS_SELECTOR, "kat-link[data-testid='prep-modal-link']")
                        explicit_wait()
                        infoprep.click()
                        catprep = self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='prep-guidance-prep-category-dropdown']")
                        explicit_wait()
                        catprep.click()
                        catprep.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='NONE']").click()
                        explicit_wait()
                        self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[variant='primary']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                        explicit_wait()
                        self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-save-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                    except:
                        try:
                            # self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='prep-category-update-btn']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                            shadow_host = self.driver.find_element(By.CSS_SELECTOR,"kat-button[data-testid='prep-category-update-btn']")
                            shadow_root = shadow_host.shadow_root
                            shadow_root.find_element(By.CSS_SELECTOR,"button.button").click()
                            explicit_wait()

                            # self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-save-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                            shadow_host = self.driver.find_element(By.CSS_SELECTOR,"kat-button[data-testid='packing-template-save-button']")
                            shadow_root = shadow_host.shadow_root
                            shadow_root.find_element(By.CSS_SELECTOR,"button.button").click()
                        except:                        
                            pass
                try:
                    WebDriverWait(cols[0], 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR , "span[data-testid='prep-fee-text']"), "Prep not required"))
                except:
                    try:
                        WebDriverWait(cols[0], 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR , "span[data-testid='prep-fee-text']"), "Unit prep: By seller"))
                    except:
                        try:
                            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-cancel-button']").find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                            error = True
                            errorlist.append(sku + " Prep Failed")
                        except:
                            error = True
                            errorlist.append(sku + " Prep Failed")
                            pass
                # except:
                #     try:
                #         WebDriverWait(cols[0], 10).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR , "span[data-testid='prep-fee-text']"), "Unit prep: By seller"))
                #     except:
                #         pass
                
                
                try:
                    numunit = cols[0].find_element(By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-units-input']").find_element(By.CSS_SELECTOR, "input[name='numOfUnits']")
                    if not numunit.is_enabled():
                        errorlist.append(sku + " unit number disabled")
                        error = True                     
                except:
                    errorlist.append(sku + " unit number disabled")
                    error = True

                now = datetime.now()
                maxdate = now + timedelta(days=105)
                strexpiry = item['expiry'].strip()
                dformat = '%Y-%m-%d %H:%M:%S'
                dateinput = True
                if  strexpiry == 'None' or strexpiry == 'N/A':
                    dexpiry = now + timedelta(days=365)
                    notelist.append('{}: date expiry is empty, it will be adjusted to now +1 year'.format(xlssku))

                else:
                    try:
                        dexpiry = datetime.strptime(strexpiry, dformat)
                    except ValueError:
                        dateinput = False
                        error = True
                        errorlist.append("{}: wrong date expiry value".format(xlssku))                

                if dateinput == True:
                    if dexpiry < maxdate:
                        dexpiry = now + timedelta(days=365)
                        notelist.append("{}: date expiry is less than 105 days, it will be adjusted to now +1 year".format(xlssku))
                        # print(xlssku, "date expiry adjusted to now +1 year ", end="")
                        

                try:
                    expiry = dexpiry.strftime('%m/%d/%Y')
                except:
                    expiry = strexpiry

                try:
                    # breakpoint()
                    inputexpiry = cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']")
                    if inputexpiry.is_enabled():
                        inputexpiry.send_keys(expiry)
                        inputexpiry.send_keys(Keys.TAB)
                except:
                    pass


                errormsg = ''
                try:
                    errormsg = cols[0].find_element(By.CSS_SELECTOR, "kat-label[data-testid='sku-readiness-expiration-date-error']").text
                except:
                    pass
                if errormsg != '':
                    errorlist.append(errormsg)
                    error = True
                
                boxes = "".join(str(x) for x in item['boxes'])
                if boxes.isnumeric() == False:
                    error = True
                    errmsg = "{}: Boxes value is wrong".format(xlssku)
                    # print(errmsg)
                    errorlist.append(errmsg)
                
                if str(item['total']).isnumeric() == False:
                    error = True
                    errmsg = "{}: Total value is wrong".format(xlssku)
                    # print(errmsg)
                    errorlist.append(errmsg)

                # if error:
                #     break
                

            if error:
                print("Failed")
                # print("".join(errorlist))
                idxdel.append(dlist['name'])
                logger2.critical(dlist['name'] + " Skipped..")
                logger2.critical("Error:")
                logger2.critical("\n".join(errorlist))
                if len(notelist) != 0:
                    logger2.critical("Info:")
                    logger2.critical("\n".join(notelist))
                # del self.datalist[idx]
            else:
                logger2.critical(dlist['name'] + " OK..")
                if len(notelist) != 0:
                    logger2.critical("Info:")
                    logger2.critical("\n".join(notelist))
                print("Passed")
            
            logger2.critical("-------------------------")
        # input('pause')
        # print(idxdel)
       
        # time.sleep(5)

    @property
    def workbook(self):
        return self.__workbook
   
    @property
    def worksheet(self):
        return self.__worksheet

    @property
    def datalist(self):
        return self.__datalist

    @datalist.setter
    def datalist(self, value):
        self.__datalist = value

    @property
    def datajson(self):
        return json.dumps(self.datalist)  

    @property
    def xlsfile(self):
        return self.__xlsfile
    
    @xlsfile.setter    
    def xlsfile(self, value):
        self.__xlsfile = value

    @property
    def file_delimeter(self):
        return self.__delimeter

    @property
    def driver(self):
        return self.__driver
    
    @property
    def profile(self):
        return self.__profile

def main():
    parser = argparse.ArgumentParser(description="Amazon Shipment Check")
    parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
    parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    parser.add_argument('-profile', '--profile', type=str,help="Chrome Profile Selected")
    args = parser.parse_args()
    if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
        input('input the right XLSX or XLSM file')
        sys.exit()
    isExist = os.path.exists(args.xlsinput)
    if not isExist:
        input(args.xlsinput + " does not exist")
        sys.exit()


    filename = "amazon-shipping-{}-{}-{}-{}.log"
    # the second handler is a file handler
    foldername = os.path.dirname(args.xlsinput)
    foldername = foldername + file_delimeter() + "Shipment Reports"
    isExist = os.path.exists(foldername)
    if not isExist:
        os.makedirs(foldername) 
    
    
    file_handler = logging.FileHandler('logs/'+ filename.format("error", Path(args.xlsinput).stem, args.sheetname, date.today()) )
    file_handler.setLevel(logging.ERROR)
    file_handler_format = '%(asctime)s | %(levelname)s | %(lineno)d: %(message)s'
    file_handler.setFormatter(logging.Formatter(file_handler_format))
    logger.addHandler(file_handler)
    reportfilename = foldername + file_delimeter() + filename.format("report", Path(args.xlsinput).stem, args.sheetname, date.today())
    if os.path.exists(reportfilename):
        os.remove(reportfilename)
    file_handler2 = logging.FileHandler(reportfilename)
    file_handler2.setLevel(logging.CRITICAL)
    # file_handler2_format = '%(asctime)s | %(levelname)s: %(message)s'
    file_handler2_format = '%(message)s'
    file_handler2.setFormatter(logging.Formatter(file_handler2_format))
    logger2.addHandler(file_handler2)


    logger2.critical("###### Start ######")
    logger2.critical("Filename: {}".format(args.xlsinput))
    logger2.critical("Sheet Name:{}".format(args.sheetname))
    # maxrun = 10
    # for i in range(1, maxrun+1):
    #     if i > 1:
    #         print("Process will be reapeated")
    #     try:    
    shipment = AmazonShipmentCheck(xlsfile=args.xlsinput, sname=args.sheetname, profile=args.profile)
    # print(shipment.datajson)
    # raise
    # if len(shipment.datalist) == 0:
    #     break
    shipment.data_sanitizer()
            
        # except Exception as e:
        #     # exit()
        #     logger.error(e)
        #     print("There is an error, check logs/amazonship-err.log")
        #     if i == maxrun:
        #         logger.error("Execution Limit reached, Please check the script")
        #     continue
        
        # break
    # killAllChrome()
    print("Report File generated: {}".format (reportfilename))
    input("End Process..")    
if __name__ == '__main__':
    main()
