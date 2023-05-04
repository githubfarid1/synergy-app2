from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
from selenium.webdriver.support.select import Select
import sys
import fitz
import shutil
import os
import argparse
import time
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import unicodedata as ud
from sys import platform
from os import walk
import json
from random import randint
from datetime import date
import warnings
import glob

def clearlist(*args):
    for varlist in args:
        varlist.clear()

def explicit_wait():
    time.sleep(randint(2, 3))

def clear_screan():
        try:
            os.system("cls")
        except:
            os.system("clear")

class AmazonShipment:
    def __init__(self, xlsfile, sname, chrome_data, download_folder) -> None:
        try:
            self.__workbook = load_workbook(filename=xlsfile, read_only=False, data_only=True)
            self.__worksheet = self.__workbook[sname]
        except:
            input("XLSX file or Sheet name not found")
            sys.exit()
        self.__datajson = json.loads("{}")
        self.__datalist = []
        self.__data_generator()
        self.__chrome_data = chrome_data
        self.__download_folder = download_folder
        self.__xlsfile = xlsfile
        self.__delimeter = "/"    
        if platform == "win32":
            self.__delimeter = "\\"

        warnings.filterwarnings("ignore", category=UserWarning) 
        options = webdriver.ChromeOptions()
        # options.add_argument("--headless")
        options.add_argument("user-data-dir={}".format(self.chrome_data)) #Path to your chrome profile
        options.add_argument('--no-sandbox')
        options.add_argument("--log-level=3")
        options.add_argument("--window-size=1200, 900")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        # profile = {"plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
        #             "download.default_directory": self.download_folder, "download.extensions_to_open": "applications/pdf", 
        #             'profile.default_content_setting_values.automatic_downloads': 1}
        profile = {"download.default_directory": self.download_folder}
        options.add_experimental_option("prefs", profile)
        self.__driver = webdriver.Chrome(service=Service(CM().install()), options=options)
        self.__data_sanitizer()

    def parse(self):
        url = "https://sellercentral.amazon.com/fba/sendtoamazon?ref=fbacentral_nav_fba"
        self.driver.get(url)
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']")))
        except:
            try:
                self.driver.find_element(By.CSS_SELECTOR, "input[id='signInSubmit']").click()
                
            except:
                input("Please click `Chrome Tester` menu, then login manually, then close the browser and try the script again")
                sys.exit()
        explicit_wait()
        self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))

        defsubmitter = self.driver.find_element(By.CSS_SELECTOR, "div[class='textBlock-60ch break-words']").text
        first = True
        for idx, dlist in enumerate(self.datalist):
            # if idx <= 21:
            #     continue
            original_window = self.driver.current_window_handle
            submitter = dlist['submitter'].split("(")[0].strip()
            addresstmp = dlist['address']
            addresslist = addresstmp[addresstmp.find("(")+1:addresstmp.find(")")].strip().split(" ")
            address = addresslist[0] + " " + addresslist[1]# + " " + addresslist[2]
            # print(defsubmitter)
            print('#' * 5, dlist['name'], 'Start Process..', '#' * 5)
            if defsubmitter.find(submitter) != -1 and defsubmitter.find(address) != -1:
                # print('sama')
                print("Ship from label OK")
                pass
            else:
                print("Ship from Label Choosing..")
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "a[data-testid='ship-from-another-address-link']").click()
                ck = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='selected-address-tile']")))
                selects = self.driver.find_elements(By.CSS_SELECTOR, "div[class='address-tile']")
                for sel  in selects:
                    txt = sel.find_element(By.CSS_SELECTOR, "div[class='tile-address']").text
                    # print(txt)
                    if txt.find(submitter) != -1 and txt.find(address) != -1:
                        sel.find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                        # input('pause')
                        break
                explicit_wait()
                defsubmitter = self.driver.find_element(By.CSS_SELECTOR, "div[class='textBlock-60ch break-words']").text
                print("Ship from label OK")
            # PENTING
            # UNTUK CHROME DEVELOPER TOOLS AGAR BISA DEBUG SELECT
            # setTimeout(() => {debugger;}, 3000)                         
            explicit_wait()
            # if first == True:
            self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").click()
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='MSKU']").click()
            explicit_wait()
                # first = False
            for item in dlist['items']:
                self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").clear()
                xlssku = item['id'].upper()
                print('searching', xlssku, '..')
                self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").send_keys(xlssku)
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "a[data-testid='search-input-link']").click()
                explicit_wait()
                cols = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='sku-row-information-details']")
                try:
                    sku = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='msku']").find_element(By.CSS_SELECTOR, "span").text
                except:
                    input(xlssku + " not found!")
                    sys.exit()
                if xlssku != sku:
                    input(sku + " not found!")
                    sys.exit()
                else:
                    print(xlssku, 'has Found')
                error = ''
                try:
                    error = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='sku-action-info']").find_element(By.CSS_SELECTOR, "span[data-testid='sku-action-error-text']").text
                except:
                    pass
                if error != '':
                    input(error)
                    sys.exit()
                
                try:
                    infoprep = cols[0].find_element(By.CSS_SELECTOR, "kat-link[data-testid='sku-action-info-prep-missing-link']")
                    infoprep.click()
                    print(xlssku, "Prep..")
                    catprep = self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='prep-guidance-prep-category-dropdown']")
                    catprep.click()
                    catprep.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='NONE']").click()
                    explicit_wait()
                    self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[variant='primary']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                    self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='packing-template-save-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()                    
                except:
                    print(xlssku, "Prep is not required..")
                    pass
                explicit_wait()
                print(xlssku, "Input the unit number")
                # input("wait")
                try:
                    # WebDriverWait(cols[0], 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-units-input']")))
                    numunit = cols[0].find_element(By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-units-input']").find_element(By.CSS_SELECTOR, "input[name='numOfUnits']")
                    # numbox = cols[0].find_element(By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-boxes-input']").find_element(By.CSS_SELECTOR, "input[name='numOfBoxes']")
                    if numunit.is_enabled():                     
                        numunit.send_keys(item['total'])
                    else:
                        # if numbox.is_enabled():
                        #     numbox.send_keys(item['total'])
                        # else:
                        #     input(xlssku, "Can Not Input unit or box number, data error")
                        #     sys.exit()

                        print(xlssku, "Unit number disabled")
                        sys.exit()

                    explicit_wait()
                except:
                    input(xlssku + " Can Not Input unit number, data error")
                    sys.exit()
                try:
                    cols[0].find_element(By.CSS_SELECTOR, "kat-button[data-testid='skureadiness-confirm-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                except:
                    pass
                explicit_wait()
                try:                
                    expiry = "{}/{}/{}".format(item['expiry'][5:7], item['expiry'][8:10],item['expiry'][0:4])
                    inputexpiry = cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']").find_element(By.CSS_SELECTOR, "input")
                    if inputexpiry.is_enabled():
                        print(xlssku, "Input the date expired")
                        inputexpiry.send_keys(expiry)
                        inputexpiry.send_keys(Keys.TAB)
                    # cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']").find_element(By.CSS_SELECTOR, "input").send_keys(expiry)
                    # cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']").find_element(By.CSS_SELECTOR, "input").send_keys(Keys.TAB)
                        explicit_wait()
                        cols[0].find_element(By.CSS_SELECTOR, "kat-button[data-testid='skureadiness-confirm-button']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
                except:
                    pass

                explicit_wait()
                error = ''
                try:
                    error = cols[0].find_element(By.CSS_SELECTOR, "kat-label[data-testid='sku-readiness-expiration-date-error']").text
                except:
                    pass
                if error != '':
                    input(error)
                    sys.exit()
            # break
            print(dlist['name'], 'Packaging..')
            self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='step1-prep-fees-and-continue-button']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='step1-continue']").find_element(By.CSS_SELECTOR, "button[class='primary']").click()
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='pack-group-controls']")))
            # input("wait")
            print('Input box count, weight, dimension..')
            if dlist['boxcount'] == 1:
                self.driver.find_element(By.CSS_SELECTOR, "kat-label[text='Everything will fit into one box']").click()
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-input-method-verify-button']").click()
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='pack-group-cli-single-box-webform']")))
                weight = dlist['weightboxes'][0]
                dimension = dlist['dimensionboxes'][0].split("x")
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-width-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dimension[0])
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-height-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dimension[1])
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-length-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dimension[2])
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-single-box-weight-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(weight)
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-single-box-confirm-btn']").click()
                explicit_wait()
                error = ''
                try:
                    error = self.driver.find_element(By.CSS_SELECTOR, "kat-alert[data-testid='pack-mixed-unit-error-results']").text
                except:
                    pass
                if error != '':
                    input(error)
                    sys.exit()

                # try:
                #     WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-alert[data-testid='pack-group-cli-warning-results']")))
                #     driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='confirm-and-continue']").click()
                # except:
                #     pass


            else:
                self.driver.find_element(By.CSS_SELECTOR, "kat-label[text='Multiple boxes will be needed']").click()
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-input-method-verify-button']").click()
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-input[data-testid='cli-multi-box-webform-intial-container-quantity-input']")))
                explicit_wait()
                # input("pause")
                self.driver.find_element(By.CSS_SELECTOR, "kat-input[data-testid='cli-multi-box-webform-intial-container-quantity-input']").find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dlist['boxcount'])
                explicit_wait()
                self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='cli-multi-box-open-webform-btn']").find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                explicit_wait()
                cols = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='sku-quantity-inputs']").find_elements(By.CSS_SELECTOR, "div[class='flo-athens-border-bottom sku-input-child']")
                # print(cols)
                for col in cols:
                    explicit_wait()
                    tsku = col.find_element(By.CSS_SELECTOR, "div[data-testid='sku-information']").find_element(By.CSS_SELECTOR,"span[class='text-primary']").text.strip()
                    for item in dlist['items']:
                        txlssku = item['id'].strip().upper()
                        if tsku == txlssku:
                            cinputs = col.find_element(By.CSS_SELECTOR, "div[class='sku-quantity-wrapper']").find_elements(By.CSS_SELECTOR, "div[class='kat-input-padding-bottom-0 sku-input-katal-box']")
                            for idx, cinput in enumerate(cinputs):
                                cinput.find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(item['boxes'][idx])


                for i in range(0,len(dlist['dimensionboxes'])-1 ):
                    self.driver.find_element(By.CSS_SELECTOR, "div[class='bwd-add-dimension']").find_element(By.CSS_SELECTOR,"kat-link[data-testid='bwd-add-dimension-link']").click()
                cinputs = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='box-dimensions-labels']").find_elements(By.CSS_SELECTOR, "div[data-testid='box-dimensions-label']")
                for idx, cinput in enumerate(cinputs):
                    dimlist = dlist['dimensionboxes'][idx].split("x")
                    xinputs = cinput.find_elements(By.CSS_SELECTOR, "input[type='number']")
                    xinputs[0].send_keys(dimlist[0])
                    xinputs[1].send_keys(dimlist[1])
                    xinputs[2].send_keys(dimlist[2])
                    explicit_wait()

                bwdinput = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='bwd-input']") 
                cinputs = bwdinput.find_element(By.CSS_SELECTOR, "div[data-testid='box-weight-row']").find_elements(By.CSS_SELECTOR, "div[data-testid='weight-input-box']")
                for idx, cinput in enumerate(cinputs):
                    cinput.find_element(By.CSS_SELECTOR, "input[type='number']").send_keys(dlist['weightboxes'][idx])

                cinputs = bwdinput.find_elements(By.CSS_SELECTOR, "div[data-testid='bwd-input-child']")
                for idx, cinput in enumerate(cinputs):
                    xchecks = cinput.find_elements(By.CSS_SELECTOR, "kat-checkbox[data-testid='dimension-checkbox']")
                    xchecks[idx].click()

                self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='modal-confirm-button']").click()

                try:
                    WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-alert[data-testid='pack-group-cli-warning-results']")))
                    self.driver.find_element(By.CSS_SELECTOR, "kat-modal-footer[data-testid='modal-footer']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='modal-confirm-button']").click()
                except:
                    pass

                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[id='skudetails']")))
                explicit_wait()


            self.driver.find_element(By.CSS_SELECTOR, "div[id='stepFooter']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='confirm-and-continue']").click()
            WebDriverWait(self.driver, 60).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='confirm-shipping-body-component']")))

            print("input Send By Date, Shipping mode..")
            todays_date = date.today()
            todays_str = "{}/{}/{}".format(str(todays_date.month), str(todays_date.day), str(todays_date.year))
            self.driver.find_element(By.CSS_SELECTOR, "kat-date-picker[id='sendByDatePicker']").find_element(By.CSS_SELECTOR, "input").clear()
            self.driver.find_element(By.CSS_SELECTOR, "kat-date-picker[id='sendByDatePicker']").find_element(By.CSS_SELECTOR, "input").send_keys(todays_str)
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-date-picker[id='sendByDatePicker']").find_element(By.CSS_SELECTOR, "input").send_keys(Keys.ESCAPE)

            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='original-shipment']")))
            explicit_wait()

            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "h6[data-testid='shipping-mode-title-spd")))

            self.driver.find_element(By.CSS_SELECTOR, "h6[data-testid='shipping-mode-title-spd']").click()
            explicit_wait()
            print(dlist['name'], 'Saving the Shipping data')
            self.driver.find_element(By.CSS_SELECTOR, "div[id='stepFooter']").find_element(By.CSS_SELECTOR, "kat-button[data-testid='confirm-spd-shipping']").click()
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='print-label-dropdown']")))
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='print-label-dropdown']").click()
            explicit_wait()
            print("Downloading PDF File to", self.download_folder)
            self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='print-label-dropdown']").find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='PackageLabel_Letter_2']").click()
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='print-box-labels-button']").click()
            explicit_wait()
            # driver.close()
            self.driver.switch_to.window(original_window)
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button[data-testid='proceed-tracking-details-button']")))
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-button[data-testid='proceed-tracking-details-button']").click()
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='spd-tracking-table']")))
            explicit_wait()
            print(dlist['name'], 'Saving to XLSX file..')
            tracks = self.driver.find_element(By.CSS_SELECTOR, "div[data-testid='spd-tracking-table']").find_elements(By.CSS_SELECTOR,"kat-table-row[class='tracking-id-row']")
            fbalabel = []
            trackid = []
            for track in tracks:
                trs = track.find_elements(By.CSS_SELECTOR, "kat-table-cell")
                fbalabel.append(trs[1].text)
                trackid.append(trs[2].text)
            boxes = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
            for i in range(dlist['begin'], dlist['end']):
                if self.worksheet['B{}'.format(i)].value == 'Shipment ID':
                    for idx, item in enumerate(fbalabel):
                        self.worksheet['{}{}'.format(boxes[idx], i)].value = item

                if self.worksheet['B{}'.format(i)].value == 'Tracking Number':
                    for idx, item in enumerate(trackid):
                        self.worksheet['{}{}'.format(boxes[idx], i)].value = item
            self.workbook.save(self.xlsfile)
            print(dlist['name'], 'Saved to', self.xlsfile)
            print(dlist['name'], 'Extract PDF..')
            self.__extract_pdf(dlist=dlist)
            print('#' * 5, dlist['name'], "End Process", '#' * 5)
            explicit_wait()
            print("Processing next shipping..")
            explicit_wait()
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']")))
            explicit_wait()
            self.driver.find_element(By.CSS_SELECTOR, "kat-link[data-testid='start-new-link']").click()
            
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
            explicit_wait()
        input('All Process Done...')

    def __extract_pdf(self, dlist):
        list_of_files = glob.glob(self.download_folder + self.file_delimeter + "*.pdf" ) # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        foldername = latest_file.replace(".pdf",'')
        isExist = os.path.exists(foldername)
        if not isExist:
            os.makedirs(foldername)
        boxes = dlist['nameboxes']
        
        for idx, tmpbox in enumerate(boxes):
            mfile = fitz.open(latest_file)
            for i in range(idx, len(boxes)-1):
                mfile.delete_page(-1)
            
            for i in range(0, idx):
                mfile.delete_page(0)


            fname = "{}{}{}.pdf".format(foldername, self.file_delimeter,  tmpbox.strip())
            mfile.save(fname)

    def __data_generator(self):
        shipmentlist = []
        for i in range(2, self.worksheet.max_row + 1):
            shipment_row = str(self.worksheet['A{}'.format(i)].value)
            
            if shipment_row.find('Shipment') != -1:
                submitter_row = self.worksheet['B{}'.format(i)].value
                address_row = self.worksheet['B{}'.format(i+1)].value
                if submitter_row != None and address_row != None:
                    shipmentlist.append({'begin':i})
                    if len(shipmentlist) >= 1:
                        shipmentlist[len(shipmentlist)-2]['end'] = i - 1
                else:
                    shipmentlist[len(shipmentlist)-1]['end'] = i - 1
                    break
        # print(shipmentlist)
        for index, shipmentdata in enumerate(shipmentlist):
            shipmentlist[index]['submitter'] = self.worksheet['B{}'.format(shipmentdata['begin'])].value
            shipmentlist[index]['address'] = self.worksheet['B{}'.format(shipmentdata['begin']+1)].value
            shipmentlist[index]['name'] = self.worksheet['A{}'.format(shipmentdata['begin'])].value
            boxes = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
            boxcount = 0
            for box in boxes:
                
                if self.worksheet['{}{}'.format(box, shipmentdata['begin']+1)].value != None:
                    boxcount += 1
                else:
                    break
            if boxcount == 0:
                del shipmentlist[index]
                continue
            shipmentlist[index]['boxcount'] = boxcount
            start = shipmentdata['begin'] + 2
            shipmentlist[index]['weightboxes'] = []
            shipmentlist[index]['dimensionboxes'] = []
            shipmentlist[index]['nameboxes'] = []
            shipmentlist[index]['items'] = []

            # get weightboxes
            rowsearch = 0
            for i in range(start, shipmentdata['end']):
                if self.worksheet['B{}'.format(i)].value == 'Weight':
                    rowsearch = i
                    break
            
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['weightboxes'].append(self.worksheet['{}{}'.format(box, rowsearch)].value)

            # get dimensionboxes
            rowsearch = 0
            for i in range(start, shipmentdata['end']):
                if self.worksheet['B{}'.format(i)].value == 'Dimensions':
                    rowsearch = i
                    break
            
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['dimensionboxes'].append(self.worksheet['{}{}'.format(box, rowsearch)].value)

            #get nameboxes
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                shipmentlist[index]['nameboxes'].append(str(self.worksheet['{}{}'.format(box, shipmentdata['begin'])].value))

            ti = -1
            for i in range(start, shipmentdata['end']):
                ti += 1
                if self.worksheet['A{}'.format(i)].value == None or str(self.worksheet['A{}'.format(i)].value).strip() == '':
                    break
                # shipmentlist[index]['items'].append()
                dict = {
                    'id': self.worksheet['A{}'.format(i)].value,
                    'name': self.worksheet['B{}'.format(i)].value,
                    'total': self.worksheet['C{}'.format(i)].value,
                    'expiry': str(self.worksheet['D{}'.format(i)].value),
                    'boxes':[],

                }
                shipmentlist[index]['items'].append(dict)
                for ke, box in enumerate(boxes):
                    if ke == boxcount:
                        break
                    if self.worksheet['{}{}'.format(box, i)].value == None or str(self.worksheet['{}{}'.format(box, i)].value).strip() == '':
                        shipmentlist[index]['items'][ti]['boxes'].append(0)
                    else:                           
                        shipmentlist[index]['items'][ti]['boxes'].append(self.worksheet['{}{}'.format(box, i)].value)

            # pass
        
        self.datajson =  json.dumps(shipmentlist)
        self.datalist = shipmentlist

    def __data_sanitizer(self):
        clear_screan()
        print('Checking Excel Data..')
        url = 'https://sellercentral.amazon.com/fba/sendtoamazon?ref=fbacentral_nav_fba'
        self.driver.get(url)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-testid='sku-list']")))
        self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").click()
        explicit_wait()
        self.driver.find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='search-dropdown']").find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-value='MSKU']").click()
        explicit_wait()
        for idx, dlist in enumerate(self.datalist):
            print(dlist['name'], "... ", end="")
            error = False
            errorlist = []

            submitter = dlist['submitter'].split("(")[0].strip()
            addresstmp = dlist['address']
            addresslist = addresstmp[addresstmp.find("(")+1:addresstmp.find(")")].strip().split(" ")
            address = addresslist[0] + " " + addresslist[1]# + " " + addresslist[2]

            self.driver.find_element(By.CSS_SELECTOR, "a[data-testid='ship-from-another-address-link']").click()
            ck = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class='selected-address-tile']")))
            selects = self.driver.find_elements(By.CSS_SELECTOR, "div[class='address-tile']")
            address_found = False
            for sel  in selects:
                txt = sel.find_element(By.CSS_SELECTOR, "div[class='tile-address']").text
                # print(txt)
                if txt.find(submitter) != -1 and txt.find(address) != -1:
                    sel.find_element(By.CSS_SELECTOR, "button[class='secondary']").click()
                    # input('pause')
                    address_found = True
                    break
            if not address_found:
                errorlist.append("Address or Submitter not Found")
                error = True
            # else:
            #     print('address found')

            for item in dlist['items']:
                self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").clear()
                xlssku = item['id'].upper()
                self.driver.find_element(By.CSS_SELECTOR,"kat-input[data-testid='search-input']").find_element(By.CSS_SELECTOR, "input").send_keys(xlssku)
                # explicit_wait()
                searchinput = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[data-testid='search-input-link']")))
                searchinput.click()
                explicit_wait()
                cols = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='sku-row-information-details']")
                sku = ''
                try:
                    sku = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='msku']").find_element(By.CSS_SELECTOR, "span").text
                except:
                    error = True
                    errorlist.append(xlssku + ' Not Found')

                if xlssku != sku:
                    errorlist.append(sku + ' Not Match')
                    error = True
                errormsg = ''
                try:
                    errormsg = cols[0].find_element(By.CSS_SELECTOR, "div[data-testid='sku-action-info']").find_element(By.CSS_SELECTOR, "span[data-testid='sku-action-error-text']").text
                except:
                    pass
                if errormsg != '':
                    errorlist.append(errormsg)
                    error = True
                # explicit_wait()
                individual = WebDriverWait(cols[0], 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-dropdown[data-testid='packing-template-dropdown']")))

                # individual = cols[0].find_element(By.CSS_SELECTOR, "kat-dropdown[data-testid='packing-template-dropdown']")
                if individual.text.find('Individual units') == -1:
                    individual.click()
                    explicit_wait()
                    individual.find_element(By.CSS_SELECTOR, "div[class='select-options']").find_element(By.CSS_SELECTOR, "div[class='option-inner-container']").find_element(By.CSS_SELECTOR, "div[data-name='Individual units']").click()

                try:
                    numunit = cols[0].find_element(By.CSS_SELECTOR, "kat-input[data-testid='sku-readiness-number-of-units-input']").find_element(By.CSS_SELECTOR, "input[name='numOfUnits']")
                    if not numunit.is_enabled():
                        errorlist.append(sku + " unit number disabled")
                        error = True                     
                except:
                    errorlist.append(sku + " unit number disabled")
                    error = True

                expiry = "{}/{}/{}".format(item['expiry'][5:7], item['expiry'][8:10],item['expiry'][0:4])
                try:
                    inputexpiry = cols[0].find_element(By.CSS_SELECTOR, "kat-date-picker[id='expirationDatePicker']").find_element(By.CSS_SELECTOR, "input")
                    if inputexpiry.is_enabled():
                        inputexpiry.send_keys(expiry)
                        inputexpiry.send_keys(Keys.TAB)
                except:
                    pass

                errormsg = ''
                try:
                    errormsg = cols[0].find_element(By.CSS_SELECTOR, "kat-label[data-testid='sku-readiness-expiration-date-error']").text
                except:
                    pass
                if errormsg != '':
                    errorlist.append(errormsg)
                    error = True
                
                if error:
                    break

            if error:
                print("Deleted")
                print(errorlist)
                del self.datalist[idx]
            else:
                print("Passed")
            
    @property
    def workbook(self):
        return self.__workbook
   
    @property
    def worksheet(self):
        return self.__worksheet

    @property
    def datalist(self):
        return self.__datalist

    @datalist.setter
    def datalist(self, value):
        self.__datalist = value

    @property
    def datajson(self):
        return self.__datajson
    
    @datajson.setter    
    def datajson(self, value):
        self.__datajson = value

    @property
    def chrome_data(self):
        return self.__chrome_data
    
    @chrome_data.setter    
    def chrome_data(self, value):
        self.__chrome_data = value

    @property
    def download_folder(self):
        return self.__download_folder
    
    @download_folder.setter    
    def download_folder(self, value):
        self.__download_folder = value

    @property
    def xlsfile(self):
        return self.__xlsfile
    
    @xlsfile.setter    
    def xlsfile(self, value):
        self.__xlsfile = value

    @property
    def file_delimeter(self):
        return self.__delimeter

    @property
    def driver(self):
        return self.__driver

def main():
    parser = argparse.ArgumentParser(description="Amazon Shipment")
    parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
    parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    parser.add_argument('-output', '--pdfoutput', type=str,help="PDF output folder")
    parser.add_argument('-cdata', '--chromedata', type=str,help="Chrome User Data Directory")
    args = parser.parse_args()
    if args.xlsinput[-5:] != '.xlsx':
        input('2nd File input have to XLSX file')
        sys.exit()
    isExist = os.path.exists(args.xlsinput)
    if not isExist:
        input(args.xlsinput + " does not exist")
        sys.exit()
    isExist = os.path.exists(args.chromedata)
    if isExist == False :
        input('Please check Chrome User Data Directory')
        sys.exit()
    isExist = os.path.exists(args.pdfoutput)
    if not isExist:
        input(args.pdfoutput + " folder does not exist")
        sys.exit()

    # input(args)
    # exit()
    shipment = AmazonShipment(xlsfile=args.xlsinput, sname=args.sheetname, chrome_data=args.chromedata, download_folder=args.pdfoutput)
    print(shipment.datalist)
    # shipment.parse()
    # shipment.extract_pdf(shipment.datalist[0])
    # input('wait')


if __name__ == '__main__':
    main()
