from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
import sys
import os
import argparse
import time
from openpyxl import Workbook, load_workbook
from sys import platform
import json
from datetime import date, datetime, timedelta
import warnings
import logging
from pathlib import Path
import amazon_lib as lib

report = logging.getLogger()
report.setLevel(logging.NOTSET)
logger = logging.getLogger()
logger.setLevel(logging.NOTSET)

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

class AmazonReview:
    def __init__(self, xlsfile, sname, profilename) -> None:
        try:
            self.__workbook = load_workbook(filename=xlsfile, read_only=False, data_only=True)
            self.__worksheet = self.__workbook[sname]
        except Exception as e:
            logger.error(e)
            input("XLSX file or Sheet name not found")
            sys.exit()
        self.__datajson = json.loads("{}")
        self.__datalist = []
        self.__profilename = profilename
        self.__xlsfile = xlsfile
        self.__delimeter = "/" 
        if platform == "win32":
            self.__delimeter = "\\"
        lib.clear_screan()
        self.__driver = self.__browser_init()
        self.__data_generator()

    def __browser_init(self):
        warnings.filterwarnings("ignore", category=UserWarning)
        options = webdriver.ChromeOptions()
        # options.add_argument("--headless")
        options.add_argument("user-data-dir={}".format(getProfiles()[self.profilename]['chrome_user_data']))
        options.add_argument("profile-directory={}".format(getProfiles()[self.profilename]['chrome_profile']))
        options.add_argument('--no-sandbox')
        options.add_argument("--log-level=3")
        # options.add_argument("--window-size=1200, 900")
        options.add_argument('--start-maximized')
        options.add_argument("--disable-notifications")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        # return webdriver.Chrome(service=Service(CM(version="114.0.5735.90").install()), options=options)
        return webdriver.Chrome(service=Service(executable_path=os.path.join(os.getcwd(), "chromedriver", "chromedriver.exe")), options=options)

    def __data_generator(self):
        print("Data Mounting... ", end="")
        itemlist = []
        for i in range(2, self.worksheet.max_row + 1):
            sku_row = str(self.worksheet['A{}'.format(i)].value)
            if sku_row == 'None':
                break
            else:
                mydict = {
                    "sku": sku_row,
                    "asin": str(self.worksheet['B{}'.format(i)].value),
                    "name": str(self.worksheet['C{}'.format(i)].value),
                }
                itemlist.append(mydict)
        self.datalist = itemlist
        print("Passed")

    def parse(self):
        dateformat = "%m/%d/%Y"
        for dlist in self.datalist:
            url = "https://sellercentral.amazon.ca/orders-v3/search?shipByDate=all&sort=order_date_asc&date-range=last-30&page=1&q={}&qt=asin".format(dlist['asin'])
            # report.critical("#" * 50)
            print("ASIN: ", dlist['asin'])
            report.critical("ASIN: {}".format(dlist['asin']))
            self.driver.get(url)
            original_window = self.driver.current_window_handle
            while True:
                try:
                    WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "table#orders-table")))
                    time.sleep(3)
                except:
                    break
                
                rows = self.driver.find_element(By.CSS_SELECTOR, "table#orders-table").find_element(By.CSS_SELECTOR, "tbody").find_elements(By.CSS_SELECTOR, "tr")
                for row in rows:
                    try:
                        today = date.today()
                        dorderstr = row.find_elements(By.CSS_SELECTOR, "td")[1].find_element(By.CSS_SELECTOR, "div.cell-body").find_elements(By.CSS_SELECTOR, "div")[2].text
               
                        dorder = datetime.strptime(dorderstr, dateformat)
                        dorder = dorder.date()
                        dorder10 = dorder + timedelta(days=9)
                        dorder25 = dorder + timedelta(days=24)
                        orderid = row.find_elements(By.CSS_SELECTOR, "td")[2].find_element(By.CSS_SELECTOR, "div.cell-body-title a").text
                        # print(dorder,dorder10, today)
                        if today >= dorder10 and today <= dorder25:
                            link = "https://sellercentral.amazon.ca/messaging/reviews?orderId={}&marketplaceId=ATVPDKIKX0DER".format(orderid)
                            self.driver.execute_script("window.open('%s', '_blank')" % link)
                            self.driver.switch_to.window(self.driver.window_handles[-1])
                            time.sleep(2)
                            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div#ayb-reviews")))
                            try:
                                self.driver.find_element(By.CSS_SELECTOR, "div#ayb-reviews h4.ayb-reviews-description")
                                self.driver.find_element(By.CSS_SELECTOR, "div#ayb-reviews kat-button[label='Yes']").click()
                                print(orderid, "Request Review sent")
                                report.critical(orderid + ": Request Review sent")
                                report.critical("Url: " + link)
                            except:
                                try:
                                    message = self.driver.find_element(By.CSS_SELECTOR, "div#ayb-reviews div.ayb-request-review-error-description").text
                                    print(orderid, "Request Review Disable")
                                    report.critical(orderid + ": Request Review Disable")
                                    report.critical("Message: " + message)
                                    report.critical("Url: " + link)
                                except:
                                    print(orderid, "Not Detected")
                                    report.critical(orderid + ": Not Detected")
                                    report.critical("Url: " + link)
                            time.sleep(2)
                            for handle in self.driver.window_handles:
                                if handle != original_window:
                                    self.driver.switch_to.window(handle)
                                    self.driver.close()
                            self.driver.switch_to.window(original_window)
                    except:
                        continue
                try:
                    next = self.driver.find_element(By.CSS_SELECTOR, "li[class='a-last']")
                    next.click()
                except:
                    break
            report.critical("-" * 100)    

    @property
    def workbook(self):
        return self.__workbook
   
    @property
    def worksheet(self):
        return self.__worksheet

    @property
    def datalist(self):
        return self.__datalist

    @datalist.setter
    def datalist(self, value):
        self.__datalist = value

    @property
    def datajson(self):
        return json.dumps(self.datalist)  
    @property
    def driver(self):
        return self.__driver

    @property
    def profilename(self):
        return self.__profilename

    @profilename.setter
    def profilename(self, value):
        self.__profilename = value

def main():
    parser = argparse.ArgumentParser(description="Amazon Shipment")
    parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
    parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    parser.add_argument('-profile', '--profile', type=str,help="Chrome Profile name")

    args = parser.parse_args()
    if args.xlsinput[-5:] != '.xlsx':
        input('2nd File input have to XLSX file')
        sys.exit()
    isExist = os.path.exists(args.xlsinput)
    if not isExist:
        input(args.xlsinput + " does not exist")
        sys.exit()

    filename = "amazon-review-request-{}-{}-{}-{}.log"
    reportfilename = 'logs/'+ filename.format("report", Path(args.xlsinput).stem, args.sheetname, date.today())
    if os.path.exists(reportfilename):
        os.remove(reportfilename)
   
    file_handler = logging.FileHandler('logs/amazonreview-err.log')
    file_handler.setLevel(logging.ERROR)
    file_handler_format = '%(asctime)s | %(levelname)s | %(lineno)d: %(message)s'
    file_handler.setFormatter(logging.Formatter(file_handler_format))
    logger.addHandler(file_handler)

    file_handler = logging.FileHandler(reportfilename)
    file_handler.setLevel(logging.CRITICAL)
    # file_handler_format = '%(asctime)s | %(levelname)s: %(message)s'
    file_handler_format = '%(message)s'
    file_handler.setFormatter(logging.Formatter(file_handler_format))
    report.addHandler(file_handler)

    report.critical("###### START ######")
    report.critical("Filename: {}\nSheet Name:{}".format(args.xlsinput, args.sheetname))
    report.critical("\n")
    review = AmazonReview(xlsfile=args.xlsinput, sname=args.sheetname, profilename=args.profile)
    review.parse()
    report.critical("\n")
    report.critical("###### END ######")
if __name__ == '__main__':
    main()
