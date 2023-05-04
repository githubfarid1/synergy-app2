import argparse
import sys, os, time
from openpyxl import Workbook, load_workbook
from urllib.parse import urlparse
from bs4 import BeautifulSoup


def main():
    parser = argparse.ArgumentParser(description="Amazon Shipment Check")
    parser.add_argument('-xls', '--xlsinput', type=str,help="XLSX File Input")
    parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    args = parser.parse_args()
    if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
        input('input the right XLSX or XLSM file')
        sys.exit()
    isExist = os.path.exists(args.xlsinput)
    if not isExist:
        input(args.xlsinput + " does not exist")
        sys.exit()
    workbook = load_workbook(filename=args.xlsinput, read_only=False, keep_vba=True, data_only=True)
    worksheet = workbook[args.sheetname]
    cw = os.getcwd()
    path = '/walmart'
    ourPath = cw + os.path.join(path)
    os.chdir(ourPath)
        

    for i in range(2, worksheet.max_row + 1):
        
        url = worksheet[f'A{i}'].value
        domain = urlparse(url).netloc
        if domain == 'www.walmart.ca':
            os.system("scrapy crawl superstore -a url={}".format(url))

if __name__ == '__main__':
    main()