import fitz
import os
from openpyxl import Workbook, load_workbook
import sys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
# from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from PyPDF2 import PdfMerger, PdfFileReader, PdfFileWriter
from sys import platform
from datetime import date, datetime, timedelta
import time
import glob
import shutil
import easyocr
from pathlib import Path
from pdf2image import convert_from_path
import numpy
import pandas as pd
import json
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def file_delimeter():
    dm = "/" 
    if platform == "win32":
        dm = "\\"
    return dm

def extract_pdf(box, shipment_id, label, download_folder):
    pdffile = "{}{}package-{}.pdf".format(download_folder, file_delimeter(), shipment_id)
    foldername = "{}{}combined".format(download_folder, file_delimeter()) 
    isExist = os.path.exists(foldername)
    if not isExist:
        os.makedirs(foldername)
    white = fitz.utils.getColor("white")
    mfile = fitz.open(pdffile)
    fname = "{}{}{}.pdf".format(foldername, file_delimeter(),  box.strip())
    tmpname = "{}{}{}.pdf".format(foldername, file_delimeter(), "tmp")

    found = False
    pfound = 0
    for i in range(0, mfile.page_count):
        page = mfile[i]
        plist = page.search_for(label)
        if len(plist) != 0:
            found = True
            pfound = i
            break
    if found:
        single = fitz.open()
        single.insert_pdf(mfile, from_page=pfound, to_page=pfound)
        mfile.close()
        single.save(tmpname)
        mfile = fitz.open(tmpname)
        page = mfile[0]
        page.insert_text((550.2469787597656, 100.38037109375), "Box:{}".format(str(box)), rotate=90, color=white)
        page.set_rotation(90)
        mfile.save(fname)


def scan_web(download_folder, xlsfile, sname):
    workbook = load_workbook(filename=xlsfile, read_only=False, data_only=True)
    # print(xlsfile, sname)
    worksheet = workbook[sname]

    config = getConfig()
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir={}".format(config['chrome_user_data'])) 
    options.add_argument("profile-directory={}".format(config['chrome_profile']))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    # options.add_argument("--window-size=1200, 900")
    options.add_argument('--start-maximized')
    options.add_argument("--disable-notifications")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    profile = {"plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
                "download.default_directory": download_folder, # disable karena kadang gak jalan di PC lain. Jadi downloadnya tetap ke folder download default
                "download.extensions_to_open": "applications/pdf",
                "download.prompt_for_download": False,
                'profile.default_content_setting_values.automatic_downloads': 1,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome                    
                }
    options.add_experimental_option("prefs", profile)
    driver = webdriver.Chrome(service=Service(CM().install()), options=options)
    # url = "https://sellercentral.amazon.ca/gp/ssof/shipping-queue.html/ref=xx_fbashipq_dnav_xx#fbashipment"
    url = "https://sellercentral.amazon.ca/fba/sendtoamazon?wf=wf8db48d52-7b85-4c4c-8268-d42f5f642730"
    driver.get(url)
    input("pause")
    tabs = driver.find_elements(By.CSS_SELECTOR, "div[data-testid='shipment-tracking-tab']")
    # tabcount = 0
    shiplist = []
    for tab in tabs:
        # tabcount += 1
        shipment_id = tab.find_elements(By.CSS_SELECTOR, "div")[3].text.replace("Shipment ID:","").strip()
        tab.click()
        time.sleep(2)
        tracks = driver.find_element(By.CSS_SELECTOR, "div[data-testid='spd-tracking-table']").find_elements(By.CSS_SELECTOR,"kat-table-row[class='tracking-id-row']")
        dtmp = []
        for track in tracks:
            trs = track.find_elements(By.CSS_SELECTOR, "kat-table-cell")
            dict = {
                'shipmentid': shipment_id,
                'label':trs[1].text,
                'trackid': trs[2].text,
                'weight': trs[4].text,
                'dimension': trs[5].text,

            }
            dtmp.append(dict)
        shiplist.append(dtmp)
    input(shiplist)
    boxcols = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
    begin = 164
    end = 186
    shipname = "Shipment 8"
    
    stmp = []
    for boxcol in boxcols:
        dimension = ""
        weight = ""
        box = str(worksheet['{}{}'.format(boxcol, begin)].value)
        
        # for ship in shiplist:
        #     for s in ship:
        #         # print(s['label'], s['trackid'])
        #         if box != 'None':
        #             # print(stmp)
        #             dimrow = 0
        #             for i in range(begin, end):
        #                 if worksheet['B{}'.format(i)].value == 'Weight':
        #                     weight = worksheet['{}{}'.format(boxcol, i)].value
                        
        #                 if worksheet['B{}'.format(i)].value == 'Dimensions':
        #                     dimension = worksheet['{}{}'.format(boxcol, i)].value
        #                     dimrow = i
        #                 dimension = dimension.replace(" ","")
        #                 dimship = s['dimension'].replace(" ","")
        #             # print(s['weight'], weight, dimension, dimship)
        #             if int(s['weight']) == int(weight) and dimension == dimship:
        #                 print("boxed:",boxcol, s['trackid'])
        #                 if not s['trackid'] in stmp and str(worksheet['{}{}'.format(boxcol, dimrow+2)].value) == 'None':
        #                     print("boxx:",boxcol, s['trackid'])
        #                     stmp.append(s['trackid'])
        #                     # extract_pdf(box=box, shipment_id=s['shipmentid'], label=s['label'], download_folder=download_folder)
        #                     # print(boxcol, dimrow+1, s['label'])
        #                     # print(boxcol, dimrow+2, s['trackid'])

        #                     worksheet['{}{}'.format(boxcol, dimrow+1)].value = s['label']
        #                     worksheet['{}{}'.format(boxcol, dimrow+2)].value = s['trackid']



    stmp = []
    for ship in shiplist:
        for s in ship:
            print(s['label'], s['trackid'])
            for boxcol in boxcols:
                dimension = ""
                weight = ""
                box = str(worksheet['{}{}'.format(boxcol, begin)].value)
                # print(box)
                if box != 'None':
                    dimrow = 0
                    for i in range(begin, end):
                        if worksheet['B{}'.format(i)].value == 'Weight':
                            weight = worksheet['{}{}'.format(boxcol, i)].value
                        
                        if worksheet['B{}'.format(i)].value == 'Dimensions':
                            dimension = worksheet['{}{}'.format(boxcol, i)].value
                            dimrow = i
                        dimension = dimension.replace(" ","")
                        dimship = s['dimension'].replace(" ","")

                    if int(s['weight']) == int(weight) and dimension == dimship:
                        if not s['trackid'] in stmp and str(worksheet['{}{}'.format(boxcol, dimrow+2)].value) == 'None':
                            stmp.append(s['trackid'])
                            extract_pdf(box=box, shipment_id=s['shipmentid'], label=s['label'], download_folder=download_folder)
                            worksheet['{}{}'.format(boxcol, dimrow+1)].value = s['label']
                            worksheet['{}{}'.format(boxcol, dimrow+2)].value = s['trackid']

    workbook.save(xlsfile)
    print(shipname, 'Saved to', xlsfile)
    print(shipname, 'Extract PDF..')


def join_pdfs(pdfoutput_folder):
    merger = PdfMerger()
    print("Merging PDF files in one PDF File..", end=" ", flush=True)
    time.sleep(1)
    file_delimeter = "/" 
    if platform == "win32":
        file_delimeter = "\\"
    time.sleep(1)
    now = datetime.now()
    fname = "{} {}.pdf".format(now.strftime("%B %d"), "Labels")
    pdfoutput_folder_combined = pdfoutput_folder + file_delimeter + "combined"
    tmpname = "{}{}{}.pdf".format(pdfoutput_folder_combined, file_delimeter, "tmp")
    isExist = os.path.exists(tmpname)
    if isExist:
        os.remove(tmpname)
    resultfile = pdfoutput_folder + file_delimeter + fname
    pdffiles = glob.glob(pdfoutput_folder_combined + file_delimeter + "*.pdf")
    if len(pdffiles) != 0:
        dictfiles = {}
        for pdffile in pdffiles:
            try:
                basefilename = os.path.basename(pdffile)
                dictfiles[int(basefilename.replace(".pdf",""))] = pdffile
            except:
                continue
        sortedfiles = dict(sorted(dictfiles.items()))

        for file in sortedfiles:
            merger.append(sortedfiles[file])
        merger.write(resultfile)
        print("Finished")
        return resultfile
    else:
        print("No pdf files was found")
        return ""

def add_page_numbers(pdffile):
    print("Add page numbering...", end=" ", flush=True)
    time.sleep(1)
    tmpfile = "__tmp.pdf" 
    shutil.copy(pdffile, tmpfile)
    doc = fitz.open(tmpfile)
    for i in range(0, doc.page_count):
        page = doc[i]
        page.insert_text((590.2469787597656, 400.38037109375), "{}".format(str(i+1)), rotate=90, fontsize=9)
    doc.save(pdffile)
    doc.close()
    os.remove(tmpfile)
    print("Finished")

    
def generate_xls_from_pdf(fileinput, addressfile):
    print("Generate new XLS file from PDF File...", end=" ", flush=True)
    addressdict = pd.read_csv(addressfile, usecols=['Consignee', 'Address']).to_dict('index')
    pdfFileObject = open(fileinput, 'rb')
    pdfReader = PdfFileReader(pdfFileObject)
    reader = easyocr.Reader(['en'], gpu=False, verbose=False)
    # print(" No. Of Pages :", pdfReader.numPages)
    filepath = fileinput[:-4] + ".xlsx"
    wb = Workbook()
    ws = wb.active

    ws['A1'].value = 'Box'
    ws['B1'].value = 'Tracking ID'
    ws['C1'].value = 'Weight'
    ws['D1'].value = 'Consignee'
    ws['E1'].value = 'Address'
    ws['F1'].value = 'Distributor'
    ws['G1'].value = 'Consignee'
    ws['H1'].value = 'Consignee Check'

    for i in range(0, pdfReader.numPages):
        if os.path.exists(Path("pdftmp.pdf")):
            os.remove(Path("pdftmp.pdf"))
        pdfWriter = PdfFileWriter()
        pdf = pdfReader.getPage(i)
        pdfWriter.addPage(pdf)
        with Path("pdftmp.pdf").open(mode="wb") as output_file:
            pdfWriter.write(output_file)
        images = convert_from_path(Path('pdftmp.pdf'))
        imgcrop = images[0].crop(box = (180,750,750,900))
        res = reader.readtext(numpy.array(imgcrop)  , detail = 0)
        tracking_id = res[1].strip().replace('TRACKING #','').replace(' ','').replace(":","")
        # print(tracking_id)
        lines = pdf.extract_text(space_width=200).split("\n")
        if platform == "win32":
            weight = lines[1].split('-')[1].strip().split("lb")[0]
        else:
            weight = lines[0].split('-')[1].strip().replace("lb","")    

        to = lines[7]
        address = lines[8]
        shipper = lines[2]
        code = lines[12]
        box = lines[-2].replace(":", " ")
        consignee = ""
        for idx, addict in addressdict.items():
            if str(addict['Address'])[0:21].lower() == address[0:21].lower():
                consignee = addict['Consignee']
                break

        ws['A{}'.format(i + 2)].value = box
        ws['B{}'.format(i + 2)].value = tracking_id
        ws['C{}'.format(i + 2)].value = weight
        ws['D{}'.format(i + 2)].value = to
        ws['E{}'.format(i + 2)].value = address
        ws['F{}'.format(i + 2)].value = shipper
        ws['G{}'.format(i + 2)].value = code
        ws['H{}'.format(i + 2)].value = consignee
    wb.save(filepath)
    print("Finished")



if __name__ == "__main__":
    download_folder = "C:\\App\\data-tester\\March 06x"
    xlsfile = "C:\\App\\data-tester\\Shipment transfer.xlsx"
    sname = "March 06"
    # scan_web(download_folder=download_folder, xlsfile=xlsfile, sname=sname)
    # pdfoutput = "/home/farid/dev/python/synergy-github/data/sample/Feb 21"
    addressfile = Path("address.csv")
    resultfile = join_pdfs(pdfoutput_folder=download_folder)
    if resultfile != "":
        add_page_numbers(resultfile)
        generate_xls_from_pdf(resultfile, addressfile)
    input("End Process..")    

