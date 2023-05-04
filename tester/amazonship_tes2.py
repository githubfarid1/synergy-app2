from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
# from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
import sys
import fitz
import os
import argparse
import time
from openpyxl import Workbook, load_workbook
# import unicodedata as ud
from sys import platform
import json
from random import randint
from datetime import date, datetime, timedelta
import warnings
import logging
from pathlib import Path
import xlwings as xw
import shutil
dfolder = r"C:/Users/User/OneDrive/01 - Shipment Creation/April 27th Shipment Labels/res"
xltmp = r"C:/Users/User/OneDrive/01 - Shipment Creation/April 27th Shipment Labels/xlstmp.xlsm"
xltmp2 = r"C:/Users/User/OneDrive/01 - Shipment Creation/April 27th Shipment Labels/xlstmp2.xlsm"

sname = "Shipment summary"
workbook = load_workbook(filename=xltmp, read_only=False, keep_vba=True, data_only=True)
worksheet = workbook[sname]
xlbook = xw.Book(xltmp2)
xlsheet = xlbook.sheets[sname]


def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def browser_init(download_folder):
    config = getConfig()
    warnings.filterwarnings("ignore", category=UserWarning)
    options = webdriver.ChromeOptions()
    # options = Options()
    # options.add_argument("--headless")
    options.add_argument("user-data-dir={}".format(config['chrome_user_data'])) 
    options.add_argument("profile-directory={}".format(config['chrome_profile']))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    # options.add_argument("--window-size=1200, 900")
    options.add_argument('--start-maximized')
    options.add_argument("--disable-notifications")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    profile = {"plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
                "download.default_directory": download_folder, # disable karena kadang gak jalan di PC lain. Jadi downloadnya tetap ke folder download default
                "download.extensions_to_open": "applications/pdf",
                "download.prompt_for_download": False,
                'profile.default_content_setting_values.automatic_downloads': 1,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome                    
                }
    options.add_experimental_option("prefs", profile)
    return webdriver.Chrome(service=Service(CM().install()), options=options)



def data_generator():
    print("Data Mounting... ", end="")
    shipmentlist = []
    for i in range(2, worksheet.max_row + 1):
        shipment_row = str(worksheet['A{}'.format(i)].value)
        if shipment_row.find('Shipment') != -1:
            # print(shipment_row, i)
            startrow = i
            y = i
            shipment_empty = True
            while True:
                y += 1

                # skip if shipment_id was filled
                if ''.join(str(worksheet['B{}'.format(y)].value)).strip() == 'Shipment ID':
                    if ''.join(str(worksheet['E{}'.format(y)].value)).strip() != 'None':
                        shipment_empty = False

                if str(worksheet['B{}'.format(y)].value) == 'Tracking Number':
                    endrow = y + 1
                    i = y + 1
                    break
            if shipment_empty == True:
                shipmentlist.append({'begin':startrow, 'end':endrow})
            else:
                print(shipment_row + " Skipped")

    # print(json.dumps(shipmentlist))
    for index, shipmentdata in enumerate(shipmentlist):
        shipmentlist[index]['submitter'] = worksheet['B{}'.format(shipmentdata['begin'])].value
        shipmentlist[index]['address'] = worksheet['B{}'.format(shipmentdata['begin']+1)].value
        shipmentlist[index]['name'] = worksheet['A{}'.format(shipmentdata['begin'])].value
        boxes = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
        boxcount = 0
        for box in boxes:
            
            if worksheet['{}{}'.format(box, shipmentdata['begin'])].value != None:
                boxcount += 1
            else:
                break
        if boxcount == 0:
            del shipmentlist[index]
            continue
        shipmentlist[index]['boxcount'] = boxcount
        start = shipmentdata['begin'] + 2
        shipmentlist[index]['weightboxes'] = []
        shipmentlist[index]['dimensionboxes'] = []
        shipmentlist[index]['nameboxes'] = []
        shipmentlist[index]['items'] = []

        # get weightboxes
        rowsearch = 0
        for i in range(start, shipmentdata['end']):
            if worksheet['B{}'.format(i)].value == 'Weight':
                rowsearch = i
                break
        
        for ke, box in enumerate(boxes):
            if ke == boxcount:
                break
            shipmentlist[index]['weightboxes'].append(worksheet['{}{}'.format(box, rowsearch)].value)

        # get dimensionboxes
        rowsearch = 0
        for i in range(start, shipmentdata['end']):
            if worksheet['B{}'.format(i)].value == 'Dimensions':
                rowsearch = i
                break
        
        for ke, box in enumerate(boxes):
            if ke == boxcount:
                break
            shipmentlist[index]['dimensionboxes'].append(worksheet['{}{}'.format(box, rowsearch)].value)

        #get nameboxes
        for ke, box in enumerate(boxes):
            if ke == boxcount:
                break
            shipmentlist[index]['nameboxes'].append(str(worksheet['{}{}'.format(box, shipmentdata['begin'])].value))

        ti = -1
        for i in range(start, shipmentdata['end']):
            ti += 1
            if worksheet['A{}'.format(i)].value == None or str(worksheet['A{}'.format(i)].value).strip() == '':
                break
            # shipmentlist[index]['items'].append()
            
            dict = {
                'id': worksheet['A{}'.format(i)].value,
                'name': worksheet['B{}'.format(i)].value,
                'total': worksheet['C{}'.format(i)].value,
                'expiry': str(worksheet['D{}'.format(i)].value),
                'boxes':[],

            }

            shipmentlist[index]['items'].append(dict)
            for ke, box in enumerate(boxes):
                if ke == boxcount:
                    break
                if worksheet['{}{}'.format(box, i)].value == None or str(worksheet['{}{}'.format(box, i)].value).strip() == '':
                    shipmentlist[index]['items'][ti]['boxes'].append(0)
                else:                           
                    shipmentlist[index]['items'][ti]['boxes'].append(worksheet['{}{}'.format(box, i)].value)

    
    #cleansing
    idxdel = []
    for index, shipmentdata in enumerate(shipmentlist):
        try:
            cheat = shipmentdata['name']
        except:
            idxdel.append(index)
    
    for idx in idxdel:
        for index, shipmentdata in enumerate(shipmentlist):
            try:
                cheat = shipmentdata['name']
            except:
                del shipmentlist[index]
            
        # pass
    
    return shipmentlist
    # print("Passed")

datalist = data_generator()
# url = "https://sellercentral.amazon.ca/fba/sendtoamazon/enter_tracking_details_step?wf=wf7a0b0552-5b73-4916-8dd9-6d444822a00c"
url = "https://sellercentral.amazon.ca/fba/sendtoamazon/enter_tracking_details_step?wf=wf7f6f7d0f-ac84-4cf7-abe1-51da678f647f"
dlist = datalist[1]
print(dlist)
driver = browser_init(dfolder)
driver.get(url)
input("pause")
tabs = driver.find_elements(By.CSS_SELECTOR, "div[data-testid='shipment-tracking-tab']")
# tabcount = 0
# print(tabs)
shiplist = []
for tab in tabs:
    # tabcount += 1
    shipment_id = tab.find_elements(By.CSS_SELECTOR, "div")[3].text.replace("Shipment ID:","").strip()
    tab.click()
    time.sleep(1)
    tracks = driver.find_element(By.CSS_SELECTOR, "div[data-testid='spd-tracking-table']").find_elements(By.CSS_SELECTOR,"kat-table-row[class='tracking-id-row']")
    dtmp = []
    for track in tracks:
        trs = track.find_elements(By.CSS_SELECTOR, "kat-table-cell")
        dict = {
            'shipmentid': shipment_id,
            'label':trs[1].text,
            'trackid': trs[2].text,
            'weight': trs[4].text,
            'dimension': trs[5].text,

        }
        dtmp.append(dict)
    shiplist.append(dtmp)

boxcols = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
stmp = []
for ship in shiplist:
    for s in ship:
        for boxcol in boxcols:
            print(boxcol)
            dimension = ""
            weight = ""
            box = str(xlsheet['{}{}'.format(boxcol, dlist['begin'])].value)
            # print(box)
            if box != 'None':
                dimrow = 0
                for i in range(dlist['begin'], dlist['end']):
                    if xlsheet['B{}'.format(i)].value == 'Weight':
                        weight = xlsheet['{}{}'.format(boxcol, i)].value
                    
                    if xlsheet['B{}'.format(i)].value == 'Dimensions':
                        dimension = xlsheet['{}{}'.format(boxcol, i)].value
                        dimrow = i
                    dimension = dimension.replace(" ","")
                    dimship = s['dimension'].replace(" ","")

                if int(s['weight']) == int(weight) and dimension == dimship:
                    if not s['trackid'] in stmp and str(xlsheet['{}{}'.format(boxcol, dimrow+2)].value) == 'None':
                        stmp.append(s['trackid'])
                        # __extract_pdf(box=box, shipment_id=s['shipmentid'], label=s['label'])
                        # worksheet['{}{}'.format(boxcol, dimrow+1)].value = s['label']
                        # worksheet['{}{}'.format(boxcol, dimrow+2)].value = s['trackid']
                        # restup = (f"{boxcol}{dimrow+1}", s['label'], f"{boxcol}{dimrow+2}", s['trackid'])
                        # reslist.append(restup)
                        xlsheet[f"{boxcol}{dimrow+1}"].value = s['label']
                        xlsheet[f"{boxcol}{dimrow+2}"].value = s['trackid']
                        input(s['label'] + " " + s['trackid'])
