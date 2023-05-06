from tkinter import *
from tkinter import ttk
from functools import partial
from tkinter import filedialog
from tkinter import messagebox
import tkinter
from tkcalendar import Calendar, DateEntry
from pathlib import Path
import os
import sys
from sys import platform
from subprocess import Popen
import openpyxl
import git
import warnings

warnings.filterwarnings("ignore", category=UserWarning)
if platform == "linux" or platform == "linux2":
    pass
elif platform == "win32":
	from subprocess import CREATE_NEW_CONSOLE
import json

VERSION = "1.14-new"
def getConfig():
	file = open("setting.json", "r")
	config = json.load(file)
	return config

def getProfiles():
	file = open("chrome_profiles.json", "r")
	config = json.load(file)
	return config

def run_module(comlist):
	if platform == "linux" or platform == "linux2":
		comlist[:0] = ["--"]
		comlist[:0] = ["gnome-terminal"]
		# print(comlist)
		Popen(comlist)
	elif platform == "win32":
		Popen(comlist, creationflags=CREATE_NEW_CONSOLE)
	
	comall = ''
	for com in comlist:
		comall += com + " "
	print(comall)

def main():
	window = Window()
	window.mainloop()
	
class Window(Tk):
	def __init__(self) -> None:
		super().__init__()
		self.title('Synergy Script ' + VERSION)
		self.resizable(0, 0)
		self.grid_propagate(False)
		width = 700
		height = 650
		swidth = self.winfo_screenwidth()
		sheight = self.winfo_screenheight()
		newx = int((swidth/2) - (width/2))
		newy = int((sheight/2) - (height/2))
		self.geometry(f"{width}x{height}+{newx}+{newy}")
		self.columnconfigure(0, weight=1)
		# self.columnconfigure(1, weight=1)

		self.rowconfigure(0, weight=1)
		# self.rowconfigure(1, weight=1)
		# self.rowconfigure(2, weight=1)
		
		exitButton = ttk.Button(self, text="Exit", command=lambda:self.destroy())
		pullButton = ttk.Button(self, text='Update Script', command=lambda:self.gitPull())
		settingButton = ttk.Button(self, text='Chrome Profiles', command=lambda:self.chromeProfile())
		
		exitButton.grid(row=2, column=0, sticky=(E), padx=20, pady=5)
		# pullButton.grid(row=2, column=0, sticky=(E, N, S), padx=20, pady=5)
		pullButton.grid(row = 2, column = 0, sticky = (W), padx=20, pady=10)
		settingButton.grid(row = 2, column = 0, sticky = (N, S), padx=20, pady=10)


		config = getConfig()
		mainFrame = MainFrame(self)
		mainFrame.grid(column=0, row=0, sticky=(N, E, W, S))
		# settingFrame = SettingFrame(self)
		# settingFrame.grid(column=0, row=1, sticky=(N, E, W, S))
	def gitPull(self):
		git_dir = os.getcwd() 
		g = git.cmd.Git(git_dir)
		g.pull()		
		messagebox.showinfo(title='Info', message='the scripts has updated..')

	def chromeProfile(self):
		settingFrame = ChromeProfilesFrame(self)
		settingFrame.grid(column=0, row=0, sticky=(N, E, W, S))

class ChromeProfilesFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.columnconfigure(1, weight=1)
		self.columnconfigure(2, weight=1)
		# self.columnconfigure(3, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		self.rowconfigure(6, weight=1)
				
		titleLabel = TitleLabel(self, 'Chrome Profiles')
		closeButton = CloseButton(self)
		profiles = getProfiles()
		profileList = []
		for profile in profiles:
			profileList.append(ttk.Button(self, text=profile, command=lambda pro=profiles[profile]:self.chromeTester(pro)))
		# layout
		titleLabel.grid(column = 0, row = 0, sticky=(W, E, N, S), padx=15, pady=5, columnspan=4)
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		colnum = 0
		rownum = 1
		for profile in profileList:
			if colnum == 3:
				colnum = 0
				rownum += 1
			profile.grid(column = colnum, row = rownum, sticky=(W, E, N, S), padx=15, pady=5)
			colnum += 1

	def chromeTester(self, profile):
		chromeUserData = profile['chrome_user_data']
		chromeProfile = profile['chrome_profile']
		if platform == "linux" or platform == "linux2":
			CHROME = "google-chrome"
		elif platform == "win32":
			CHROME = "chrome.exe"
		if chromeUserData != '':
			Popen([CHROME, "https://google.com","--user-data-dir={}".format(chromeUserData), "--profile-directory={}".format(chromeProfile)])

class FileChooserFrame(ttk.Frame):
	def __init__(self, window, **kwargs):
		super().__init__(window)
		self.__filename = StringVar()
		# FOR EXCEL SHEET DISPLAY
		try: 
			sheetlist = kwargs['sheetlist']
		except:
			sheetlist = None

		fileLabel = ttk.Label(self, textvariable=self.__filename, foreground="red")
		label1 = ttk.Label(self, text=kwargs['label'])
		chooseButton = ttk.Button(self, text="...", command=lambda:self.chooseButtonClick(kwargs['btype'], filetypes=kwargs['filetypes'], sheetlist=sheetlist))
		self.rowconfigure(0, weight=1)
		self.columnconfigure(0, weight=1)
		self.columnconfigure(1, weight=1)
		self.columnconfigure(2, weight=1)
		# self.config(width=70, height=10)
		label1.grid(row=0, column=0, sticky=(W))
		fileLabel.grid(row=0, column=1, sticky=(W,E, N,S))
		chooseButton.grid(row=0, column=2, sticky=(E))
		
	@property
	def filename(self):
		return self.__filename.get()

	@filename.setter
	def filename(self, value):
		self.__filename.set(value)

	def chooseButtonClick(self, btype, **kwargs):
		if btype == 'folder':
			filenametmp = filedialog.askdirectory(title='Select Folder')
		else:
			filenametmp = filedialog.askopenfilename(title='Select File', filetypes=kwargs['filetypes'])

		if filenametmp != ():
			self.filename = filenametmp
			if kwargs['sheetlist'] != None:
				wb = openpyxl.load_workbook(filenametmp)
				if type(kwargs['sheetlist']) == tuple:
					for idx, sl in enumerate(kwargs['sheetlist']):
						kwargs['sheetlist'][idx]['values'] = wb.sheetnames
						kwargs['sheetlist'][idx].current(0)
				else:
					kwargs['sheetlist']['values'] = wb.sheetnames
					kwargs['sheetlist'].current(0)

class FileChooserMultipleFrame(ttk.Frame):
	def __init__(self, window, **kwargs):
		super().__init__(window)
		label1 = ttk.Label(self, text=kwargs['label'])
		chooseButton = ttk.Button(self, text="...", command=lambda:self.chooseButtonClick(filetypes=kwargs['filetypes']))
		delButton = ttk.Button(self, text="Del", command=lambda:self.delButtonClick())
		scrollbarx = Scrollbar(self, orient=HORIZONTAL)
		
		self.__filenames = tkinter.StringVar()
		
		self.__filelist = tkinter.Listbox(self,listvariable=self.__filenames, height=6, selectmode=tkinter.EXTENDED, xscrollcommand=scrollbarx.set, width=56)
		scrollbarx.config(command=self.__filelist.xview)
		# filelist.bind
		self.rowconfigure(0, weight=1)
		self.columnconfigure(0, weight=1)
		self.columnconfigure(1, weight=1)
		self.columnconfigure(2, weight=1)
		# self.config(width=70, height=10)
		label1.grid(row=0, column=0, sticky=(W))
		self.__filelist.grid(row=0, column=1, sticky=(W,E, N,S))
		scrollbarx.grid(row=1, column=1, sticky=(W,E, N,S))
		chooseButton.grid(row=0, column=2, sticky=(E))
		delButton.grid(row=2, column=1, sticky=(E))
	@property
	def filenames(self):
		return self.__filenames.get()

	def delButtonClick(self):
		self.__filelist.delete(ANCHOR)

	def chooseButtonClick(self, **kwargs):
		filenametmp = filedialog.askopenfilenames(title='Select File', filetypes=kwargs['filetypes'])
		if filenametmp != '':
			for fl in filenametmp:
				self.__filelist.insert(tkinter.END, fl)
		
class SettingFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		config = getConfig()
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')
		titleLabel = TitleLabel(self, text="Settings")
		chromeFolder = FileChooserFrame(self, btype="folder", label="Chrome User Data Folder:", filetypes=())
		profile = Entry(self, width=45)
		profile.insert(0, config['chrome_profile'])
		profile.setvar(config['chrome_profile'])
		profile.config(state="disabled")
		saveButton = ttk.Button(self, text='Save Setting', command=lambda:self.saveSetting(chrome_user_data=chromeFolder.filename, chrome_profile=profile.get()))
		labelprofile = Label(self, text="Profile Name:")

		chromeFolder.grid(column=0, row=1, sticky=(W,E), padx=20)
		labelprofile.grid(column = 0, row = 2, sticky=(W), padx=20)
		profile.grid(column = 0, row = 2, pady=10)
		
		chromeFolder.filename = config['chrome_user_data']
		chromeButton = ttk.Button(self, text='Chrome Tester', command = lambda: self.openChrome())
		pullButton = ttk.Button(self, text='Update Script', command=lambda:self.gitPull())

		#layout
		titleLabel.grid(column = 0, row = 0, sticky=(W, E, N, S), padx=15, pady=5)
		pullButton.grid(column = 0, row = 3, sticky = (W), padx=20, pady=10)
		chromeButton.grid(column = 0, row = 3, sticky = (E), padx=20, pady=10)
		saveButton.grid	(column = 0, row = 3, sticky = (N, S), padx=20, pady=10)
	
	
	def saveSetting(self, **kwarg):
		if platform == "win32":
			cud = kwarg['chrome_user_data'].replace("/", "\\")
		else:
			cud = kwarg['chrome_user_data']

		newfile = open("setting.json", "w")
		dict = {
			"chrome_user_data": cud,
			"chrome_profile": kwarg['chrome_profile'],
		}
		json.dump(dict, newfile)
		newfile.close()
		messagebox.showinfo(title='Info', message='File Setting Saved')

	def openChrome(self):
		chromeUserData = getConfig()['chrome_user_data']
		chromeProfile = getConfig()['chrome_profile']
		if platform == "linux" or platform == "linux2":
			CHROME = "google-chrome"
		elif platform == "win32":
			CHROME = "chrome.exe"
		if chromeUserData != '':
			Popen([CHROME, "https://amazon.com","--user-data-dir={}".format(chromeUserData), "--profile-directory={}".format(chromeProfile)])

	def gitPull(self):
		git_dir = os.getcwd() 
		g = git.cmd.Git(git_dir)
		g.pull()		
		messagebox.showinfo(title='Info', message='the scripts has updated..')

class MainFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		# self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		framestyle = ttk.Style()
		framestyle.configure('TFrame', background='#C1C1CD')
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove', style='TFrame')
		
		# self.place(anchor=CENTER)
		self.columnconfigure(0, weight=1)
		self.columnconfigure(1, weight=1)
		self.columnconfigure(2, weight=1)
		self.columnconfigure(3, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		self.rowconfigure(6, weight=1)
		self.rowconfigure(7, weight=1)
		self.rowconfigure(8, weight=1)
		self.rowconfigure(9, weight=1)
		
		titleLabel = TitleLabel(self, 'Main Menu')
		# pdfconvertButton = PdfConvertButton(self, window)
		pdfconvertButton = FrameButton(self, window, text="PDF Converter", class_frame=PdfConvertFrame)
		scrapeBySellerButton = FrameButton(self, window, text="Scrape By Amazon Seller", class_frame=ScrapeBySellerAmazonFrame)
		scrapeWalmartButton = FrameButton(self, window, text="Scrape Walmart", class_frame=ScrapeWalmartFrame)
		trackingUpdateButton = FrameButton(self, window, text="Amazon Tracking Update", class_frame=TrackingUpdateFrame)
		statisticsButton = FrameButton(self, window, text="Statistics", class_frame=StatisticsFrame)
		dykShippedButton = FrameButton(self, window, text="DYK Shipped", class_frame=DykShippedFrame)
		costLookupButton = FrameButton(self, window, text="Cost Lookup", class_frame=CostLookupFrame)
		fdaEntryButton = FrameButton(self, window, text="FDA Entry", class_frame=FdaEntryFrame)
		fdaPdfButton = FrameButton(self, window, text="FDA PDF", class_frame=FdaPdfFrame)
		amazonShipmentButton = FrameButton(self, window, text="Amazon Shipment", class_frame=AmazonShippingFrame)
		amazonShipmentCheckButton = FrameButton(self, window, text="Amazon Shipment Check", class_frame=AmazonShippingCheckFrame)
		fdaEntryPdfButton = FrameButton(self, window, text="FDA Entry + PDF", class_frame=FdaEntryPdfFrame)
		amazonLabelXlsButton = FrameButton(self, window, text="Amazon Join PDF", class_frame=AmazonJoinPdfFrame)
		amazonReviewButton = FrameButton(self, window, text="Amazon Review Request", class_frame=AmazonReviewFrame)
		canadaPostPdfButton = FrameButton(self, window, text="Canada Post PDF Convert", class_frame=CanadaPostPdfFrame)
		# amazonAllButton = FrameButton(self, window, text="Amazon Shipment + FDA", class_frame=AmazonAllFrame)
		canadaPostButton = FrameButton(self, window, text="Canada Post Tracker", class_frame=CanadaPostFrame)
		walmartstButton = FrameButton(self, window, text="Walmart Scraper", class_frame=WalmartstFrame)
		superstoreButton = FrameButton(self, window, text="Superstore Scraper", class_frame=SuperstoreFrame)


		# layout
		titleLabel.grid(column = 0, row = 0, sticky=(W, E, N, S), padx=15, pady=5, columnspan=3)
		pdfconvertButton.grid(column = 0, row = 1, sticky=(W, E, N, S), padx=15, pady=5)
		scrapeBySellerButton.grid(column = 1, row = 1, sticky=(W, E, N, S), padx=15, pady=5)
		trackingUpdateButton.grid(column = 2, row = 1, sticky=(W, E, N, S), padx=15, pady=5)
		scrapeWalmartButton.grid(column = 0, row = 2, sticky=(W, E, N, S), padx=15, pady=5)
		statisticsButton.grid(column = 1, row = 2, sticky=(W, E, N, S), padx=15, pady=5)
		dykShippedButton.grid(column = 2, row = 2, sticky=(W, E, N, S), padx=15, pady=5)
		costLookupButton.grid(column = 0, row = 3, sticky=(W, E, N, S), padx=15, pady=5)
		fdaEntryButton.grid(column = 1, row = 3, sticky=(W, E, N, S), padx=15, pady=5)
		fdaPdfButton.grid(column = 2, row = 3, sticky=(W, E, N, S), padx=15, pady=5)
		amazonShipmentButton.grid(column = 0, row = 4, sticky=(W, E, N, S), padx=15, pady=5)
		fdaEntryPdfButton.grid(column = 1, row = 4, sticky=(W, E, N, S), padx=15, pady=5)
		amazonShipmentCheckButton.grid(column = 2, row = 4, sticky=(W, E, N, S), padx=15, pady=5)
		amazonLabelXlsButton.grid(column = 0, row = 5, sticky=(W, E, N, S), padx=15, pady=5)
		amazonReviewButton.grid(column = 1, row = 5, sticky=(W, E, N, S), padx=15, pady=5)
		canadaPostPdfButton.grid(column = 2, row = 5, sticky=(W, E, N, S), padx=15, pady=5)
		# amazonAllButton.grid(column = 0, row = 6, sticky=(W, E, N, S), padx=15, pady=5)
		canadaPostButton.grid(column = 0, row = 6, sticky=(W, E, N, S), padx=15, pady=5)
		# amazonShipmentButton['state'] = DISABLED
		# fdaEntryPdfButton['state'] = DISABLED
		walmartstButton.grid(column = 1, row = 6, sticky=(W, E, N, S), padx=15, pady=5)
		superstoreButton.grid(column = 2, row = 6, sticky=(W, E, N, S), padx=15, pady=5)

class PdfConvertFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		
		# populate
		titleLabel = TitleLabel(self, text="Pdf Converter")
		closeButton = CloseButton(self)
		
		pdfInputFile = FileChooserFrame(self, btype="file", label="Select PDF Input File:", filetypes=(("pdf files", "*.pdf"),("all files", "*.*")))
		xlsOutputFile = FileChooserFrame(self, btype="file", label="Select XLSX Output File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=pdfInputFile.filename, output=xlsOutputFile.filename))

		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		pdfInputFile.grid(column = 0, row = 1, sticky = (W,E))
		xlsOutputFile.grid(column = 0, row = 2, sticky = (W,E))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		# self.runButton.state(['disabled'])

	def run_process(self, **kwargs):
		if kwargs['input'] == "" or kwargs['output'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "pdfconvert.py", "-i", kwargs['input'], "-o", kwargs['output'])	
			# Popen([PYLOC, "pdfconvert.py", "-i", kwargs['input'], "-o", kwargs['output']], creationflags=CREATE_NEW_CONSOLE)
			run_module(comlist=[PYLOC, "modules/pdfconvert.py", "-i", kwargs['input'], "-o", kwargs['output']])

class ScrapeBySellerAmazonFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		# populate
		titleLabel = TitleLabel(self, text="Scrape By Amazon Seller")
		closeButton = CloseButton(self)
	
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select XLSX Input File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename))
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W,E))

	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "scrapebyseller.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data'])	
			# Popen([PYLOC, "scrapebyseller.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']], creationflags=CREATE_NEW_CONSOLE)
			run_module(comlist=[PYLOC, "modules/scrapebyseller.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']])

class TrackingUpdateFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		# populate
		titleLabel = TitleLabel(self, text="Amazon Tracking Update")
		closeButton = CloseButton(self)
		
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select XLSM or XLSX Input File:", filetypes=(("excel files", "*.xlsm *.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W, E))
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))

	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "trackupdate.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data'])	
			# Popen([PYLOC, "trackupdate.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']], creationflags=CREATE_NEW_CONSOLE)
			run_module(comlist=[PYLOC, "modules/trackupdate.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']])

class StatisticsFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		labelclist = Label(self, text="Amazon Store:")
		countries = ["US", "CA"]
		vl = StringVar()
		titleLabel = TitleLabel(self, text="Statistics")
		closeButton = CloseButton(self)
		clist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		# for country in countries:
		clist['values'] = [country for country in countries]
		clist.current(0)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select XLSX Input File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename, country=clist.get() ))
		

		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W,E))
		labelclist.grid(column = 0, row = 2, sticky=(W))
		clist.grid(column = 0, row = 2, pady=10)
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))

	def run_process(self, **kwargs):
		print(kwargs)
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			run_module(comlist=[PYLOC, "modules/statistic.py", "-i", kwargs['input'], "-c", kwargs['country'], "-d", getConfig()['chrome_user_data']])

class CanadaPostFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		titleLabel = TitleLabel(self, text="Canada Post Tracker")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select CSV Input File:", filetypes=(("CSV files", "*.csv"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename))
		
			# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W, E))
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))

	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			run_module(comlist=[PYLOC, "modules/cposttracker.py", "-i", kwargs['input']])

class DykShippedFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		# populate
		titleLabel = TitleLabel(self, text="DYK Shipped")
		closeButton = CloseButton(self)
		xlsOutputFile = FileChooserFrame(self, btype="file", label="Select XLSX Output File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(output=xlsOutputFile.filename))
		# layout

		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsOutputFile.grid(column = 0, row = 1, sticky = (W, E))
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))

	def run_process(self, **kwargs):
		if kwargs['output'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "dykshipped.py", "-o", kwargs['output'], "-d", getConfig()['chrome_user_data'])	
			# Popen([PYLOC, "dykshipped.py", "-o", kwargs['output'], "-d", getConfig()['chrome_user_data']], creationflags=CREATE_NEW_CONSOLE)
			run_module(comlist=[PYLOC, "modules/dykshipped.py", "-o", kwargs['output'], "-d", getConfig()['chrome_user_data']])

class CostLookupFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		# populate
		titleLabel = TitleLabel(self, text="Cost Lookup")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select XLSX Input File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename))

		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W, E))
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))

	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "costlookup.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data'])	
			# Popen([PYLOC, "costlookup.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']], creationflags=CREATE_NEW_CONSOLE)
			run_module(comlist=[PYLOC, "modules/costlookup.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']])

class ScrapeWalmartFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		# populate
		titleLabel = TitleLabel(self, text="Scrape Walmart")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select XLSX Input File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename))

		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W,E))
		runButton.grid(column = 0, row = 3, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))


	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "walmart.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data'])	
			# Popen([PYLOC, "walmart.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']], creationflags=CREATE_NEW_CONSOLE)
			run_module(comlist=[PYLOC, "modules/walmart.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data']])

class FdaEntryFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)

		# populate
		titleLabel = TitleLabel(self, text="FDA Entry")
		closeButton = CloseButton(self)
	
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select XLSX Input File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		labeldate = Label(self, text="Anticipated Date Arrival:")
		labelsname = Label(self, text="Sheet Name:")
		sheetName = Entry(self, width=45)
		dateArrival = DateEntry(self, width= 20, date_pattern='mm/dd/yyyy')
		
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename, sheetname = sheetName, datearrival=dateArrival, pdffolder=pdfFolder.filename))
		pdfFolder = FileChooserFrame(self, btype="folder", label="PDF output Folder:", filetypes=())

		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W,E))
		labelsname.grid(column = 0, row = 2, sticky=(W))
		sheetName.grid(column = 0, row = 2, pady=10)
		pdfFolder.grid(column = 0, row = 3, sticky = (W,E))
		dateArrival.grid(column=0, row = 4)
		labeldate.grid(column = 0, row = 4, sticky=(W))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))

	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		elif kwargs['pdffolder'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the pdf folder')
		elif kwargs['sheetname'].get() == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have filled the sheetname')
		else:
			pdffolder = kwargs['pdffolder']
			if platform == "win32":
				pdffolder = pdffolder.replace("/", "\\")
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed it.')
			run_module(comlist=[PYLOC, "modules/fda.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data'], "-s", kwargs['sheetname'].get(), "-dt", str(kwargs['datearrival'].get_date()), "-o", pdffolder])

class FdaPdfFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		
		# populate
		titleLabel = TitleLabel(self, text="FDA PDF Extractor")
		closeButton = CloseButton(self)

		pdfInputFiles = FileChooserMultipleFrame(self, label="Select Input PDF File:", filetypes=(("pdf files", "*.pdf"),("all files", "*.*")))

		# xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input XLSX File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")))
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=sheetlist)

		outputfolder = FileChooserFrame(self, btype="folder", label="Output PDF Folder:", filetypes=())

		labelsname = Label(self, text="Sheet Name:")
		
		# sheetName = Entry(self, width=45)
		# runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(pdfinput=pdfInputFiles.filenames, xlsinput=xlsInputFile.filename, sname=sheetName, pdfoutput=outputfolder.filename))
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(pdfinput=pdfInputFiles.filenames, xlsinput=xlsInputFile.filename, sname=sheetlist, pdfoutput=outputfolder.filename))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		pdfInputFiles.grid(column = 0, row = 1, sticky = (W,E))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname.grid(column = 0, row = 3, sticky=(W))
		# sheetName.grid(column = 0, row = 3, pady=10)
		outputfolder.grid(column = 0, row = 4, sticky = (W,E))
		sheetlist.grid(column=0, row = 3, pady=10)

		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		# self.runButton.state(['disabled'])

	def run_process(self, **kwargs):
		if kwargs['pdfinput'] == "" or kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# print(PYLOC, "fdapdf.py", "-i", kwargs['input'])
			# Popen([PYLOC, "fdapdf.py", "-i", kwargs['input']], creationflags=CREATE_NEW_CONSOLE)
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/fdapdf.py", "-pdf", kwargs['pdfinput'], "-xls", kwargs['xlsinput'], "-sname", kwargs['sname'].get(), "-output", kwargs['pdfoutput'] ])

class AmazonShippingFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		
		# populate
		titleLabel = TitleLabel(self, text="Amazon Shipment")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=sheetlist)

		outputfolder = FileChooserFrame(self, btype="folder", label="Output PDF Folder:", filetypes=())

		labelsname = Label(self, text="Sheet Name:")
		
		# sheetName = Entry(self, width=45)
		
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(xlsinput=xlsInputFile.filename, sname=sheetlist, pdfoutput=outputfolder.filename))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname.grid(column = 0, row = 3, sticky=(W))
		# sheetName.grid(column = 0, row = 3, pady=10)
		outputfolder.grid(column = 0, row = 4, sticky = (W,E))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		sheetlist.grid(column=0, row = 3, pady=10)

		# self.runButton.state(['disabled'])

	def run_process(self, **kwargs):
		if kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			pdffolder = kwargs['pdfoutput']
			if platform == "win32":
				pdffolder = pdffolder.replace("/", "\\")

			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/amazonship.py", "-xls", kwargs['xlsinput'], "-sname", kwargs['sname'].get(), "-output", pdffolder, "-cdata",  getConfig()['chrome_user_data']])

class FdaEntryPdfFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S), columnspan=2)
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		# populate
		titleLabel = TitleLabel(self, text="FDA Entry + PDF Generator")
		closeButton = CloseButton(self)
	
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=sheetlist)

		labeldate = Label(self, text="Anticipated Date Arrival:")
		labelsname = Label(self, text="Sheet Name:")
		# sheetName = Entry(self, width=45)
		dateArrival = DateEntry(self, width= 20, date_pattern='mm/dd/yyyy')
		
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(input=xlsInputFile.filename, sheetname = sheetlist, datearrival=dateArrival, pdffolder=pdfFolder.filename))
		pdfFolder = FileChooserFrame(self, btype="folder", label="PDF output Folder:", filetypes=())

		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 1, sticky = (W,E))
		labelsname.grid(column = 0, row = 2, sticky=(W))
		# sheetName.grid(column = 0, row = 2, pady=10)
		pdfFolder.grid(column = 0, row = 3, sticky = (W,E))
		dateArrival.grid(column=0, row = 4)
		labeldate.grid(column = 0, row = 4, sticky=(W))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		sheetlist.grid(column=0, row = 2, pady=10)
		
	def run_process(self, **kwargs):
		if kwargs['input'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		elif kwargs['pdffolder'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the pdf folder')
		elif kwargs['sheetname'].get() == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have filled the sheetname')
		else:
			pdffolder = kwargs['pdffolder']
			if platform == "win32":
				pdffolder = pdffolder.replace("/", "\\")
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed it.')
			run_module(comlist=[PYLOC, "modules/autofdapdf.py", "-i", kwargs['input'], "-d", getConfig()['chrome_user_data'], "-s", kwargs['sheetname'].get(), "-dt", str(kwargs['datearrival'].get_date()), "-o", pdffolder])

class AmazonShippingCheckFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		
		# populate
		titleLabel = TitleLabel(self, text="Amazon Shipment Check")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=sheetlist)

		labelsname = Label(self, text="Sheet Name:")
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(xlsinput=xlsInputFile.filename, sname=sheetlist))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname.grid(column = 0, row = 3, sticky=(W))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		sheetlist.grid(column=0, row = 3, pady=10)
	def run_process(self, **kwargs):
		if kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/amazonshipcheck.py", "-xls", kwargs['xlsinput'], "-sname", kwargs['sname'].get()])

class AmazonJoinPdfFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		
		# populate
		titleLabel = TitleLabel(self, text="Amazon Shipping Join PDF")
		closeButton = CloseButton(self)
		sourcefolder = FileChooserFrame(self, btype="folder", label="PDF Source Folder:", filetypes=())
		outputfolder = FileChooserFrame(self, btype="folder", label="PDF Output Folder:", filetypes=())

		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(sourcefolder=sourcefolder.filename, outputfolder=outputfolder.filename))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		sourcefolder.grid(column = 0, row = 2, sticky = (W,E))
		outputfolder.grid(column = 0, row = 3, sticky = (W,E))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))


	def run_process(self, **kwargs):
		if kwargs['sourcefolder'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the folder')
		elif kwargs['outputfolder'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the folder')

		else:
			sourcefolder = kwargs['sourcefolder']
			outputfolder = kwargs['outputfolder']
			if platform == "win32":
				sourcefolder = sourcefolder.replace("/", "\\")
				outputfolder = outputfolder.replace("/", "\\")

			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/amazonshipjoin.py", "-sourcefolder", sourcefolder, "-outputfolder", outputfolder])

class AmazonReviewFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		
		# populate
		titleLabel = TitleLabel(self, text="Amazon Review Request")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input XLSX File:", filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*")), sheetlist=sheetlist)

		labelsname = Label(self, text="Sheet Name:")
		
		# sheetName = Entry(self, width=45)
		
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(xlsinput=xlsInputFile.filename, sname=sheetlist))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname.grid(column = 0, row = 3, sticky=(W))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		sheetlist.grid(column=0, row = 3, pady=10)

		# self.runButton.state(['disabled'])

	def run_process(self, **kwargs):
		if kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/amazonreview.py", "-xls", kwargs['xlsinput'], "-sname", kwargs['sname'].get()])

class CanadaPostPdfFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		
		# populate
		titleLabel = TitleLabel(self, text="Canada Post PDF Converter")
		closeButton = CloseButton(self)
		pdfInputFiles = FileChooserMultipleFrame(self, label="Select Input PDF File:", filetypes=(("pdf files", "*.pdf"),("all files", "*.*")))
		outputfolder = FileChooserFrame(self, btype="folder", label="Output CSV Folder:", filetypes=())

		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(pdfinput=pdfInputFiles.filenames, pdfoutput=outputfolder.filename))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		pdfInputFiles.grid(column = 0, row = 1, sticky = (W,E))
		outputfolder.grid(column = 0, row = 4, sticky = (W,E))

		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		# self.runButton.state(['disabled'])

	def run_process(self, **kwargs):
		if kwargs['pdfinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			# messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/cpostconvert.py", "-pdf", kwargs['pdfinput'], "-output", kwargs['pdfoutput'] ])

class AmazonAllFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist1 = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		sheetlist2 = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		sheetlist3 = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		dateArrival = DateEntry(self, width= 20, date_pattern='mm/dd/yyyy')
		
		# populate
		titleLabel = TitleLabel(self, text="Amazon Shipment + FDA")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=(sheetlist1, sheetlist2, sheetlist3))

		outputfolder = FileChooserFrame(self, btype="folder", label="Output PDF Folder:", filetypes=())

		labelsname1 = Label(self, text="Shipment Sheet:")
		labelsname2 = Label(self, text="Prior Notice Sheet:")
		labelsname3 = Label(self, text="Tracking Update Sheet:")
		labeldate = Label(self, text="Anticipated Date Arrival:")
		
		# sheetName = Entry(self, width=45)
		
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(xlsinput=xlsInputFile.filename, shipsheet=sheetlist1, pnsheet=sheetlist2, tracksheet=sheetlist3, pdfoutput=outputfolder.filename, datearrival=dateArrival))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname1.grid(column = 0, row = 3, sticky=(W))
		labelsname2.grid(column = 0, row = 4, sticky=(W))
		labelsname3.grid(column = 0, row = 5, sticky=(W))
		labeldate.grid(column = 0, row = 7, sticky=(W))

		sheetlist1.grid(column=0, row = 3, pady=10)
		sheetlist2.grid(column=0, row = 4, pady=10)
		sheetlist3.grid(column=0, row = 5, pady=10)
		dateArrival.grid(column=0, row = 7)

		outputfolder.grid(column = 0, row = 6, sticky = (W,E))
		runButton.grid(column = 0, row = 8, sticky = (E))
		closeButton.grid(column = 0, row = 9, sticky = (E, N, S))


		# self.runButton.state(['disabled'])

	def run_process(self, **kwargs):
		if kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			pdffolder = kwargs['pdfoutput']
			if platform == "win32":
				pdffolder = pdffolder.replace("/", "\\")

			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/amazonall.py", "-xls", kwargs['xlsinput'], "-shipsheet", kwargs['shipsheet'].get(), "-pnsheet", kwargs['pnsheet'].get(), "-tracksheet", kwargs['tracksheet'].get(), "-output", pdffolder, "-cdata",  getConfig()['chrome_user_data'], "-dt", str(kwargs['datearrival'].get_date())])

class WalmartstFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		
		# populate
		titleLabel = TitleLabel(self, text="Walmart Price Monitor")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=sheetlist)

		labelsname = Label(self, text="Sheet Name:")
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(xlsinput=xlsInputFile.filename, sname=sheetlist))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname.grid(column = 0, row = 3, sticky=(W))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		sheetlist.grid(column=0, row = 3, pady=10)
	def run_process(self, **kwargs):
		if kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/walmart_superstore.py", "-xls", kwargs['xlsinput'], "-sname", kwargs['sname'].get(), "-module", "walmart"])

class SuperstoreFrame(ttk.Frame):
	def __init__(self, window) -> None:
		super().__init__(window)
		# configure
		self.grid(column=0, row=0, sticky=(N, E, W, S))
		self.config(padding="20 20 20 20", borderwidth=1, relief='groove')

		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.rowconfigure(3, weight=1)
		self.rowconfigure(4, weight=1)
		self.rowconfigure(5, weight=1)
		sheetlist = ttk.Combobox(self, textvariable=StringVar(), state="readonly")
		
		# populate
		titleLabel = TitleLabel(self, text="Superstore Price Monitor")
		closeButton = CloseButton(self)
		xlsInputFile = FileChooserFrame(self, btype="file", label="Select Input Excel File:", filetypes=(("Excel files", "*.xlsx *.xlsm"),("all files", "*.*")), sheetlist=sheetlist)

		labelsname = Label(self, text="Sheet Name:")
		runButton = ttk.Button(self, text='Run Process', command = lambda:self.run_process(xlsinput=xlsInputFile.filename, sname=sheetlist))
		
		# layout
		titleLabel.grid(column = 0, row = 0, sticky = (W, E, N, S))
		xlsInputFile.grid(column = 0, row = 2, sticky = (W,E))
		labelsname.grid(column = 0, row = 3, sticky=(W))
		runButton.grid(column = 0, row = 5, sticky = (E))
		closeButton.grid(column = 0, row = 6, sticky = (E, N, S))
		sheetlist.grid(column=0, row = 3, pady=10)
	def run_process(self, **kwargs):
		if kwargs['xlsinput'] == "": 
			messagebox.showwarning(title='Warning', message='Please make sure you have choosed the files')
		else:
			messagebox.showwarning(title='Warning', message='This process will update the excel file. make sure you have closed the file.')
			run_module(comlist=[PYLOC, "modules/walmart_superstore.py", "-xls", kwargs['xlsinput'], "-sname", kwargs['sname'].get(), "-module", "superstore"])

class CloseButton(ttk.Button):
	def __init__(self, parent):
		super().__init__(parent)
		self.config(text = '< Back', command=lambda : parent.destroy())
		
class FrameButton(ttk.Button):
	def __init__(self, parent, window, **kwargs):
		super().__init__(parent)
		# object attributes
		self.text = kwargs['text']
		# configure
		self.config(text = self.text, command = lambda : kwargs['class_frame'](window))

class TitleLabel(ttk.Label):
	def __init__(self, parent, text):
		super().__init__(parent)
		font_tuple = ("Comic Sans MS", 20, "bold")
		self.config(text=text, font=font_tuple, anchor="center")

if __name__ == "__main__":
	if platform == "linux" or platform == "linux2":
		PYLOC = "python"
	elif platform == "win32":
		PYLOC = "python.exe"
	
	isExist = os.path.exists("setting.json")
	if not isExist:
		dir = os.getcwd()
		newfile = open("setting.json", "w")
		if platform == "linux" or platform == "linux2":
			dict = {
				"chrome_user_data": "{}/user-data".format(dir),
				"chrome_profile": "Default"
			}
		elif platform == "win32":
			dict = {

				"chrome_user_data": "{}\\AppData\\Local\\Google\Chrome\\User Data".format(os.getenv('USERPROFILE')),
				"chrome_profile": "Default"

			}

		json.dump(dict, newfile)
		newfile.close()
	main()